from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime
import base64
import json
import re
import io
import pypdf

# For Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Admin password for import (from .env)
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'Liana2018')
OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY', '')

# Gmail SMTP Configuration
SMTP_EMAIL = os.environ.get('SMTP_EMAIL', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))

# ============ Models ============

class FinancingRates(BaseModel):
    """Taux de financement pour chaque terme (36-96 mois)"""
    rate_36: float
    rate_48: float
    rate_60: float
    rate_72: float
    rate_84: float
    rate_96: float

class VehicleProgram(BaseModel):
    """
    Structure d'un programme de financement véhicule
    
    RÈGLES IMPORTANTES:
    - Option 1 (Consumer Cash): Rabais AVANT TAXES + taux d'intérêt (variables selon le véhicule)
    - Option 2 (Alternative): Rabais = $0 + taux réduits (peut être None si non disponible)
    - Bonus Cash: Rabais APRÈS TAXES (comme comptant), combinable avec Option 1 OU 2
    - Les taux sont FIXES (non variables)
    """
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    brand: str
    model: str
    trim: Optional[str] = None
    year: int
    
    # Option 1: Consumer Cash (rabais AVANT taxes) + taux Consumer Cash
    consumer_cash: float = 0  # Rabais avant taxes
    option1_rates: FinancingRates
    
    # Option 2: Alternative Consumer Cash (généralement $0) + taux réduits
    # None = Option 2 non disponible pour ce véhicule
    option2_rates: Optional[FinancingRates] = None
    
    # Bonus Cash: Rabais APRÈS taxes (combinable avec Option 1 ou 2)
    bonus_cash: float = 0
    
    # Métadonnées du programme
    program_month: int = Field(default_factory=lambda: datetime.utcnow().month)
    program_year: int = Field(default_factory=lambda: datetime.utcnow().year)
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class VehicleProgramCreate(BaseModel):
    brand: str
    model: str
    trim: Optional[str] = None
    year: int
    consumer_cash: float = 0
    option1_rates: FinancingRates
    option2_rates: Optional[FinancingRates] = None
    bonus_cash: float = 0
    program_month: Optional[int] = None
    program_year: Optional[int] = None

class VehicleProgramUpdate(BaseModel):
    brand: Optional[str] = None
    model: Optional[str] = None
    trim: Optional[str] = None
    year: Optional[int] = None
    consumer_cash: Optional[float] = None
    option1_rates: Optional[FinancingRates] = None
    option2_rates: Optional[FinancingRates] = None
    bonus_cash: Optional[float] = None

class CalculationRequest(BaseModel):
    vehicle_price: float
    program_id: Optional[str] = None

class PaymentComparison(BaseModel):
    term_months: int
    # Option 1: Consumer Cash + taux
    option1_rate: float
    option1_monthly: float
    option1_total: float
    option1_rebate: float  # Consumer Cash (avant taxes)
    # Option 2: Taux réduits (si disponible)
    option2_rate: Optional[float] = None
    option2_monthly: Optional[float] = None
    option2_total: Optional[float] = None
    # Meilleure option
    best_option: Optional[str] = None
    savings: Optional[float] = None

class CalculationResponse(BaseModel):
    vehicle_price: float
    consumer_cash: float
    bonus_cash: float
    brand: str
    model: str
    trim: Optional[str]
    year: int
    comparisons: List[PaymentComparison]

class ProgramPeriod(BaseModel):
    month: int
    year: int
    count: int

class ImportRequest(BaseModel):
    password: str
    programs: List[Dict[str, Any]]
    program_month: int
    program_year: int

# ============ CRM Models ============

class Submission(BaseModel):
    """Soumission client avec suivi"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    # Client info
    client_name: str
    client_phone: str
    client_email: str
    # Vehicle info
    vehicle_brand: str
    vehicle_model: str
    vehicle_year: int
    vehicle_price: float
    # Financing info
    term: int
    payment_monthly: float
    payment_biweekly: float = 0
    payment_weekly: float = 0
    selected_option: str = "1"
    rate: float = 0
    # Tracking
    submission_date: datetime = Field(default_factory=datetime.utcnow)
    reminder_date: Optional[datetime] = None  # Default: 24h after submission
    reminder_done: bool = False
    status: str = "pending"  # pending, contacted, converted, lost
    notes: str = ""
    # For comparison
    program_month: int = 0
    program_year: int = 0

class SubmissionCreate(BaseModel):
    client_name: str
    client_phone: str
    client_email: str
    vehicle_brand: str
    vehicle_model: str
    vehicle_year: int
    vehicle_price: float
    term: int
    payment_monthly: float
    payment_biweekly: float = 0
    payment_weekly: float = 0
    selected_option: str = "1"
    rate: float = 0
    program_month: int = 0
    program_year: int = 0

class ReminderUpdate(BaseModel):
    reminder_date: datetime
    notes: Optional[str] = None

class BetterOffer(BaseModel):
    """Meilleure offre trouvée pour un client"""
    submission_id: str
    client_name: str
    client_phone: str
    client_email: str
    vehicle: str
    old_payment: float
    new_payment: float
    savings_monthly: float
    savings_total: float
    term: int
    approved: bool = False
    email_sent: bool = False

# ============ Utility Functions ============

def calculate_monthly_payment(principal: float, annual_rate: float, months: int) -> float:
    """Calcule le paiement mensuel avec la formule d'amortissement"""
    if principal <= 0 or months <= 0:
        return 0
    if annual_rate == 0:
        return round(principal / months, 2)
    
    monthly_rate = annual_rate / 100 / 12
    payment = principal * (monthly_rate * (1 + monthly_rate) ** months) / ((1 + monthly_rate) ** months - 1)
    return round(payment, 2)

def get_rate_for_term(rates: FinancingRates, term: int) -> float:
    """Obtient le taux pour un terme spécifique"""
    rate_map = {
        36: rates.rate_36,
        48: rates.rate_48,
        60: rates.rate_60,
        72: rates.rate_72,
        84: rates.rate_84,
        96: rates.rate_96
    }
    return rate_map.get(term, 4.99)

# ============ Routes ============

@api_router.get("/")
async def root():
    return {"message": "Vehicle Financing Calculator API - v4 (avec historique et Bonus Cash)"}

# Ping endpoint for keep-alive
@api_router.get("/ping")
async def ping():
    """Endpoint pour garder le serveur actif (keep-alive)"""
    return {"status": "ok", "message": "Server is alive"}

# Get PDF info (page count)
@api_router.post("/pdf-info")
async def get_pdf_info(file: UploadFile = File(...)):
    """Récupère les informations du PDF (nombre de pages)"""
    try:
        contents = await file.read()
        pdf_reader = pypdf.PdfReader(io.BytesIO(contents))
        total_pages = len(pdf_reader.pages)
        
        return {
            "success": True,
            "total_pages": total_pages,
            "filename": file.filename
        }
    except Exception as e:
        logger.error(f"Error reading PDF: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur lors de la lecture du PDF: {str(e)}"
        }

# Get available program periods (for history)
@api_router.get("/periods", response_model=List[ProgramPeriod])
async def get_periods():
    """Récupère les périodes de programmes disponibles (pour l'historique)"""
    pipeline = [
        {"$group": {
            "_id": {"month": "$program_month", "year": "$program_year"},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id.year": -1, "_id.month": -1}}
    ]
    periods = await db.programs.aggregate(pipeline).to_list(100)
    return [ProgramPeriod(month=p["_id"]["month"], year=p["_id"]["year"], count=p["count"]) for p in periods]

# Programs CRUD
@api_router.post("/programs", response_model=VehicleProgram)
async def create_program(program: VehicleProgramCreate):
    program_dict = program.dict()
    if program_dict.get("program_month") is None:
        program_dict["program_month"] = datetime.utcnow().month
    if program_dict.get("program_year") is None:
        program_dict["program_year"] = datetime.utcnow().year
    program_obj = VehicleProgram(**program_dict)
    await db.programs.insert_one(program_obj.dict())
    return program_obj

@api_router.get("/programs", response_model=List[VehicleProgram])
async def get_programs(month: Optional[int] = None, year: Optional[int] = None):
    """
    Récupère les programmes de financement
    Si month/year sont fournis, filtre par période
    Sinon, retourne la période la plus récente
    """
    if month and year:
        query = {"program_month": month, "program_year": year}
    else:
        # Trouver la période la plus récente
        latest = await db.programs.find_one(sort=[("program_year", -1), ("program_month", -1)])
        if latest:
            query = {"program_month": latest.get("program_month"), "program_year": latest.get("program_year")}
        else:
            query = {}
    
    programs = await db.programs.find(query).to_list(1000)
    return [VehicleProgram(**p) for p in programs]

@api_router.get("/programs/{program_id}", response_model=VehicleProgram)
async def get_program(program_id: str):
    program = await db.programs.find_one({"id": program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    return VehicleProgram(**program)

@api_router.put("/programs/{program_id}", response_model=VehicleProgram)
async def update_program(program_id: str, update: VehicleProgramUpdate):
    program = await db.programs.find_one({"id": program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    update_data = {k: v for k, v in update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    await db.programs.update_one({"id": program_id}, {"$set": update_data})
    updated = await db.programs.find_one({"id": program_id})
    return VehicleProgram(**updated)

@api_router.delete("/programs/{program_id}")
async def delete_program(program_id: str):
    result = await db.programs.delete_one({"id": program_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    return {"message": "Program deleted successfully"}

# Import programs with password protection
@api_router.post("/import")
async def import_programs(request: ImportRequest):
    """
    Importe des programmes de financement (protégé par mot de passe)
    Remplace tous les programmes pour le mois/année spécifié
    """
    if request.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    
    # Supprimer les programmes existants pour cette période
    await db.programs.delete_many({
        "program_month": request.program_month,
        "program_year": request.program_year
    })
    
    # Insérer les nouveaux programmes
    inserted = 0
    for prog_data in request.programs:
        # Assurer que les taux sont au bon format
        if prog_data.get("option1_rates") and isinstance(prog_data["option1_rates"], dict):
            prog_data["option1_rates"] = FinancingRates(**prog_data["option1_rates"]).dict()
        if prog_data.get("option2_rates") and isinstance(prog_data["option2_rates"], dict):
            prog_data["option2_rates"] = FinancingRates(**prog_data["option2_rates"]).dict()
        
        prog_data["program_month"] = request.program_month
        prog_data["program_year"] = request.program_year
        prog_data["bonus_cash"] = prog_data.get("bonus_cash", 0)
        
        prog = VehicleProgram(**prog_data)
        await db.programs.insert_one(prog.dict())
        inserted += 1
    
    return {"message": f"Importé {inserted} programmes pour {request.program_month}/{request.program_year}"}

# Calculate financing options
@api_router.post("/calculate", response_model=CalculationResponse)
async def calculate_financing(request: CalculationRequest):
    """
    Calcule et compare les options de financement
    
    Option 1: Prix - Consumer Cash (rabais avant taxes), avec taux Option 1
    Option 2: Prix complet, avec taux réduits (si disponible)
    
    Note: Bonus Cash est affiché mais non inclus dans le calcul du financement
    (car appliqué après taxes, comme comptant)
    """
    vehicle_price = request.vehicle_price
    
    if not request.program_id:
        raise HTTPException(status_code=400, detail="Program ID is required")
    
    program = await db.programs.find_one({"id": request.program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    program_obj = VehicleProgram(**program)
    consumer_cash = program_obj.consumer_cash
    bonus_cash = program_obj.bonus_cash
    
    comparisons = []
    terms = [36, 48, 60, 72, 84, 96]
    
    for term in terms:
        # Option 1: Avec Consumer Cash (rabais avant taxes) + taux Option 1
        option1_rate = get_rate_for_term(program_obj.option1_rates, term)
        principal1 = vehicle_price - consumer_cash  # Rabais avant taxes
        monthly1 = calculate_monthly_payment(principal1, option1_rate, term)
        total1 = round(monthly1 * term, 2)
        
        comparison = PaymentComparison(
            term_months=term,
            option1_rate=option1_rate,
            option1_monthly=monthly1,
            option1_total=total1,
            option1_rebate=consumer_cash
        )
        
        # Option 2: Sans rabais + taux réduits (si disponible)
        if program_obj.option2_rates:
            option2_rate = get_rate_for_term(program_obj.option2_rates, term)
            principal2 = vehicle_price  # Pas de rabais
            monthly2 = calculate_monthly_payment(principal2, option2_rate, term)
            total2 = round(monthly2 * term, 2)
            
            comparison.option2_rate = option2_rate
            comparison.option2_monthly = monthly2
            comparison.option2_total = total2
            
            # Déterminer la meilleure option (coût total le plus bas)
            if total1 < total2:
                comparison.best_option = "1"
                comparison.savings = round(total2 - total1, 2)
            elif total2 < total1:
                comparison.best_option = "2"
                comparison.savings = round(total1 - total2, 2)
            else:
                comparison.best_option = "1"
                comparison.savings = 0
        
        comparisons.append(comparison)
    
    return CalculationResponse(
        vehicle_price=vehicle_price,
        consumer_cash=consumer_cash,
        bonus_cash=bonus_cash,
        brand=program_obj.brand,
        model=program_obj.model,
        trim=program_obj.trim,
        year=program_obj.year,
        comparisons=comparisons
    )

# Seed initial data from PDF pages 20-21 (Février 2026)
@api_router.post("/seed")
async def seed_data():
    """
    Seed les données initiales à partir du PDF Février 2026
    Pages 20 (2026) et 21 (2025)
    """
    # Clear existing data
    await db.programs.delete_many({})
    
    # Taux standard 4.99% pour la plupart des véhicules
    std = {"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}
    
    # Mois/Année du programme
    prog_month = 2
    prog_year = 2026
    
    programs_data = [
        # ==================== 2026 MODELS (Page 20) ====================
        
        # CHRYSLER 2026
        {"brand": "Chrysler", "model": "Grand Caravan", "trim": "SXT", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49}, 
         "bonus_cash": 0},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "PHEV", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49}, 
         "bonus_cash": 0},
        
        # Pacifica (excluding PHEV): Taux bas (0% jusqu'à 72mo, 1.99% à 84mo, 3.49% à 96mo), PAS d'Option 2
        {"brand": "Chrysler", "model": "Pacifica", "trim": "(excluding PHEV)", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 3.49},
         "option2_rates": None, "bonus_cash": 0},
        
        # JEEP COMPASS 2026 - Taux Option 2 corrigés
        {"brand": "Jeep", "model": "Compass", "trim": "Sport", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49}, 
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "North", "year": 2026,
         "consumer_cash": 3500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "North w/ Altitude Package (ADZ)", "year": 2026,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Trailhawk", "year": 2026,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Limited", "year": 2026,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        # JEEP CHEROKEE 2026
        {"brand": "Jeep", "model": "Cherokee", "trim": "Base (KMJL74)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Cherokee", "trim": "(excluding Base)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # JEEP WRANGLER 2026 - Corrigé avec Option 2
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door (JL) non Rubicon", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door Rubicon (JL)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99}, 
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (excl. 392 et 4xe)", "year": 2026,
         "consumer_cash": 5250, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99}, 
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door MOAB 392", "year": 2026,
         "consumer_cash": 6000, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # JEEP GLADIATOR 2026
        {"brand": "Jeep", "model": "Gladiator", "trim": "Sport S, Willys, Sahara, Willys '41", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Gladiator", "trim": "(excl. Sport S, Willys, Sahara, Willys '41)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49}, 
         "bonus_cash": 0},
        
        # JEEP GRAND CHEROKEE 2026 - Taux variables Option 1 et Option 2 corrigés
        {"brand": "Jeep", "model": "Grand Cherokee/L", "trim": "Laredo/Laredo X", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 1.99, "rate_48": 2.99, "rate_60": 3.49, "rate_72": 3.99, "rate_84": 4.49, "rate_96": 4.99},
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Grand Cherokee/L", "trim": "Altitude", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 1.99, "rate_48": 2.99, "rate_60": 3.49, "rate_72": 3.99, "rate_84": 4.49, "rate_96": 4.99},
         "option2_rates": None,
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Grand Cherokee/L", "trim": "Limited/Limited Reserve/Summit", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 1.99, "rate_48": 2.99, "rate_60": 3.49, "rate_72": 3.99, "rate_84": 4.49, "rate_96": 4.99},
         "option2_rates": None,
         "bonus_cash": 0},
        
        # JEEP GRAND WAGONEER 2026
        {"brand": "Jeep", "model": "Grand Wagoneer/L", "trim": None, "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.99}, 
         "bonus_cash": 0},
        
        # DODGE DURANGO 2026
        {"brand": "Dodge", "model": "Durango", "trim": "SXT, GT, GT Plus", "year": 2026,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Dodge", "model": "Durango", "trim": "GT Hemi V8 Plus, GT Hemi V8 Premium", "year": 2026,
         "consumer_cash": 9000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Dodge", "model": "Durango", "trim": "SRT Hellcat", "year": 2026,
         "consumer_cash": 15500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        # DODGE CHARGER 2026
        {"brand": "Dodge", "model": "Charger", "trim": "2-Door & 4-Door (ICE)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # RAM 2026 - Corrigé
        {"brand": "Ram", "model": "ProMaster", "trim": None, "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Tradesman, Express, Warlock", "year": 2026,
         "consumer_cash": 6500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Big Horn", "year": 2026,
         "consumer_cash": 6000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Sport, Rebel", "year": 2026,
         "consumer_cash": 8250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Laramie (DT6P98)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Laramie, Limited, Longhorn, Tungsten, RHO (excl. DT6P98)", "year": 2026,
         "consumer_cash": 11500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500 Power Wagon Crew Cab", "trim": "(DJ7X91 2UP)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "Gas Models", "year": 2026,
         "consumer_cash": 7000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "Diesel Models", "year": 2026,
         "consumer_cash": 5000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 0.99, "rate_60": 0.99, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "Chassis Cab", "trim": None, "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # ==================== 2025 MODELS (Page 21) ====================
        
        # CHRYSLER 2025 - Corrigé
        {"brand": "Chrysler", "model": "Grand Caravan", "trim": "SXT", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "Hybrid", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 1.99, "rate_72": 2.99, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "Select Models (excl. Hybrid)", "year": 2025,
         "consumer_cash": 0, 
         "option1_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 1.99, "rate_72": 2.99, "rate_84": 3.99, "rate_96": 4.99},
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "(excl. Select & Hybrid)", "year": 2025,
         "consumer_cash": 750, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        # JEEP COMPASS 2025 - Corrigé
        {"brand": "Jeep", "model": "Compass", "trim": "Sport", "year": 2025,
         "consumer_cash": 5500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Compass", "trim": "North", "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0.99, "rate_60": 1.99, "rate_72": 1.99, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Altitude, Trailhawk, Trailhawk Elite", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Limited", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        # JEEP WRANGLER 2025 - Corrigé
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (JL) 4xe (JLXL74)", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (JL) 4xe (excl. JLXL74)", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door (JL) non Rubicon", "year": 2025,
         "consumer_cash": 750, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door Rubicon (JL)", "year": 2025,
         "consumer_cash": 8500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door Rubicon w/ 2.0L", "year": 2025,
         "consumer_cash": 0, 
         "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (JL) (excl. Rubicon 2.0L & 4xe)", "year": 2025,
         "consumer_cash": 8500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        # JEEP GLADIATOR 2025
        {"brand": "Jeep", "model": "Gladiator", "trim": None, "year": 2025,
         "consumer_cash": 11000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        # JEEP GRAND CHEROKEE 2025 - Corrigé
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "4xe (WL)", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Laredo (WLJH74 2*A)", "year": 2025,
         "consumer_cash": 6000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Altitude (WLJH74 2*B)", "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Summit (WLJT74 23S)", "year": 2025,
         "consumer_cash": 0, 
         "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "(WL) (excl. Laredo, Altitude, Summit, 4xe)", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        # JEEP GRAND CHEROKEE L 2025 - Corrigé
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Laredo (WLJH75 2*A)", "year": 2025,
         "consumer_cash": 6000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Altitude (WLJH75 2*B)", "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Overland (WLJS75)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "(WL) (excl. Laredo, Altitude, Overland)", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        # JEEP WAGONEER 2025 - Corrigé
        {"brand": "Jeep", "model": "Wagoneer/L", "trim": None, "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Wagoneer/L", "trim": None, "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wagoneer S", "trim": "Limited & Premium (BEV)", "year": 2025,
         "consumer_cash": 8000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        # DODGE DURANGO 2025 - Corrigé
        {"brand": "Dodge", "model": "Durango", "trim": "GT, GT Plus", "year": 2025,
         "consumer_cash": 8000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Durango", "trim": "R/T, R/T Plus, R/T 20th Anniversary", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Durango", "trim": "SRT Hellcat", "year": 2025,
         "consumer_cash": 16000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        # DODGE CHARGER 2025 - Corrigé
        {"brand": "Dodge", "model": "Charger Daytona", "trim": "R/T (BEV)", "year": 2025,
         "consumer_cash": 3000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Charger Daytona", "trim": "R/T Plus (BEV)", "year": 2025,
         "consumer_cash": 5000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Charger Daytona", "trim": "Scat Pack (BEV)", "year": 2025,
         "consumer_cash": 7000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        # DODGE HORNET 2025
        {"brand": "Dodge", "model": "Hornet", "trim": "RT (PHEV)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Hornet", "trim": "RT Plus (PHEV)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Hornet", "trim": "GT (Gas)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Hornet", "trim": "GT Plus (Gas)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        # RAM 2025 - Corrigé
        {"brand": "Ram", "model": "ProMaster", "trim": None, "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Tradesman, Warlock, Express (DT)", "year": 2025,
         "consumer_cash": 9250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Big Horn (DT) w/ Off-Roader Value Package (4KF)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Big Horn (DT) (excl. Off-Roader)", "year": 2025,
         "consumer_cash": 9250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Sport, Rebel (DT)", "year": 2025,
         "consumer_cash": 10000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Laramie, Limited, Longhorn, Tungsten, RHO (DT)", "year": 2025,
         "consumer_cash": 12250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "Gas Models (excl. Chassis Cab, Diesel)", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "6.7L High Output Diesel (ETM)", "year": 2025,
         "consumer_cash": 7000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 0.99, "rate_60": 0.99, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "Chassis Cab", "trim": None, "year": 2025,
         "consumer_cash": 5000, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # FIAT 2025
        {"brand": "Fiat", "model": "500e", "trim": "BEV", "year": 2025,
         "consumer_cash": 6000,
         "option1_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 1.99, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "option2_rates": None, "bonus_cash": 5000},
    ]
    
    for prog_data in programs_data:
        if prog_data.get("option2_rates") and isinstance(prog_data["option2_rates"], dict):
            prog_data["option2_rates"] = FinancingRates(**prog_data["option2_rates"]).dict()
        if isinstance(prog_data["option1_rates"], dict):
            prog_data["option1_rates"] = FinancingRates(**prog_data["option1_rates"]).dict()
        prog_data["program_month"] = prog_month
        prog_data["program_year"] = prog_year
        prog = VehicleProgram(**prog_data)
        await db.programs.insert_one(prog.dict())
    
    return {"message": f"Seeded {len(programs_data)} programs for {prog_month}/{prog_year}"}

# ============ Excel Generation Function ============

def generate_excel_from_programs(programs: List[Dict[str, Any]], program_month: int, program_year: int) -> bytes:
    """Génère un fichier Excel selon le format du PDF Stellantis"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programmes"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    option1_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")  # Rouge pour Option 1
    option2_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")  # Bleu pour Option 2
    bonus_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Vert pour Bonus
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Month names
    month_names = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    # Title Row 1
    ws.merge_cells('A1:S1')
    ws['A1'] = f"PROGRAMMES DE FINANCEMENT RETAIL - {month_names[program_month].upper()} {program_year}"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Sub-header Row 2 - Group headers
    ws.merge_cells('A2:D2')
    ws['A2'] = "VÉHICULE"
    ws['A2'].font = header_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = header_alignment
    
    # Option 1 header with rabais
    ws.merge_cells('E2:K2')
    ws['E2'] = "OPTION 1 - Consumer Cash + Taux Standard"
    ws['E2'].font = header_font
    ws['E2'].fill = option1_fill
    ws['E2'].alignment = header_alignment
    
    # Option 2 header with rabais
    ws.merge_cells('L2:R2')
    ws['L2'] = "OPTION 2 - Alternative Consumer Cash + Taux Réduit"
    ws['L2'].font = header_font
    ws['L2'].fill = option2_fill
    ws['L2'].alignment = header_alignment
    
    # Bonus header
    ws['S2'] = "BONUS"
    ws['S2'].font = header_font
    ws['S2'].fill = bonus_fill
    ws['S2'].alignment = header_alignment
    
    # Detail Headers Row 3
    headers = [
        "Marque", "Modèle", "Version", "Année",  # Véhicule (A-D)
        "Rabais ($)", "36m", "48m", "60m", "72m", "84m", "96m",  # Option 1: Rabais + Taux (E-K)
        "Rabais ($)", "36m", "48m", "60m", "72m", "84m", "96m",  # Option 2: Rabais + Taux (L-R)
        "Bonus ($)"  # Bonus (S)
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = header_alignment
        cell.border = thin_border
        
        if col <= 4:
            cell.fill = header_fill
        elif col <= 11:  # Option 1
            cell.fill = option1_fill
        elif col <= 18:  # Option 2
            cell.fill = option2_fill
        else:  # Bonus
            cell.fill = bonus_fill
    
    # Data rows
    for row_idx, prog in enumerate(programs, 4):
        opt1_rates = prog.get("option1_rates") or {}
        opt2_rates = prog.get("option2_rates") or {}
        
        consumer_cash = prog.get("consumer_cash", 0) or 0
        alt_consumer_cash = prog.get("alt_consumer_cash", 0) or 0  # Alternative Consumer Cash pour Option 2
        bonus_cash = prog.get("bonus_cash", 0) or 0
        
        # Format rate display
        def format_rate(rate):
            if rate is None or rate == "-":
                return "-"
            try:
                r = float(rate)
                return f"{r:.2f}%" if r > 0 else "0%"
            except:
                return "-"
        
        data = [
            prog.get("brand", ""),
            prog.get("model", ""),
            prog.get("trim", "") or "",
            prog.get("year", ""),
            # Option 1: Rabais + Taux
            f"${consumer_cash:,.0f}" if consumer_cash else "-",
            format_rate(opt1_rates.get("rate_36")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_48")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_60")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_72")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_84")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_96")) if opt1_rates else "-",
            # Option 2: Rabais + Taux
            f"${alt_consumer_cash:,.0f}" if alt_consumer_cash else "-",
            format_rate(opt2_rates.get("rate_36")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_48")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_60")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_72")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_84")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_96")) if opt2_rates else "-",
            # Bonus
            f"${bonus_cash:,.0f}" if bonus_cash else "-",
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Color coding for columns
            if col >= 5 and col <= 11:  # Option 1
                cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            elif col >= 12 and col <= 18:  # Option 2
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            elif col == 19:  # Bonus
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # Adjust column widths
    column_widths = [12, 18, 28, 7, 12, 8, 8, 8, 8, 8, 8, 12, 8, 8, 8, 8, 8, 8, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Row height for headers
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 25
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def send_excel_email(excel_data: bytes, admin_email: str, program_month: int, program_year: int, program_count: int):
    """Envoie le fichier Excel par email"""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    
    month_names = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    msg = MIMEMultipart()
    msg['From'] = SMTP_EMAIL
    msg['To'] = admin_email
    msg['Subject'] = f"CalcAuto AiPro - Extraction PDF {month_names[program_month]} {program_year}"
    
    body = f"""
Bonjour,

L'extraction du PDF des programmes de financement est terminée.

📊 Résumé:
• Période: {month_names[program_month]} {program_year}
• Programmes extraits: {program_count}

Le fichier Excel est joint à cet email pour vérification.

⚠️ IMPORTANT: Veuillez vérifier les données dans le fichier Excel avant de confirmer l'import dans l'application.

---
CalcAuto AiPro
    """
    
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # Attach Excel file
    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment.set_payload(excel_data)
    encoders.encode_base64(attachment)
    filename = f"programmes_{month_names[program_month].lower()}_{program_year}.xlsx"
    attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
    msg.attach(attachment)
    
    # Send email
    try:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        logger.error(f"Error sending Excel email: {str(e)}")
        return False

# ============ PDF Import with AI ============

class PDFExtractRequest(BaseModel):
    password: str
    program_month: int
    program_year: int

class ProgramPreview(BaseModel):
    """Preview of extracted program for validation"""
    brand: str
    model: str
    trim: Optional[str] = None
    year: int
    consumer_cash: float = 0
    bonus_cash: float = 0
    option1_rates: Dict[str, float]
    option2_rates: Optional[Dict[str, float]] = None
    
class ExtractedDataResponse(BaseModel):
    success: bool
    message: str
    programs: List[Dict[str, Any]] = []
    raw_text: str = ""

class SaveProgramsRequest(BaseModel):
    password: str
    programs: List[Dict[str, Any]]
    program_month: int
    program_year: int

@api_router.post("/verify-password")
async def verify_password(password: str = Form(...)):
    """Vérifie le mot de passe admin"""
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    return {"success": True, "message": "Mot de passe vérifié"}

@api_router.post("/extract-pdf", response_model=ExtractedDataResponse)
async def extract_pdf(
    file: UploadFile = File(...),
    password: str = Form(...),
    program_month: int = Form(...),
    program_year: int = Form(...),
    start_page: int = Form(1),
    end_page: int = Form(9999)
):
    """
    Extrait les données de financement d'un PDF via OpenAI GPT-4
    Retourne les programmes pour prévisualisation/modification avant sauvegarde
    
    start_page et end_page permettent de limiter l'extraction à certaines pages
    (indexation commence à 1)
    """
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    
    if not OPENAI_API_KEY:
        raise HTTPException(status_code=500, detail="Clé OpenAI non configurée")
    
    try:
        import tempfile
        import os as os_module
        import base64
        from openai import OpenAI
        import PyPDF2
        
        # Save uploaded PDF temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        try:
            # Extract text ONLY from specified pages using PyPDF2
            pdf_text = ""
            with open(tmp_path, 'rb') as pdf_file:
                reader = PyPDF2.PdfReader(pdf_file)
                total_pages = len(reader.pages)
                
                # Convert to 0-based index and validate range
                start_idx = max(0, start_page - 1)  # Convert 1-based to 0-based
                end_idx = min(total_pages, end_page)  # Keep as-is (exclusive end)
                
                logger.info(f"PDF has {total_pages} pages. Extracting pages {start_page} to {end_page} (indices {start_idx} to {end_idx})")
                
                # Only extract the specified pages
                for page_num in range(start_idx, end_idx):
                    page = reader.pages[page_num]
                    page_text = page.extract_text()
                    pdf_text += f"\n--- PAGE {page_num + 1} ---\n{page_text}\n"
                
                logger.info(f"Extracted {end_idx - start_idx} pages, total text length: {len(pdf_text)} characters")
            
            # Use OpenAI to extract structured data
            client = OpenAI(api_key=OPENAI_API_KEY)
            
            extraction_prompt = f"""EXTRAIS TOUS LES VÉHICULES de ce PDF de programmes de financement FCA Canada.

TEXTE COMPLET DU PDF:
{pdf_text}

=== FORMAT DES LIGNES DU PDF ===
Chaque ligne suit ce format:
VÉHICULE [Consumer Cash $X,XXX] [6 taux Option1] [6 taux Option2] [Bonus Cash]

- Si tu vois "- - - - - -" = option non disponible (null)
- Si tu vois "P" avant un montant = c'est quand même le montant
- 6 taux = 36M, 48M, 60M, 72M, 84M, 96M

=== BONUS CASH - TRÈS IMPORTANT ===
Le Bonus Cash est dans la DERNIÈRE COLONNE (colonne verte "Bonus Cash").
- CHAQUE véhicule a son PROPRE montant de Bonus Cash (peut être 0, $1000, $3000, etc.)
- Le Bonus Cash n'est PAS le même pour tous les véhicules!
- Si la colonne Bonus Cash est vide ou montre "-", alors bonus_cash = 0
- Les montants typiques sont: $0, $1,000, $3,000

=== EXEMPLES AVEC BONUS CASH ===

2026 MODELS (généralement PAS de Bonus Cash):
"Grand Caravan SXT    4.99%... " → bonus_cash: 0
"Ram 1500 Big Horn    $6,000  4.99%..." → bonus_cash: 0

2025 MODELS (souvent avec Bonus Cash de $1,000 à $3,000):
"Compass North  $7,500  4.99%...   $1,000" → bonus_cash: 1000
"Ram 1500 Sport  $10,000  4.99%...  $3,000" → bonus_cash: 3000
"Ram 2500/3500 Gas Models  $9,500  4.99%...  -" → bonus_cash: 0 (pas de bonus!)
"Ram Chassis Cab  $5,000  4.99%...  -" → bonus_cash: 0

=== RÈGLE IMPORTANTE POUR RAM 2025 ===
- Ram 1500 2025: bonus_cash = $3,000
- Ram 2500/3500 2025: bonus_cash = 0 (PAS de bonus!)
- Ram ProMaster 2025: bonus_cash = 0
- Ram Chassis Cab 2025: bonus_cash = 0

=== MARQUES À EXTRAIRE ===
- CHRYSLER: Grand Caravan, Pacifica
- JEEP: Compass, Cherokee, Wrangler, Gladiator, Grand Cherokee, Grand Wagoneer
- DODGE: Durango, Charger, Hornet
- RAM: ProMaster, 1500, 2500, 3500, Chassis Cab

=== ANNÉES ===
- "2026 MODELS" → year: 2026
- "2025 MODELS" → year: 2025
Extrais les véhicules des DEUX sections!

=== JSON REQUIS ===
{{
    "programs": [
        {{
            "brand": "Chrysler",
            "model": "Grand Caravan", 
            "trim": "SXT",
            "year": 2026,
            "consumer_cash": 0,
            "bonus_cash": 0,
            "option1_rates": {{"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}},
            "option2_rates": null
        }},
        ... TOUS les autres véhicules ...
    ]
}}

EXTRAIS ABSOLUMENT TOUS LES VÉHICULES DES SECTIONS 2026 ET 2025. 
VÉRIFIE LE BONUS CASH POUR CHAQUE VÉHICULE INDIVIDUELLEMENT!"""

            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "Tu extrais TOUS les véhicules d'un PDF FCA Canada. CHAQUE ligne = 1 entrée. N'oublie AUCUN véhicule. Sections 2026 ET 2025. JSON valide uniquement."},
                    {"role": "user", "content": extraction_prompt}
                ],
                temperature=0.1,
                max_tokens=16000,
                response_format={"type": "json_object"}
            )
            
            response_text = response.choices[0].message.content.strip()
            
            # Clean up response (remove markdown code blocks if present)
            if response_text.startswith("```"):
                lines = response_text.split("\n")
                response_text = "\n".join(lines[1:-1] if lines[-1] == "```" else lines[1:])
                response_text = response_text.strip()
            if response_text.endswith("```"):
                response_text = response_text[:-3].strip()
            
            # Clean common JSON issues
            response_text = response_text.replace('\n', ' ').replace('\r', '')
            response_text = re.sub(r',\s*}', '}', response_text)  # Remove trailing commas
            response_text = re.sub(r',\s*]', ']', response_text)  # Remove trailing commas in arrays
            
            try:
                data = json.loads(response_text)
                programs = data.get("programs", [])
            except json.JSONDecodeError as e:
                # Try to fix common issues and retry
                try:
                    # Find the programs array and try to parse it
                    programs_match = re.search(r'"programs"\s*:\s*\[(.*)\]', response_text, re.DOTALL)
                    if programs_match:
                        programs_str = programs_match.group(1)
                        # Try to parse individual objects
                        programs = []
                        obj_pattern = r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}'
                        for obj_match in re.finditer(obj_pattern, programs_str):
                            try:
                                obj = json.loads(obj_match.group())
                                programs.append(obj)
                            except:
                                continue
                        if programs:
                            data = {"programs": programs}
                        else:
                            raise e
                    else:
                        raise e
                except:
                    return ExtractedDataResponse(
                        success=False,
                        message=f"Erreur de parsing JSON: {str(e)}",
                        programs=[],
                        raw_text=response_text[:3000]
                    )
            
            # Validate and clean programs
            valid_programs = []
            for p in programs:
                # Ensure required fields exist
                if 'brand' in p and 'model' in p:
                    # Clean up rates
                    if p.get('option1_rates') and isinstance(p['option1_rates'], dict):
                        for key in ['rate_36', 'rate_48', 'rate_60', 'rate_72', 'rate_84', 'rate_96']:
                            if key not in p['option1_rates']:
                                p['option1_rates'][key] = 4.99
                    if p.get('option2_rates') and isinstance(p['option2_rates'], dict):
                        for key in ['rate_36', 'rate_48', 'rate_60', 'rate_72', 'rate_84', 'rate_96']:
                            if key not in p['option2_rates']:
                                p['option2_rates'][key] = 0
                    valid_programs.append(p)
            
            # Generate Excel and send by email
            excel_sent = False
            if EXCEL_AVAILABLE and valid_programs and SMTP_EMAIL:
                try:
                    excel_data = generate_excel_from_programs(valid_programs, program_month, program_year)
                    excel_sent = send_excel_email(excel_data, SMTP_EMAIL, program_month, program_year, len(valid_programs))
                    logger.info(f"Excel generated and sent: {excel_sent}")
                except Exception as excel_error:
                    logger.error(f"Error generating/sending Excel: {str(excel_error)}")
            
            # AUTO-SAVE: Sauvegarder automatiquement les programmes dans la base de données
            saved_count = 0
            try:
                # D'abord, supprimer les anciens programmes de la même période
                delete_result = await db.programs.delete_many({
                    "program_month": program_month,
                    "program_year": program_year
                })
                logger.info(f"Deleted {delete_result.deleted_count} old programs for {program_month}/{program_year}")
                
                # Taux par défaut (4.99% standard)
                default_rates = {
                    "rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99,
                    "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99
                }
                
                # Ensuite, ajouter les nouveaux programmes
                for prog in valid_programs:
                    # S'assurer que option1_rates n'est pas None
                    opt1 = prog.get("option1_rates")
                    if opt1 is None or not isinstance(opt1, dict):
                        opt1 = default_rates.copy()
                    
                    program_doc = {
                        "id": str(uuid.uuid4()),
                        "brand": prog.get("brand", ""),
                        "model": prog.get("model", ""),
                        "trim": prog.get("trim", ""),
                        "year": prog.get("year", program_year),
                        "consumer_cash": prog.get("consumer_cash", 0) or 0,
                        "bonus_cash": prog.get("bonus_cash", 0) or 0,
                        "alt_consumer_cash": prog.get("alt_consumer_cash", 0) or 0,
                        "option1_rates": opt1,
                        "option2_rates": prog.get("option2_rates"),
                        "program_month": program_month,
                        "program_year": program_year,
                        "created_at": datetime.utcnow().isoformat()
                    }
                    await db.programs.insert_one(program_doc)
                    saved_count += 1
                
                logger.info(f"Auto-saved {saved_count} programs for {program_month}/{program_year}")
            except Exception as save_error:
                logger.error(f"Error auto-saving programs: {str(save_error)}")
            
            return ExtractedDataResponse(
                success=True,
                message=f"Extrait et sauvegardé {len(valid_programs)} programmes" + (" - Excel envoyé par email!" if excel_sent else ""),
                programs=valid_programs,
                raw_text=""
            )
            
        finally:
            # Clean up temp file
            os_module.unlink(tmp_path)
            
    except Exception as e:
        logger.error(f"Error extracting PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur d'extraction: {str(e)}")

@api_router.post("/save-programs")
async def save_programs(request: SaveProgramsRequest):
    """
    Sauvegarde les programmes validés dans la base de données
    Remplace les programmes existants pour le mois/année spécifié
    Garde seulement les 6 derniers mois d'historique
    """
    if request.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    
    # Delete existing programs for this period
    await db.programs.delete_many({
        "program_month": request.program_month,
        "program_year": request.program_year
    })
    
    # Insert new programs
    inserted = 0
    skipped = 0
    default_rates = {"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}
    
    for prog_data in request.programs:
        # Skip invalid entries (missing brand/model)
        if not prog_data.get("brand") or not prog_data.get("model"):
            skipped += 1
            continue
            
        # Ensure option1_rates has a default value if missing
        if not prog_data.get("option1_rates"):
            prog_data["option1_rates"] = default_rates.copy()
        elif isinstance(prog_data["option1_rates"], dict):
            prog_data["option1_rates"] = FinancingRates(**prog_data["option1_rates"]).dict()
            
        # Process option2_rates if present
        if prog_data.get("option2_rates") and isinstance(prog_data["option2_rates"], dict):
            prog_data["option2_rates"] = FinancingRates(**prog_data["option2_rates"]).dict()
        
        prog_data["program_month"] = request.program_month
        prog_data["program_year"] = request.program_year
        prog_data["bonus_cash"] = prog_data.get("bonus_cash", 0)
        prog_data["consumer_cash"] = prog_data.get("consumer_cash", 0)
        
        try:
            prog = VehicleProgram(**prog_data)
            await db.programs.insert_one(prog.dict())
            inserted += 1
        except Exception as e:
            logger.warning(f"Skipped invalid program: {prog_data.get('brand')} {prog_data.get('model')} - {str(e)}")
            skipped += 1
            continue
    
    # Clean up old programs (keep only 6 months)
    await cleanup_old_programs()
    
    # Calculate brands summary for report
    brands_summary = {}
    for prog_data in request.programs:
        brand = prog_data.get("brand", "Inconnu")
        if brand not in brands_summary:
            brands_summary[brand] = 0
        brands_summary[brand] += 1
    
    # Send automatic email report
    try:
        await send_import_report_email(
            programs_count=inserted,
            program_month=request.program_month,
            program_year=request.program_year,
            brands_summary=brands_summary,
            skipped_count=skipped
        )
        logger.info(f"Import report email sent to {SMTP_EMAIL}")
    except Exception as e:
        logger.warning(f"Failed to send import report email: {str(e)}")
    
    return {
        "success": True,
        "message": f"Sauvegardé {inserted} programmes pour {request.program_month}/{request.program_year}" + (f" ({skipped} ignorés)" if skipped > 0 else "") + " - Rapport envoyé par email"
    }

async def send_import_report_email(programs_count: int, program_month: int, program_year: int, brands_summary: dict, skipped_count: int = 0):
    """Envoie automatiquement un rapport par email après l'import des programmes"""
    months_fr = {
        1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
        5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
        9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
    }
    month_name = months_fr.get(program_month, str(program_month))
    
    # Generate brands table
    brands_rows = ""
    total_programs = 0
    for brand in ['Chrysler', 'Jeep', 'Dodge', 'Ram', 'Fiat']:
        count = brands_summary.get(brand, 0)
        if count > 0:
            brands_rows += f"<tr><td style='padding: 10px; border-bottom: 1px solid #eee;'>{brand}</td><td style='padding: 10px; border-bottom: 1px solid #eee; text-align: center; font-weight: bold;'>{count}</td></tr>"
            total_programs += count
    
    # Add any other brands not in the standard list
    for brand, count in brands_summary.items():
        if brand not in ['Chrysler', 'Jeep', 'Dodge', 'Ram', 'Fiat'] and count > 0:
            brands_rows += f"<tr><td style='padding: 10px; border-bottom: 1px solid #eee;'>{brand}</td><td style='padding: 10px; border-bottom: 1px solid #eee; text-align: center; font-weight: bold;'>{count}</td></tr>"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; background: #fff; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); overflow: hidden; }}
            .header {{ background: linear-gradient(135deg, #1a5f4a 0%, #2d8f6f 100%); color: #fff; padding: 25px; text-align: center; }}
            .header h1 {{ margin: 0; font-size: 24px; }}
            .header .subtitle {{ margin-top: 8px; opacity: 0.9; }}
            .content {{ padding: 25px; }}
            .success-badge {{ background: #d4edda; color: #155724; padding: 15px; border-radius: 8px; text-align: center; margin-bottom: 20px; font-size: 16px; }}
            .success-badge strong {{ font-size: 18px; }}
            .stats-box {{ background: #f8f9fa; border-radius: 8px; padding: 20px; margin-bottom: 20px; text-align: center; }}
            .big-number {{ font-size: 56px; font-weight: bold; color: #1a5f4a; }}
            .big-label {{ color: #666; font-size: 14px; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th {{ background: #1a5f4a; color: #fff; padding: 12px; text-align: left; }}
            .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #eee; }}
            .warning-box {{ background: #fff3cd; border: 1px solid #ffc107; border-radius: 6px; padding: 12px; margin-top: 15px; color: #856404; font-size: 13px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>✅ Import Réussi!</h1>
                <div class="subtitle">CalcAuto AiPro - Rapport d'import automatique</div>
            </div>
            <div class="content">
                <div class="success-badge">
                    <strong>🎉 Programmes {month_name} {program_year}</strong><br>
                    importés avec succès!
                </div>
                
                <div class="stats-box">
                    <div class="big-number">{programs_count}</div>
                    <div class="big-label">programmes de financement</div>
                </div>
                
                <h3 style="color: #1a5f4a; border-bottom: 2px solid #1a5f4a; padding-bottom: 8px;">📊 Répartition par marque</h3>
                <table>
                    <thead>
                        <tr>
                            <th>Marque</th>
                            <th style="text-align: center;">Nombre</th>
                        </tr>
                    </thead>
                    <tbody>
                        {brands_rows}
                        <tr style="background: #f8f9fa; font-weight: bold;">
                            <td style="padding: 12px;">TOTAL</td>
                            <td style="padding: 12px; text-align: center;">{programs_count}</td>
                        </tr>
                    </tbody>
                </table>
                
                {f'<div class="warning-box">⚠️ {skipped_count} programme(s) ignoré(s) en raison de données invalides.</div>' if skipped_count > 0 else ''}
                
                <p style="margin-top: 25px; color: #666; font-size: 14px;">
                    Les nouveaux programmes sont maintenant disponibles dans l'application CalcAuto AiPro. 
                    Vos clients peuvent commencer à utiliser les nouveaux taux immédiatement.
                </p>
            </div>
            <div class="footer">
                <p style="margin: 0;"><strong>CalcAuto AiPro</strong></p>
                <p style="margin: 8px 0 0;">Rapport généré automatiquement le {datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    subject = f"✅ Import {month_name} {program_year} - {programs_count} programmes"
    
    send_email(SMTP_EMAIL, subject, html_body)

async def cleanup_old_programs():
    """Supprime les programmes de plus de 6 mois"""
    # Get all unique periods
    pipeline = [
        {"$group": {
            "_id": {"month": "$program_month", "year": "$program_year"}
        }},
        {"$sort": {"_id.year": -1, "_id.month": -1}}
    ]
    periods = await db.programs.aggregate(pipeline).to_list(100)
    
    # Keep only the 6 most recent periods
    if len(periods) > 6:
        periods_to_delete = periods[6:]
        for p in periods_to_delete:
            await db.programs.delete_many({
                "program_month": p["_id"]["month"],
                "program_year": p["_id"]["year"]
            })
            logger.info(f"Deleted old programs for {p['_id']['month']}/{p['_id']['year']}")

# ============ Email Functions ============

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import io

def send_email(to_email: str, subject: str, html_body: str, attachment_data: bytes = None, attachment_name: str = None):
    """Envoie un email via Gmail SMTP"""
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        raise Exception("Configuration SMTP manquante")
    
    msg = MIMEMultipart('alternative')
    msg['From'] = f"CalcAuto AiPro <{SMTP_EMAIL}>"
    msg['To'] = to_email
    msg['Subject'] = subject
    
    # Corps HTML
    msg.attach(MIMEText(html_body, 'html', 'utf-8'))
    
    # Pièce jointe si fournie
    if attachment_data and attachment_name:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment_data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{attachment_name}"')
        msg.attach(part)
    
    # Connexion SMTP
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        server.send_message(msg)
    
    return True

class SendCalculationEmailRequest(BaseModel):
    """Requête pour envoyer un calcul par email"""
    client_email: str
    client_name: str = ""
    vehicle_info: Dict[str, Any]
    calculation_results: Dict[str, Any]
    selected_term: int = 60
    selected_option: str = "1"
    vehicle_price: float
    payment_frequency: str = "monthly"  # monthly, biweekly, weekly
    dealer_name: str = "CalcAuto AiPro"
    dealer_phone: str = ""
    # New fields for complete email
    rates_table: Dict[str, Any] = {}  # Option 1 & 2 rates for all terms
    fees: Dict[str, float] = {}  # frais_dossier, taxe_pneus, frais_rdprm
    trade_in: Dict[str, float] = {}  # valeur_echange, montant_du

class SendReportEmailRequest(BaseModel):
    """Requête pour envoyer un rapport après import"""
    programs_count: int
    program_month: int
    program_year: int
    brands_summary: Dict[str, int]

@api_router.post("/send-calculation-email")
async def send_calculation_email(request: SendCalculationEmailRequest):
    """Envoie un calcul de financement par email - STYLE PDF CLAIR"""
    try:
        vehicle = request.vehicle_info
        calc = request.calculation_results
        term = request.selected_term
        freq = request.payment_frequency
        rates = request.rates_table
        fees = request.fees
        trade = request.trade_in
        
        # Get comparison data
        comparison = None
        for c in calc.get('comparisons', []):
            if c.get('term_months') == term:
                comparison = c
                break
        
        if not comparison:
            comparison = calc.get('comparisons', [{}])[0] if calc.get('comparisons') else {}
        
        consumer_cash = calc.get('consumer_cash', 0)
        bonus_cash = calc.get('bonus_cash', 0)
        
        option1_rate = comparison.get('option1_rate', 0)
        option2_rate = comparison.get('option2_rate', 0)
        
        principal_option1 = comparison.get('principal_option1', request.vehicle_price - consumer_cash)
        principal_option2 = comparison.get('principal_option2', request.vehicle_price)
        
        total_option1 = comparison.get('option1_total', 0)
        total_option2 = comparison.get('option2_total', 0)
        
        if freq == 'weekly':
            option1_payment = comparison.get('option1_weekly', 0)
            option2_payment = comparison.get('option2_weekly', 0)
            freq_label = "Hebdomadaire"
        elif freq == 'biweekly':
            option1_payment = comparison.get('option1_biweekly', 0)
            option2_payment = comparison.get('option2_biweekly', 0)
            freq_label = "Aux 2 semaines"
        else:
            option1_payment = comparison.get('option1_monthly', 0)
            option2_payment = comparison.get('option2_monthly', 0)
            freq_label = "Mensuel"
        
        best_option = comparison.get('best_option', '1')
        savings = comparison.get('savings', 0)
        has_option2 = option2_rate is not None and option2_rate > 0 and option2_payment > 0
        
        def fmt(val):
            return f"{val:,.0f}".replace(",", " ")
        
        def fmt2(val):
            return f"{val:,.2f}".replace(",", " ").replace(".", ",")
        
        # Build fees info
        frais_dossier = fees.get('frais_dossier', 0)
        taxe_pneus = fees.get('taxe_pneus', 0)
        frais_rdprm = fees.get('frais_rdprm', 0)
        valeur_echange = trade.get('valeur_echange', 0)
        montant_du = trade.get('montant_du', 0)
        
        # LIGHT THEME EMAIL - PDF STYLE
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{ font-family: Arial, Helvetica, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; color: #333; font-size: 14px; line-height: 1.5; }}
                .container {{ max-width: 600px; margin: 0 auto; background: #fff; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); overflow: hidden; }}
                
                .header {{ background: #1a5f4a; color: #fff; padding: 20px; text-align: center; }}
                .header h1 {{ margin: 0; font-size: 24px; }}
                .header p {{ margin: 5px 0 0; opacity: 0.9; font-size: 13px; }}
                
                .content {{ padding: 25px; }}
                
                .greeting {{ font-size: 15px; margin-bottom: 20px; }}
                
                .section {{ margin-bottom: 20px; }}
                .section-title {{ font-size: 12px; color: #666; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 10px; border-bottom: 1px solid #eee; padding-bottom: 5px; }}
                
                .vehicle-box {{ background: #f8f9fa; border: 2px solid #1a5f4a; border-radius: 8px; padding: 15px; text-align: center; }}
                .vehicle-brand {{ color: #1a5f4a; font-size: 14px; font-weight: bold; }}
                .vehicle-model {{ font-size: 20px; font-weight: bold; color: #333; margin: 5px 0; }}
                .vehicle-price {{ font-size: 28px; font-weight: bold; color: #1a5f4a; }}
                
                .info-table {{ width: 100%; border-collapse: collapse; }}
                .info-table td {{ padding: 8px 0; border-bottom: 1px solid #eee; }}
                .info-table td:first-child {{ color: #666; }}
                .info-table td:last-child {{ text-align: right; font-weight: 600; }}
                
                .rates-table {{ width: 100%; border-collapse: collapse; background: #f8f9fa; border-radius: 6px; overflow: hidden; }}
                .rates-table th {{ background: #1a5f4a; color: #fff; padding: 10px; font-size: 12px; }}
                .rates-table td {{ padding: 8px 10px; text-align: center; border-bottom: 1px solid #e0e0e0; }}
                .rates-table tr:last-child td {{ border-bottom: none; }}
                .rates-table .selected {{ background: #d4edda; font-weight: bold; }}
                .rate-opt1 {{ color: #c0392b; font-weight: 600; }}
                .rate-opt2 {{ color: #1a5f4a; font-weight: 600; }}
                
                .best-choice {{ background: #d4edda; border: 2px solid #1a5f4a; border-radius: 8px; padding: 15px; margin: 20px 0; }}
                .best-choice-title {{ font-size: 18px; font-weight: bold; color: #155724; }}
                .best-choice-savings {{ font-size: 14px; color: #155724; margin-top: 5px; }}
                
                .options-container {{ display: table; width: 100%; border-spacing: 10px 0; }}
                .option-box {{ display: table-cell; width: 50%; vertical-align: top; border: 2px solid #ddd; border-radius: 8px; padding: 15px; background: #fafafa; }}
                .option-box.winner {{ border-color: #1a5f4a; background: #f0fff4; }}
                .option-title {{ font-size: 16px; font-weight: bold; margin-bottom: 10px; }}
                .option-title.opt1 {{ color: #c0392b; }}
                .option-title.opt2 {{ color: #1a5f4a; }}
                
                .option-detail {{ display: flex; justify-content: space-between; margin-bottom: 5px; font-size: 13px; }}
                .option-detail span:first-child {{ color: #666; }}
                
                .payment-highlight {{ background: #f0f0f0; border-radius: 6px; padding: 12px; text-align: center; margin-top: 10px; }}
                .payment-highlight.opt1 {{ background: #fdf2f2; }}
                .payment-highlight.opt2 {{ background: #f0fff4; }}
                .payment-label {{ font-size: 12px; color: #666; }}
                .payment-amount {{ font-size: 24px; font-weight: bold; margin: 5px 0; }}
                .payment-amount.opt1 {{ color: #c0392b; }}
                .payment-amount.opt2 {{ color: #1a5f4a; }}
                .payment-total {{ font-size: 12px; color: #666; }}
                .payment-total strong {{ color: #333; }}
                
                .winner-badge {{ display: inline-block; background: #1a5f4a; color: #fff; font-size: 10px; padding: 3px 8px; border-radius: 10px; margin-left: 5px; }}
                
                .footer {{ background: #f8f9fa; padding: 20px; text-align: center; border-top: 1px solid #eee; }}
                .footer .dealer {{ font-size: 16px; font-weight: bold; color: #1a5f4a; }}
                .footer .disclaimer {{ font-size: 11px; color: #999; margin-top: 10px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>CalcAuto AiPro</h1>
                    <p>Soumission de financement</p>
                </div>
                
                <div class="content">
                    <p class="greeting">Bonjour{' ' + request.client_name if request.client_name else ''},</p>
                    <p style="color: #666; margin-top: -10px;">Voici le détail de votre soumission de financement:</p>
                    
                    <!-- VEHICLE -->
                    <div class="section">
                        <div class="section-title">Véhicule sélectionné</div>
                        <div class="vehicle-box">
                            <div class="vehicle-brand">{vehicle.get('brand', '')}</div>
                            <div class="vehicle-model">{vehicle.get('model', '')} {vehicle.get('trim', '') or ''} {vehicle.get('year', '')}</div>
                            <div class="vehicle-price">{fmt(request.vehicle_price)} $</div>
                        </div>
                    </div>
                    
                    <!-- RATES TABLE -->
                    <div class="section">
                        <div class="section-title">Tableau des taux</div>
                        <table class="rates-table">
                            <thead>
                                <tr>
                                    <th style="text-align: left;">Terme</th>
                                    <th>Option 1</th>
                                    <th>Option 2</th>
                                </tr>
                            </thead>
                            <tbody>
                                <tr class="{'selected' if term == 36 else ''}"><td style="text-align: left;">36 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">0%</td></tr>
                                <tr class="{'selected' if term == 48 else ''}"><td style="text-align: left;">48 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">0%</td></tr>
                                <tr class="{'selected' if term == 60 else ''}"><td style="text-align: left;">60 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">0%</td></tr>
                                <tr class="{'selected' if term == 72 else ''}"><td style="text-align: left;">72 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">1,49%</td></tr>
                                <tr class="{'selected' if term == 84 else ''}"><td style="text-align: left;">84 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">1,99%</td></tr>
                                <tr class="{'selected' if term == 96 else ''}"><td style="text-align: left;">96 mois</td><td class="rate-opt1">4,99%</td><td class="rate-opt2">3,49%</td></tr>
                            </tbody>
                        </table>
                    </div>
                    
                    <!-- DETAILS -->
                    <div class="section">
                        <div class="section-title">Détails du financement</div>
                        <table class="info-table">
                            <tr><td>Prix du véhicule</td><td>{fmt(request.vehicle_price)} $</td></tr>
                            {f"<tr><td>Rabais (avant taxes)</td><td style='color: #1a5f4a;'>-{fmt(consumer_cash)} $</td></tr>" if consumer_cash > 0 else ""}
                            {f"<tr><td>Bonus Cash (après taxes)</td><td style='color: #1a5f4a;'>-{fmt(bonus_cash)} $</td></tr>" if bonus_cash > 0 else ""}
                            {f"<tr><td>Frais dossier</td><td>{fmt2(frais_dossier)} $</td></tr>" if frais_dossier > 0 else ""}
                            {f"<tr><td>Taxe pneus</td><td>{fmt(taxe_pneus)} $</td></tr>" if taxe_pneus > 0 else ""}
                            {f"<tr><td>Frais RDPRM</td><td>{fmt(frais_rdprm)} $</td></tr>" if frais_rdprm > 0 else ""}
                            {f"<tr><td>Valeur échange</td><td>-{fmt(valeur_echange)} $</td></tr>" if valeur_echange > 0 else ""}
                            <tr><td>Terme sélectionné</td><td><strong>{term} mois</strong></td></tr>
                            <tr><td>Fréquence de paiement</td><td><strong>{freq_label}</strong></td></tr>
                        </table>
                    </div>
                    
                    <!-- BEST CHOICE BANNER -->
                    {"<div class='best-choice'><div class='best-choice-title'>🏆 Option " + best_option + " = Meilleur choix!</div><div class='best-choice-savings'>Économies de <strong>" + fmt(savings) + " $</strong> sur le coût total</div></div>" if has_option2 and savings > 0 else ""}
                    
                    <!-- OPTIONS COMPARISON -->
                    <div class="section">
                        <div class="section-title">Comparaison des options</div>
                        <table style="width: 100%; border-spacing: 10px; border-collapse: separate;">
                            <tr>
                                <td style="width: 50%; vertical-align: top; border: 2px solid {'#1a5f4a' if best_option == '1' else '#ddd'}; border-radius: 8px; padding: 15px; background: {'#f0fff4' if best_option == '1' else '#fafafa'};">
                                    <div class="option-title opt1">Option 1 {"<span class='winner-badge'>✓ MEILLEUR</span>" if best_option == '1' else ""}</div>
                                    <div style="font-size: 12px; color: #666; margin-bottom: 10px;">Rabais + Taux standard</div>
                                    <div class="option-detail"><span>Rabais:</span><span style="color: #1a5f4a; font-weight: 600;">{'-' + fmt(consumer_cash) + ' $' if consumer_cash > 0 else '$0'}</span></div>
                                    <div class="option-detail"><span>Capital financé:</span><span>{fmt(principal_option1)} $</span></div>
                                    <div class="option-detail"><span>Taux:</span><span class="rate-opt1">{option1_rate}%</span></div>
                                    <div class="payment-highlight opt1">
                                        <div class="payment-label">{freq_label}</div>
                                        <div class="payment-amount opt1">{fmt2(option1_payment)} $</div>
                                        <div class="payment-total">Total ({term} mois): <strong>{fmt(total_option1)} $</strong></div>
                                    </div>
                                </td>
                                
                                {"<td style='width: 50%; vertical-align: top; border: 2px solid " + ("#1a5f4a" if best_option == "2" else "#ddd") + "; border-radius: 8px; padding: 15px; background: " + ("#f0fff4" if best_option == "2" else "#fafafa") + ";'><div class='option-title opt2'>Option 2 " + ("<span class='winner-badge'>✓ MEILLEUR</span>" if best_option == "2" else "") + "</div><div style='font-size: 12px; color: #666; margin-bottom: 10px;'>$0 rabais + Taux réduit</div><div class='option-detail'><span>Rabais:</span><span>$0</span></div><div class='option-detail'><span>Capital financé:</span><span>" + fmt(principal_option2) + " $</span></div><div class='option-detail'><span>Taux:</span><span class='rate-opt2'>" + str(option2_rate) + "%</span></div><div class='payment-highlight opt2'><div class='payment-label'>" + freq_label + "</div><div class='payment-amount opt2'>" + fmt2(option2_payment) + " $</div><div class='payment-total'>Total (" + str(term) + " mois): <strong>" + fmt(total_option2) + " $</strong></div></div></td>" if has_option2 else "<td style='width: 50%; vertical-align: top; border: 2px solid #ddd; border-radius: 8px; padding: 15px; background: #f5f5f5; text-align: center; color: #999;'><div class='option-title' style='color: #999;'>Option 2</div><div style='padding: 30px 0;'>Non disponible<br>pour ce véhicule</div></td>"}
                            </tr>
                        </table>
                    </div>
                    
                    {f"<div style='background: #fff3cd; border: 1px solid #ffc107; border-radius: 6px; padding: 12px; margin-top: 15px;'><span style='color: #856404;'>ℹ️ Bonus Cash de {fmt(bonus_cash)} $ sera déduit après taxes (au comptant)</span></div>" if bonus_cash > 0 else ""}
                </div>
                
                <div class="footer">
                    <div class="dealer">{request.dealer_name}</div>
                    {f"<div style='color: #666; margin-top: 5px;'>{request.dealer_phone}</div>" if request.dealer_phone else ""}
                    <div class="disclaimer">
                        Ce calcul est une estimation et ne constitue pas une offre de financement officielle.<br>
                        Les taux et conditions peuvent varier selon votre dossier de crédit.
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        subject = f"Soumission - {vehicle.get('brand', '')} {vehicle.get('model', '')} {vehicle.get('year', '')}"
        
        send_email(request.client_email, subject, html_body)
        
        return {"success": True, "message": f"Email envoyé à {request.client_email}"}
        
    except Exception as e:
        logger.error(f"Erreur envoi email: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur d'envoi: {str(e)}")

@api_router.post("/send-import-report")
async def send_import_report(request: SendReportEmailRequest):
    """Envoie un rapport par email après l'import des programmes"""
    try:
        months_fr = {
            1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
            5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
            9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
        }
        month_name = months_fr.get(request.program_month, str(request.program_month))
        
        # Générer le résumé par marque
        brands_html = ""
        for brand, count in request.brands_summary.items():
            brands_html += f"<tr><td style='padding: 8px; border-bottom: 1px solid #eee;'>{brand}</td><td style='padding: 8px; border-bottom: 1px solid #eee; text-align: center; font-weight: bold;'>{count}</td></tr>"
        
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; }}
                .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
                .header {{ background: linear-gradient(135deg, #4ECDC4 0%, #44a08d 100%); color: white; padding: 30px; text-align: center; }}
                .header h1 {{ margin: 0; font-size: 24px; }}
                .content {{ padding: 30px; }}
                .success-badge {{ background: #d4edda; color: #155724; padding: 15px; border-radius: 8px; text-align: center; margin-bottom: 20px; }}
                .stats-box {{ background: #f8f9fa; border-radius: 8px; padding: 20px; margin: 20px 0; }}
                .big-number {{ font-size: 48px; font-weight: bold; color: #1a1a2e; text-align: center; }}
                .big-label {{ text-align: center; color: #666; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
                th {{ background: #1a1a2e; color: white; padding: 10px; text-align: left; }}
                .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>✅ Import Réussi</h1>
                    <p>CalcAuto AiPro - Rapport d'import</p>
                </div>
                <div class="content">
                    <div class="success-badge">
                        <strong>🎉 Programmes {month_name} {request.program_year} importés avec succès!</strong>
                    </div>
                    
                    <div class="stats-box">
                        <div class="big-number">{request.programs_count}</div>
                        <div class="big-label">programmes de financement</div>
                    </div>
                    
                    <h3>Répartition par marque:</h3>
                    <table>
                        <tr>
                            <th>Marque</th>
                            <th style="text-align: center;">Nombre</th>
                        </tr>
                        {brands_html}
                    </table>
                    
                    <p style="margin-top: 20px; color: #666;">
                        Les nouveaux programmes sont maintenant disponibles dans l'application CalcAuto AiPro.
                    </p>
                </div>
                <div class="footer">
                    <p>Ce rapport a été généré automatiquement après l'import.</p>
                    <p>Date: {datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        subject = f"✅ Import {month_name} {request.program_year} - {request.programs_count} programmes"
        
        send_email(SMTP_EMAIL, subject, html_body)
        
        return {"success": True, "message": "Rapport envoyé par email"}
        
    except Exception as e:
        logger.error(f"Erreur envoi rapport: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur d'envoi: {str(e)}")

@api_router.post("/test-email")
async def test_email():
    """Teste la configuration email"""
    try:
        html_body = """
        <div style="font-family: Arial; padding: 20px; text-align: center;">
            <h1 style="color: #4ECDC4;">✅ Test Email Réussi!</h1>
            <p>Votre configuration Gmail SMTP fonctionne correctement.</p>
            <p style="color: #666;">CalcAuto AiPro</p>
        </div>
        """
        send_email(SMTP_EMAIL, "🧪 Test CalcAuto AiPro - Email OK", html_body)
        return {"success": True, "message": f"Email de test envoyé à {SMTP_EMAIL}"}
    except Exception as e:
        logger.error(f"Erreur test email: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# ============ CRM Endpoints ============

@api_router.post("/submissions")
async def create_submission(submission: SubmissionCreate):
    """Créer une nouvelle soumission avec rappel automatique 24h"""
    from datetime import timedelta
    
    # Set default reminder to 24h from now
    reminder_date = datetime.utcnow() + timedelta(hours=24)
    
    new_submission = Submission(
        client_name=submission.client_name,
        client_phone=submission.client_phone,
        client_email=submission.client_email,
        vehicle_brand=submission.vehicle_brand,
        vehicle_model=submission.vehicle_model,
        vehicle_year=submission.vehicle_year,
        vehicle_price=submission.vehicle_price,
        term=submission.term,
        payment_monthly=submission.payment_monthly,
        payment_biweekly=submission.payment_biweekly,
        payment_weekly=submission.payment_weekly,
        selected_option=submission.selected_option,
        rate=submission.rate,
        program_month=submission.program_month,
        program_year=submission.program_year,
        reminder_date=reminder_date
    )
    
    await db.submissions.insert_one(new_submission.dict())
    
    return {"success": True, "submission": new_submission.dict(), "message": "Soumission enregistrée - Rappel dans 24h"}

@api_router.get("/submissions")
async def get_submissions(search: Optional[str] = None, status: Optional[str] = None):
    """Récupérer toutes les soumissions avec recherche optionnelle"""
    query = {}
    
    if search:
        # Search by name or phone
        query["$or"] = [
            {"client_name": {"$regex": search, "$options": "i"}},
            {"client_phone": {"$regex": search, "$options": "i"}}
        ]
    
    if status:
        query["status"] = status
    
    submissions = await db.submissions.find(query).sort("submission_date", -1).to_list(500)
    
    # Convert ObjectId to string and datetime to ISO format
    for sub in submissions:
        if "_id" in sub:
            sub["_id"] = str(sub["_id"])
        if "submission_date" in sub and sub["submission_date"]:
            sub["submission_date"] = sub["submission_date"].isoformat()
        if "reminder_date" in sub and sub["reminder_date"]:
            sub["reminder_date"] = sub["reminder_date"].isoformat()
    
    return submissions

@api_router.get("/submissions/reminders")
async def get_reminders():
    """Récupérer les soumissions avec rappels dus ou à venir"""
    now = datetime.utcnow()
    
    # Get submissions where reminder is due and not done
    reminders_due = await db.submissions.find({
        "reminder_done": False,
        "reminder_date": {"$lte": now}
    }).sort("reminder_date", 1).to_list(100)
    
    # Get upcoming reminders (next 7 days)
    from datetime import timedelta
    next_week = now + timedelta(days=7)
    
    reminders_upcoming = await db.submissions.find({
        "reminder_done": False,
        "reminder_date": {"$gt": now, "$lte": next_week}
    }).sort("reminder_date", 1).to_list(100)
    
    # Format dates
    for sub in reminders_due + reminders_upcoming:
        if "_id" in sub:
            sub["_id"] = str(sub["_id"])
        if "submission_date" in sub and sub["submission_date"]:
            sub["submission_date"] = sub["submission_date"].isoformat()
        if "reminder_date" in sub and sub["reminder_date"]:
            sub["reminder_date"] = sub["reminder_date"].isoformat()
    
    return {
        "due": reminders_due,
        "upcoming": reminders_upcoming,
        "due_count": len(reminders_due),
        "upcoming_count": len(reminders_upcoming)
    }

@api_router.put("/submissions/{submission_id}/reminder")
async def update_reminder(submission_id: str, reminder: ReminderUpdate):
    """Mettre à jour la date de rappel"""
    update_data = {"reminder_date": reminder.reminder_date, "reminder_done": False}
    
    if reminder.notes:
        update_data["notes"] = reminder.notes
    
    result = await db.submissions.update_one(
        {"id": submission_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Soumission non trouvée")
    
    return {"success": True, "message": "Rappel mis à jour"}

@api_router.put("/submissions/{submission_id}/done")
async def mark_reminder_done(submission_id: str, new_reminder_date: Optional[str] = None):
    """Marquer un rappel comme fait, optionnellement planifier un nouveau"""
    update_data = {"reminder_done": True, "status": "contacted"}
    
    if new_reminder_date:
        update_data["reminder_date"] = datetime.fromisoformat(new_reminder_date.replace('Z', '+00:00'))
        update_data["reminder_done"] = False
    
    result = await db.submissions.update_one(
        {"id": submission_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Soumission non trouvée")
    
    return {"success": True, "message": "Rappel marqué comme fait" + (" - Nouveau rappel planifié" if new_reminder_date else "")}

@api_router.put("/submissions/{submission_id}/status")
async def update_submission_status(submission_id: str, status: str):
    """Mettre à jour le statut (pending, contacted, converted, lost)"""
    if status not in ["pending", "contacted", "converted", "lost"]:
        raise HTTPException(status_code=400, detail="Statut invalide")
    
    result = await db.submissions.update_one(
        {"id": submission_id},
        {"$set": {"status": status}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Soumission non trouvée")
    
    return {"success": True, "message": f"Statut mis à jour: {status}"}

@api_router.post("/compare-programs")
async def compare_programs_with_submissions():
    """Compare les programmes actuels avec les soumissions passées pour trouver de meilleures offres"""
    from datetime import timedelta
    
    # Get current programs
    latest = await db.programs.find_one(sort=[("program_year", -1), ("program_month", -1)])
    if not latest:
        return {"better_offers": [], "count": 0}
    
    current_month = latest["program_month"]
    current_year = latest["program_year"]
    
    # Get all pending/contacted submissions
    submissions = await db.submissions.find({
        "status": {"$in": ["pending", "contacted"]},
        "$or": [
            {"program_month": {"$ne": current_month}},
            {"program_year": {"$ne": current_year}}
        ]
    }).to_list(500)
    
    better_offers = []
    
    for sub in submissions:
        # Find matching program
        program = await db.programs.find_one({
            "brand": sub["vehicle_brand"],
            "model": sub["vehicle_model"],
            "year": sub["vehicle_year"],
            "program_month": current_month,
            "program_year": current_year
        })
        
        if not program:
            continue
        
        # Calculate new payment
        term = sub["term"]
        old_payment = sub["payment_monthly"]
        
        # Get rate for this term
        opt1_rates = program.get("option1_rates", {})
        rate_key = f"rate_{term}"
        new_rate = opt1_rates.get(rate_key, 4.99)
        
        # Calculate with new program
        vehicle_price = sub["vehicle_price"]
        consumer_cash = program.get("consumer_cash", 0)
        principal = vehicle_price - consumer_cash
        
        if new_rate == 0:
            new_payment = round(principal / term, 2)
        else:
            monthly_rate = new_rate / 100 / 12
            new_payment = round(principal * (monthly_rate * (1 + monthly_rate) ** term) / ((1 + monthly_rate) ** term - 1), 2)
        
        # Check if better
        if new_payment < old_payment:
            savings_monthly = old_payment - new_payment
            savings_total = savings_monthly * term
            
            better_offer = {
                "submission_id": sub["id"],
                "client_name": sub["client_name"],
                "client_phone": sub["client_phone"],
                "client_email": sub["client_email"],
                "vehicle": f"{sub['vehicle_brand']} {sub['vehicle_model']} {sub['vehicle_year']}",
                "old_payment": old_payment,
                "new_payment": new_payment,
                "savings_monthly": round(savings_monthly, 2),
                "savings_total": round(savings_total, 2),
                "term": term,
                "approved": False,
                "email_sent": False
            }
            better_offers.append(better_offer)
    
    # Store better offers in DB for approval
    if better_offers:
        await db.better_offers.delete_many({})  # Clear old offers
        await db.better_offers.insert_many(better_offers)
        
        # Send notification email to admin
        try:
            send_better_offers_notification(better_offers)
        except Exception as e:
            logger.error(f"Error sending better offers notification: {e}")
    
    return {"better_offers": better_offers, "count": len(better_offers)}

def send_better_offers_notification(offers: List[dict]):
    """Envoie une notification par email des meilleures offres disponibles"""
    if not SMTP_EMAIL:
        return
    
    offers_html = ""
    for offer in offers:
        offers_html += f"""
        <tr>
            <td style="padding: 12px; border-bottom: 1px solid #eee;">{offer['client_name']}</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee;">{offer['vehicle']}</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right;">{offer['old_payment']:.2f}$</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right; color: #28a745; font-weight: bold;">{offer['new_payment']:.2f}$</td>
            <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right; color: #28a745;">-{offer['savings_monthly']:.2f}$/mois</td>
        </tr>
        """
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head><meta charset="utf-8"></head>
    <body style="font-family: Arial, sans-serif; padding: 20px; background: #f5f5f5;">
        <div style="max-width: 700px; margin: 0 auto; background: white; border-radius: 10px; overflow: hidden;">
            <div style="background: #1a1a2e; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0;">🔔 Meilleures Offres Disponibles</h1>
            </div>
            <div style="padding: 20px;">
                <p style="font-size: 16px;">{len(offers)} client(s) peuvent bénéficier de meilleurs taux!</p>
                
                <table style="width: 100%; border-collapse: collapse; margin: 20px 0;">
                    <tr style="background: #f8f9fa;">
                        <th style="padding: 12px; text-align: left;">Client</th>
                        <th style="padding: 12px; text-align: left;">Véhicule</th>
                        <th style="padding: 12px; text-align: right;">Ancien</th>
                        <th style="padding: 12px; text-align: right;">Nouveau</th>
                        <th style="padding: 12px; text-align: right;">Économie</th>
                    </tr>
                    {offers_html}
                </table>
                
                <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0;">
                    <strong>⚠️ Action requise:</strong><br>
                    Ouvrez l'application pour approuver l'envoi des emails aux clients.
                </div>
                
                <a href="https://deal-detail-modal.preview.emergentagent.com" style="display: inline-block; background: #4ECDC4; color: #1a1a2e; padding: 12px 24px; border-radius: 8px; text-decoration: none; font-weight: bold;">
                    Ouvrir l'application
                </a>
            </div>
        </div>
    </body>
    </html>
    """
    
    send_email(SMTP_EMAIL, f"🔔 CalcAuto - {len(offers)} client(s) à relancer!", html_body)

@api_router.get("/better-offers")
async def get_better_offers():
    """Récupérer les meilleures offres en attente d'approbation"""
    offers = await db.better_offers.find({"approved": False, "email_sent": False}).to_list(100)
    
    for offer in offers:
        if "_id" in offer:
            offer["_id"] = str(offer["_id"])
    
    return offers

@api_router.post("/better-offers/{submission_id}/approve")
async def approve_better_offer(submission_id: str):
    """Approuver et envoyer l'email au client pour une meilleure offre"""
    offer = await db.better_offers.find_one({"submission_id": submission_id})
    
    if not offer:
        raise HTTPException(status_code=404, detail="Offre non trouvée")
    
    if offer.get("email_sent"):
        return {"success": False, "message": "Email déjà envoyé"}
    
    # Send email to client
    try:
        send_client_better_offer_email(offer)
        
        # Mark as sent
        await db.better_offers.update_one(
            {"submission_id": submission_id},
            {"$set": {"approved": True, "email_sent": True}}
        )
        
        return {"success": True, "message": f"Email envoyé à {offer['client_email']}"}
    except Exception as e:
        logger.error(f"Error sending client email: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur d'envoi: {str(e)}")

def send_client_better_offer_email(offer: dict):
    """Envoie un email au client pour l'informer d'une meilleure offre"""
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head><meta charset="utf-8"></head>
    <body style="font-family: Arial, sans-serif; padding: 20px; background: #f5f5f5;">
        <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 10px; overflow: hidden;">
            <div style="background: #1a1a2e; color: white; padding: 20px; text-align: center;">
                <h1 style="margin: 0; color: #4ECDC4;">🎉 Bonne Nouvelle!</h1>
            </div>
            <div style="padding: 30px;">
                <p style="font-size: 18px;">Bonjour {offer['client_name']},</p>
                
                <p>De nouveaux programmes de financement sont disponibles et vous permettraient d'<strong>économiser</strong> sur votre {offer['vehicle']}!</p>
                
                <div style="background: #d4edda; border-radius: 10px; padding: 20px; margin: 20px 0; text-align: center;">
                    <p style="margin: 0; color: #666;">Votre paiement actuel</p>
                    <p style="font-size: 24px; margin: 5px 0; text-decoration: line-through; color: #999;">{offer['old_payment']:.2f}$/mois</p>
                    
                    <p style="margin: 15px 0 0; color: #666;">Nouveau paiement possible</p>
                    <p style="font-size: 32px; margin: 5px 0; color: #28a745; font-weight: bold;">{offer['new_payment']:.2f}$/mois</p>
                    
                    <p style="font-size: 18px; color: #28a745; margin-top: 15px;">
                        💰 Économie: {offer['savings_monthly']:.2f}$/mois ({offer['savings_total']:.2f}$ sur {offer['term']} mois)
                    </p>
                </div>
                
                <p>Contactez-nous pour profiter de cette offre!</p>
                
                <p style="color: #666; font-size: 14px; margin-top: 30px;">
                    Cordialement,<br>
                    L'équipe CalcAuto AiPro
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    send_email(offer['client_email'], f"🎉 Économisez {offer['savings_monthly']:.2f}$/mois sur votre {offer['vehicle']}!", html_body)

@api_router.post("/better-offers/{submission_id}/ignore")
async def ignore_better_offer(submission_id: str):
    """Ignorer une meilleure offre"""
    result = await db.better_offers.delete_one({"submission_id": submission_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Offre non trouvée")
    
    return {"success": True, "message": "Offre ignorée"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
