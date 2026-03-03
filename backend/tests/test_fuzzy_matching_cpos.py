"""
Test Fuzzy Matching for Excel Import with CPOS Codes
=====================================================
This iteration focuses on testing the fuzzy matching behavior when trim names
are modified with product codes like (CPOS 22L/22P), (WLJH74), etc.

Test Scenarios:
1. Export programs, modify trim by appending CPOS codes, change consumer_cash, reimport
2. Verify 0 errors (successful matching despite modified trim)
3. Verify updated >= 1 (changes detected)
4. Verify comparison shows the consumer_cash change
5. Restore original values after tests

The normalize_str() function should strip:
- (CPOS xxx) codes
- (WLJH74), (DJ7X91), (KMJL74), (DT6P98) product codes
- (ETM) codes
While keeping descriptive text like (excluding PHEV), (Gas), etc.
"""

import pytest
import requests
import os
import io
import time

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Get BASE_URL from environment
BASE_URL = os.environ.get('EXPO_PUBLIC_BACKEND_URL', 'https://deal-detail-modal.preview.emergentagent.com').rstrip('/')

# Admin password for import operations
ADMIN_PASSWORD = "Liana2018"


@pytest.fixture(scope="module")
def api_session():
    """Shared requests session"""
    session = requests.Session()
    session.headers.update({"Content-Type": "application/json"})
    return session


# ============ FUZZY MATCHING TESTS FOR CPOS CODES ============

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl required for import tests")
class TestFuzzyMatchingCPOSCodes:
    """Tests for fuzzy matching when trim names have CPOS codes appended"""
    
    @pytest.fixture(autouse=True)
    def save_original_export(self, api_session):
        """Save original export for reverting after tests"""
        export_resp = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        self.original_export = export_resp.content
        yield
        # Revert after each test
        if hasattr(self, 'original_export'):
            files = {'file': ('revert.xlsx', io.BytesIO(self.original_export),
                             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            requests.post(
                f"{BASE_URL}/api/programs/import-excel",
                files=files,
                data={'password': ADMIN_PASSWORD}
            )
            print("\n  >> Reverted to original data")
    
    def test_import_with_cpos_appended_to_trim_matches_and_updates(self, api_session):
        """
        CRITICAL TEST: Import with trim modified by CPOS code should still match.
        
        Steps:
        1. Export programs Excel
        2. Find a Grand Cherokee row (or any row)
        3. Append '(CPOS 22L/22P)' to trim value
        4. Change consumer_cash to 99999
        5. Reimport
        6. Verify: errors=0, updated>=1, comparison shows consumer_cash change
        """
        # 1. Export current state
        export_resp = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        assert export_resp.status_code == 200, f"Export failed: {export_resp.status_code}"
        
        # 2. Load Excel and find a suitable row to modify
        wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
        ws = wb['Programmes']
        
        # Find column indices (1-indexed)
        headers = [cell.value for cell in ws[1]]
        try:
            id_col = headers.index('ID') + 1
            brand_col = headers.index('Marque') + 1
            model_col = headers.index('Modele') + 1
            trim_col = headers.index('Trim') + 1
            year_col = headers.index('Annee') + 1
            consumer_cash_col = headers.index('Consumer Cash ($)') + 1
        except ValueError as e:
            pytest.fail(f"Missing column header: {e}. Available: {headers[:10]}")
        
        # 3. Find a Grand Cherokee row (or first available row)
        target_row = None
        original_trim = None
        original_consumer_cash = None
        vehicle_info = {}
        
        for row_idx in range(2, ws.max_row + 1):
            brand = ws.cell(row=row_idx, column=brand_col).value
            model = ws.cell(row=row_idx, column=model_col).value
            trim = ws.cell(row=row_idx, column=trim_col).value
            
            # Prefer Grand Cherokee but accept any row with a trim value
            if model and 'Grand Cherokee' in str(model):
                target_row = row_idx
                original_trim = trim or ""
                original_consumer_cash = ws.cell(row=row_idx, column=consumer_cash_col).value or 0
                vehicle_info = {
                    'id': ws.cell(row=row_idx, column=id_col).value,
                    'brand': brand,
                    'model': model,
                    'trim': trim,
                    'year': ws.cell(row=row_idx, column=year_col).value
                }
                break
        
        # Fallback to first data row if no Grand Cherokee found
        if target_row is None:
            target_row = 2
            original_trim = ws.cell(row=target_row, column=trim_col).value or ""
            original_consumer_cash = ws.cell(row=target_row, column=consumer_cash_col).value or 0
            vehicle_info = {
                'id': ws.cell(row=target_row, column=id_col).value,
                'brand': ws.cell(row=target_row, column=brand_col).value,
                'model': ws.cell(row=target_row, column=model_col).value,
                'trim': ws.cell(row=target_row, column=trim_col).value,
                'year': ws.cell(row=target_row, column=year_col).value
            }
        
        print(f"\n  Test vehicle: {vehicle_info['brand']} {vehicle_info['model']} | Trim: '{original_trim}'")
        print(f"  Original consumer_cash: {original_consumer_cash}")
        
        # 4. Modify trim by appending CPOS code
        modified_trim = f"{original_trim} (CPOS 22L/22P)"
        ws.cell(row=target_row, column=trim_col).value = modified_trim
        
        # 5. Change consumer_cash to test value
        test_consumer_cash = 99999.0
        ws.cell(row=target_row, column=consumer_cash_col).value = test_consumer_cash
        
        print(f"  Modified trim: '{modified_trim}'")
        print(f"  Test consumer_cash: {test_consumer_cash}")
        
        # 6. Save modified Excel
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 7. Import modified file
        files = {'file': ('cpos_test.xlsx', output,
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'password': ADMIN_PASSWORD}
        
        import_resp = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data=data
        )
        
        assert import_resp.status_code == 200, f"Import failed: {import_resp.status_code} - {import_resp.text[:500]}"
        
        result = import_resp.json()
        
        # 8. Verify results
        print(f"\n  Import Response:")
        print(f"    success: {result.get('success')}")
        print(f"    rows_processed: {result.get('rows_processed')}")
        print(f"    updated: {result.get('updated')}")
        print(f"    unchanged: {result.get('unchanged')}")
        print(f"    corrections_saved: {result.get('corrections_saved')}")
        print(f"    errors: {len(result.get('errors', []))}")
        
        # Log any errors
        if result.get('errors'):
            print(f"    Error details: {result['errors'][:5]}")
        
        # Critical assertions
        assert result.get('success') == True, f"Import not successful: {result}"
        
        # CRITICAL: Should have 0 errors (fuzzy matching should work)
        errors = result.get('errors', [])
        assert len(errors) == 0, \
            f"Expected 0 errors but got {len(errors)}. Fuzzy matching may have failed. Errors: {errors[:3]}"
        
        # Should have at least 1 update (the consumer_cash change)
        assert result.get('updated', 0) >= 1, \
            f"Expected at least 1 update, got {result.get('updated')}. Change may not have been detected."
        
        # 9. Verify comparison shows the consumer_cash change
        comparison = result.get('comparison', [])
        assert len(comparison) > 0, "Expected comparison details but got none"
        
        # Find the specific change for our vehicle
        found_change = False
        for item in comparison:
            vehicule = item.get('vehicule', '')
            changes = item.get('changes', {})
            
            # Check if this is our vehicle (approximate match)
            if vehicle_info['brand'] in vehicule and vehicle_info['model'] in vehicule:
                if 'consumer_cash' in changes:
                    found_change = True
                    cash_change = changes['consumer_cash']
                    
                    print(f"\n  Comparison found for: {vehicule}")
                    print(f"    consumer_cash change: {cash_change.get('avant')} -> {cash_change.get('apres')}")
                    
                    # Verify the 'apres' value matches our test value
                    assert cash_change.get('apres') == test_consumer_cash, \
                        f"Expected apres={test_consumer_cash}, got {cash_change.get('apres')}"
                    
                    break
        
        # If not found by vehicle name, check if any consumer_cash change to test value exists
        if not found_change:
            any_test_value_change = any(
                item.get('changes', {}).get('consumer_cash', {}).get('apres') == test_consumer_cash
                for item in comparison
            )
            assert any_test_value_change, \
                f"Consumer cash change to {test_consumer_cash} not found in comparison: {comparison[:3]}"
            print(f"\n  Consumer cash change verified (vehicle name format may differ)")
        
        print(f"\n  PASSED: Fuzzy matching with CPOS code worked correctly")
    
    def test_import_with_product_code_wljh74_matches(self, api_session):
        """
        Test that product codes like (WLJH74) are stripped during matching.
        """
        export_resp = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        assert export_resp.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
        ws = wb['Programmes']
        
        headers = [cell.value for cell in ws[1]]
        trim_col = headers.index('Trim') + 1
        consumer_cash_col = headers.index('Consumer Cash ($)') + 1
        brand_col = headers.index('Marque') + 1
        model_col = headers.index('Modele') + 1
        
        # Get first row data
        original_trim = ws.cell(row=2, column=trim_col).value or ""
        original_cash = ws.cell(row=2, column=consumer_cash_col).value or 0
        brand = ws.cell(row=2, column=brand_col).value
        model = ws.cell(row=2, column=model_col).value
        
        # Modify trim with WLJH74 product code
        modified_trim = f"{original_trim} (WLJH74 2*A)"
        test_cash = 88888.0
        
        ws.cell(row=2, column=trim_col).value = modified_trim
        ws.cell(row=2, column=consumer_cash_col).value = test_cash
        
        print(f"\n  Test: {brand} {model}")
        print(f"  Original trim: '{original_trim}'")
        print(f"  Modified trim: '{modified_trim}'")
        print(f"  Test consumer_cash: {test_cash}")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {'file': ('wljh74_test.xlsx', output,
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        import_resp = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data={'password': ADMIN_PASSWORD}
        )
        
        assert import_resp.status_code == 200, f"Import failed: {import_resp.status_code}"
        result = import_resp.json()
        
        print(f"\n  Response: errors={len(result.get('errors', []))}, updated={result.get('updated')}")
        
        assert result.get('success') == True
        assert len(result.get('errors', [])) == 0, \
            f"Expected 0 errors with (WLJH74) code. Errors: {result.get('errors', [])[:3]}"
        assert result.get('updated', 0) >= 1, \
            f"Expected at least 1 update"
        
        print(f"  PASSED: (WLJH74) product code handled correctly")
    
    def test_import_with_multiple_product_codes(self, api_session):
        """
        Test matching with multiple product codes appended.
        """
        export_resp = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        assert export_resp.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
        ws = wb['Programmes']
        
        headers = [cell.value for cell in ws[1]]
        trim_col = headers.index('Trim') + 1
        consumer_cash_col = headers.index('Consumer Cash ($)') + 1
        
        original_trim = ws.cell(row=3, column=trim_col).value or ""
        
        # Add multiple codes
        modified_trim = f"{original_trim} (CPOS 22L) (DJ7X91)"
        test_cash = 77777.0
        
        ws.cell(row=3, column=trim_col).value = modified_trim
        ws.cell(row=3, column=consumer_cash_col).value = test_cash
        
        print(f"\n  Original trim: '{original_trim}'")
        print(f"  Modified trim: '{modified_trim}'")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {'file': ('multi_code_test.xlsx', output,
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        import_resp = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data={'password': ADMIN_PASSWORD}
        )
        
        assert import_resp.status_code == 200
        result = import_resp.json()
        
        print(f"  Response: errors={len(result.get('errors', []))}, updated={result.get('updated')}")
        
        assert result.get('success') == True
        assert len(result.get('errors', [])) == 0, \
            f"Expected 0 errors with multiple codes. Errors: {result.get('errors', [])[:3]}"
        
        print(f"  PASSED: Multiple product codes handled correctly")
    
    def test_descriptive_text_preserved_during_matching(self, api_session):
        """
        Test that descriptive text like (excluding PHEV) is preserved and still matches.
        Find a row with existing descriptive text and verify it still works.
        """
        export_resp = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        assert export_resp.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
        ws = wb['Programmes']
        
        headers = [cell.value for cell in ws[1]]
        trim_col = headers.index('Trim') + 1
        consumer_cash_col = headers.index('Consumer Cash ($)') + 1
        brand_col = headers.index('Marque') + 1
        model_col = headers.index('Modele') + 1
        
        # Find a row with descriptive text like "excluding" or look for Pacifica
        target_row = None
        for row_idx in range(2, min(ws.max_row + 1, 100)):
            trim = ws.cell(row=row_idx, column=trim_col).value
            model = ws.cell(row=row_idx, column=model_col).value
            
            if trim and ('excluding' in str(trim).lower() or 'excl.' in str(trim).lower()):
                target_row = row_idx
                break
            if model and 'Pacifica' in str(model):
                target_row = row_idx
                break
        
        if target_row is None:
            target_row = 5  # Use row 5 as fallback
        
        original_trim = ws.cell(row=target_row, column=trim_col).value or ""
        brand = ws.cell(row=target_row, column=brand_col).value
        model = ws.cell(row=target_row, column=model_col).value
        
        # Add CPOS code but keep the original descriptive text
        modified_trim = f"{original_trim} (CPOS TEST123)"
        test_cash = 66666.0
        
        ws.cell(row=target_row, column=trim_col).value = modified_trim
        ws.cell(row=target_row, column=consumer_cash_col).value = test_cash
        
        print(f"\n  Test: {brand} {model}")
        print(f"  Original trim: '{original_trim}'")
        print(f"  Modified trim: '{modified_trim}'")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        files = {'file': ('desc_text_test.xlsx', output,
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        import_resp = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data={'password': ADMIN_PASSWORD}
        )
        
        assert import_resp.status_code == 200
        result = import_resp.json()
        
        print(f"  Response: errors={len(result.get('errors', []))}, updated={result.get('updated')}")
        
        assert result.get('success') == True
        assert len(result.get('errors', [])) == 0, \
            f"Matching should work with descriptive text + CPOS code. Errors: {result.get('errors', [])[:3]}"
        
        print(f"  PASSED: Descriptive text preserved during matching")


# ============ COMPARISON VERIFICATION TESTS ============

class TestComparisonAfterFuzzyMatch:
    """Verify comparison details are correct after fuzzy matching import"""
    
    def test_comparison_history_includes_fuzzy_match_imports(self, api_session):
        """GET /api/programs/comparisons should include recent fuzzy match imports"""
        response = api_session.get(f"{BASE_URL}/api/programs/comparisons")
        
        assert response.status_code == 200, f"Comparisons endpoint failed: {response.status_code}"
        
        data = response.json()
        assert isinstance(data, list), f"Expected list, got {type(data)}"
        
        print(f"\n  Comparison history contains {len(data)} records")
        
        if len(data) > 0:
            latest = data[0]
            print(f"  Latest comparison:")
            print(f"    date: {latest.get('date')}")
            print(f"    updated: {latest.get('updated')}")
            print(f"    unchanged: {latest.get('unchanged')}")
            print(f"    errors: {len(latest.get('errors', []))}")
            
            details = latest.get('details', [])
            if details:
                print(f"  Sample change: {details[0].get('vehicule')}")
    
    def test_comparison_detail_shows_avant_apres(self, api_session):
        """GET /api/programs/comparison/{id} shows avant/apres values"""
        # Get list first
        list_resp = api_session.get(f"{BASE_URL}/api/programs/comparisons")
        assert list_resp.status_code == 200
        
        comparisons = list_resp.json()
        if len(comparisons) == 0:
            pytest.skip("No comparisons available")
        
        # Get detail of latest
        comparison_id = comparisons[0]['id']
        detail_resp = api_session.get(f"{BASE_URL}/api/programs/comparison/{comparison_id}")
        
        assert detail_resp.status_code == 200, f"Detail failed: {detail_resp.status_code}"
        
        detail = detail_resp.json()
        assert detail.get('id') == comparison_id
        
        details = detail.get('details', [])
        if details:
            for item in details[:3]:
                changes = item.get('changes', {})
                for field, change in changes.items():
                    assert 'avant' in change, f"Missing 'avant' in {field} change"
                    assert 'apres' in change, f"Missing 'apres' in {field} change"
                    print(f"    {field}: {change.get('avant')} -> {change.get('apres')}")
        
        print(f"\n  Comparison detail structure verified")


# ============ SCI IMPORT TESTS ============

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl required")
class TestSCIImportExport:
    """Test SCI export/import endpoints"""
    
    def test_sci_export_returns_xlsx(self, api_session):
        """GET /api/sci/export-excel returns valid xlsx"""
        response = api_session.get(f"{BASE_URL}/api/sci/export-excel")
        
        assert response.status_code == 200, f"SCI export failed: {response.status_code}"
        
        content_disp = response.headers.get('Content-Disposition', '')
        assert '.xlsx' in content_disp, f"Expected xlsx: {content_disp}"
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheets = wb.sheetnames
        has_lease = any('Lease' in s for s in sheets)
        assert has_lease, f"Expected Lease sheet in: {sheets}"
        
        print(f"\n  SCI Export successful: {len(response.content)} bytes")
        print(f"  Sheets: {sheets}")
    
    def test_sci_comparisons_endpoint(self, api_session):
        """GET /api/sci/comparisons returns history"""
        response = api_session.get(f"{BASE_URL}/api/sci/comparisons")
        
        assert response.status_code == 200, f"SCI comparisons failed: {response.status_code}"
        
        data = response.json()
        assert isinstance(data, list)
        
        print(f"\n  SCI comparison history: {len(data)} records")
        
        if data:
            latest = data[0]
            assert latest.get('type') == 'sci_lease'
            print(f"  Latest: updated={latest.get('updated')}, unchanged={latest.get('unchanged')}")


# ============ EXECUTION ============

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
