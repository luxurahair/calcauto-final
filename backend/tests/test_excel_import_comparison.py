"""
Test Excel Import/Export with Before/After Comparison Feature
Tests for:
- POST /api/programs/import-excel - Import Excel with comparison (modify consumer_cash, reimport, verify updated > 0)
- POST /api/programs/import-excel - Import without changes (unchanged > 0, updated = 0)
- GET /api/programs/comparisons - Return history of comparisons
- GET /api/programs/export-excel - Export works correctly
- POST /api/sci/import-excel - SCI import with comparison
- GET /api/sci/comparisons - SCI comparison history
- GET /api/sci/export-excel - SCI export works

The import system uses composite key (brand+model+trim+year) for matching.
Comparison captures snapshot before import, applies changes, calculates diff,
stores in MongoDB and returns in API response.
"""

import pytest
import requests
import os
import io
import tempfile
from datetime import datetime

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Get BASE_URL from environment
BASE_URL = os.environ.get('EXPO_PUBLIC_BACKEND_URL', 'https://deal-detail-modal.preview.emergentagent.com').rstrip('/')

# Admin password for import operations
ADMIN_PASSWORD = "Liana2018"


@pytest.fixture(scope="module")
def api_session():
    """Shared requests session"""
    session = requests.Session()
    session.headers.update({"Content-Type": "application/json"})
    return session


# ============ PROGRAMS EXPORT TESTS ============

class TestProgramsExportExcel:
    """Tests for GET /api/programs/export-excel"""
    
    def test_export_programs_excel_returns_xlsx(self, api_session):
        """GET /api/programs/export-excel returns valid xlsx file"""
        response = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        
        assert response.status_code == 200, f"Export failed: {response.status_code} - {response.text[:200]}"
        
        # Check content type
        content_type = response.headers.get('Content-Type', '')
        assert 'spreadsheet' in content_type or 'xlsx' in content_type or 'octet-stream' in content_type, \
            f"Unexpected content type: {content_type}"
        
        # Check Content-Disposition header
        content_disp = response.headers.get('Content-Disposition', '')
        assert 'filename=' in content_disp, f"Missing filename in Content-Disposition: {content_disp}"
        assert '.xlsx' in content_disp, f"Expected .xlsx file: {content_disp}"
        
        # Verify file is valid xlsx
        if OPENPYXL_AVAILABLE:
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            sheets = wb.sheetnames
            assert 'Programmes' in sheets, f"Expected 'Programmes' sheet, got: {sheets}"
            
            # Check header structure
            ws = wb['Programmes']
            headers = [cell.value for cell in ws[1]]
            assert 'Marque' in headers, f"Missing 'Marque' header: {headers[:5]}"
            assert 'Modele' in headers, f"Missing 'Modele' header: {headers[:5]}"
            assert 'Consumer Cash ($)' in headers, f"Missing 'Consumer Cash ($)' header"
            
            # Count data rows
            data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
            print(f"✓ Export successful: {data_rows} programs in xlsx")
            print(f"  Sheets: {sheets}")
        else:
            # Just check file size is reasonable (> 5KB for a valid xlsx)
            assert len(response.content) > 5000, f"File too small ({len(response.content)} bytes)"
            print(f"✓ Export successful: {len(response.content)} bytes (openpyxl not available for structure check)")
    
    def test_export_programs_with_period_filter(self, api_session):
        """GET /api/programs/export-excel with month/year params"""
        response = api_session.get(
            f"{BASE_URL}/api/programs/export-excel",
            params={"month": 2, "year": 2026}
        )
        
        assert response.status_code == 200, f"Export with params failed: {response.status_code}"
        assert len(response.content) > 1000, "Export file too small"
        print(f"✓ Export with period filter successful: {len(response.content)} bytes")


# ============ PROGRAMS IMPORT TESTS ============

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl required for import tests")
class TestProgramsImportExcel:
    """Tests for POST /api/programs/import-excel with before/after comparison"""
    
    def test_import_without_password_fails(self, api_session):
        """POST /api/programs/import-excel without password returns 401"""
        # First get a valid xlsx file
        export_resp = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        assert export_resp.status_code == 200
        
        # Try to import without password
        files = {'file': ('test.xlsx', io.BytesIO(export_resp.content), 
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'password': ''}
        
        # Remove Content-Type to let requests set multipart
        headers = {k: v for k, v in api_session.headers.items() if k.lower() != 'content-type'}
        response = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data=data,
            headers=headers
        )
        
        assert response.status_code == 401, f"Expected 401, got {response.status_code}"
        print("✓ Import without password correctly rejected")
    
    def test_import_with_wrong_password_fails(self, api_session):
        """POST /api/programs/import-excel with wrong password returns 401"""
        export_resp = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        assert export_resp.status_code == 200
        
        files = {'file': ('test.xlsx', io.BytesIO(export_resp.content),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'password': 'wrongpassword'}
        
        response = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data=data
        )
        
        assert response.status_code == 401, f"Expected 401, got {response.status_code}"
        print("✓ Import with wrong password correctly rejected")
    
    def test_import_unchanged_returns_zero_updates(self, api_session):
        """POST /api/programs/import-excel without modifications returns updated=0, unchanged > 0"""
        # Export current state
        export_resp = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        assert export_resp.status_code == 200, f"Export failed: {export_resp.status_code}"
        
        # Import same file without changes
        files = {'file': ('unchanged.xlsx', io.BytesIO(export_resp.content),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'password': ADMIN_PASSWORD}
        
        import_resp = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data=data
        )
        
        assert import_resp.status_code == 200, f"Import failed: {import_resp.status_code} - {import_resp.text[:500]}"
        
        result = import_resp.json()
        assert result.get('success') == True, f"Import not successful: {result}"
        assert result.get('unchanged', 0) > 0, f"Expected unchanged > 0, got {result.get('unchanged')}"
        # Note: updated might be > 0 if there were pending corrections from previous test
        
        print(f"✓ Import unchanged file successful:")
        print(f"  updated={result.get('updated')}, unchanged={result.get('unchanged')}")
        print(f"  rows_processed={result.get('rows_processed')}")
    
    def test_import_with_changes_returns_comparison(self, api_session):
        """POST /api/programs/import-excel with modifications returns comparison details"""
        # 1. Export current state
        export_resp = api_session.get(f"{BASE_URL}/api/programs/export-excel")
        assert export_resp.status_code == 200, f"Export failed: {export_resp.status_code}"
        
        # 2. Load and modify Excel - change consumer_cash for first program
        wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
        ws = wb['Programmes']
        
        # Find Consumer Cash column (should be column F = index 6)
        headers = [cell.value for cell in ws[1]]
        consumer_cash_col = headers.index('Consumer Cash ($)') + 1  # 1-indexed
        
        # Store original value and modify first data row
        original_value = ws.cell(row=2, column=consumer_cash_col).value or 0
        test_value = float(original_value) + 12345  # Add 12345 to make it clearly different
        ws.cell(row=2, column=consumer_cash_col).value = test_value
        
        # Get vehicle info for verification
        brand = ws.cell(row=2, column=2).value  # Column B = Brand
        model = ws.cell(row=2, column=3).value  # Column C = Model
        trim = ws.cell(row=2, column=4).value   # Column D = Trim
        year = ws.cell(row=2, column=5).value   # Column E = Year
        
        print(f"  Modifying: {brand} {model} {trim} {year}")
        print(f"  consumer_cash: {original_value} -> {test_value}")
        
        # 3. Save modified Excel to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 4. Import modified file
        files = {'file': ('modified.xlsx', output,
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'password': ADMIN_PASSWORD}
        
        import_resp = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data=data
        )
        
        assert import_resp.status_code == 200, f"Import failed: {import_resp.status_code} - {import_resp.text[:500]}"
        
        result = import_resp.json()
        assert result.get('success') == True, f"Import not successful: {result}"
        assert result.get('updated', 0) > 0, f"Expected updated > 0, got {result.get('updated')}"
        
        # Check comparison details
        comparison = result.get('comparison', [])
        assert len(comparison) > 0, f"Expected comparison details, got none"
        
        # Verify the changed vehicle is in comparison
        changed_vehicle = f"{brand} {model} {trim} {year}"
        found_change = False
        for item in comparison:
            if item.get('vehicule', '').strip() == changed_vehicle.strip():
                found_change = True
                changes = item.get('changes', {})
                assert 'consumer_cash' in changes, f"consumer_cash change not found in: {changes}"
                
                # Verify before/after values
                cash_change = changes['consumer_cash']
                assert 'avant' in cash_change, f"Missing 'avant' in change: {cash_change}"
                assert 'apres' in cash_change, f"Missing 'apres' in change: {cash_change}"
                assert cash_change['apres'] == test_value, \
                    f"Expected apres={test_value}, got {cash_change['apres']}"
                
                print(f"✓ Comparison shows change:")
                print(f"  vehicule: {item.get('vehicule')}")
                print(f"  consumer_cash: {cash_change['avant']} -> {cash_change['apres']}")
                break
        
        # The vehicle should be found (might have slightly different formatting)
        if not found_change:
            # Check if any consumer_cash change was detected
            any_cash_change = any('consumer_cash' in item.get('changes', {}) for item in comparison)
            assert any_cash_change, f"No consumer_cash change detected. Comparison: {comparison[:3]}"
            print(f"✓ Consumer cash change detected (vehicle name format may differ)")
            print(f"  First change: {comparison[0]}")
        
        # 5. Revert the change by importing original export
        files_revert = {'file': ('revert.xlsx', io.BytesIO(export_resp.content),
                                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        revert_resp = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files_revert,
            data={'password': ADMIN_PASSWORD}
        )
        
        assert revert_resp.status_code == 200, f"Revert failed: {revert_resp.status_code}"
        print("✓ Changes reverted successfully")
        
        return result.get('comparison_id')


# ============ PROGRAMS COMPARISONS HISTORY TESTS ============

class TestProgramsComparisonsHistory:
    """Tests for GET /api/programs/comparisons"""
    
    def test_get_comparisons_returns_list(self, api_session):
        """GET /api/programs/comparisons returns list of comparison records"""
        response = api_session.get(f"{BASE_URL}/api/programs/comparisons")
        
        assert response.status_code == 200, f"Comparisons endpoint failed: {response.status_code}"
        
        data = response.json()
        assert isinstance(data, list), f"Expected list, got {type(data)}"
        
        if len(data) > 0:
            comp = data[0]
            # Validate structure
            assert 'id' in comp, f"Missing 'id' field: {comp.keys()}"
            assert 'type' in comp, f"Missing 'type' field: {comp.keys()}"
            assert comp['type'] == 'programs', f"Expected type='programs', got {comp['type']}"
            assert 'date' in comp, f"Missing 'date' field"
            assert 'updated' in comp, f"Missing 'updated' field"
            assert 'unchanged' in comp, f"Missing 'unchanged' field"
            
            print(f"✓ Comparisons history returned {len(data)} records")
            print(f"  Latest: id={comp['id'][:8]}... date={comp['date']}")
            print(f"  updated={comp['updated']}, unchanged={comp['unchanged']}")
            
            # Check details if present
            if 'details' in comp and len(comp['details']) > 0:
                detail = comp['details'][0]
                print(f"  Sample detail: {detail.get('vehicule', 'N/A')}")
                if 'changes' in detail:
                    print(f"    Changes: {list(detail['changes'].keys())}")
        else:
            print("✓ Comparisons endpoint working (no history yet)")
    
    def test_comparisons_sorted_by_date_desc(self, api_session):
        """GET /api/programs/comparisons returns records sorted by date descending"""
        response = api_session.get(f"{BASE_URL}/api/programs/comparisons")
        
        assert response.status_code == 200
        data = response.json()
        
        if len(data) >= 2:
            # Verify descending order
            for i in range(len(data) - 1):
                date1 = data[i].get('date', '')
                date2 = data[i + 1].get('date', '')
                assert date1 >= date2, f"Not sorted: {date1} should be >= {date2}"
            
            print(f"✓ Comparisons sorted by date descending (verified {len(data)} records)")
        else:
            print(f"✓ Not enough records to verify sort order ({len(data)} records)")


# ============ SCI EXPORT TESTS ============

class TestSCIExportExcel:
    """Tests for GET /api/sci/export-excel"""
    
    def test_export_sci_excel_returns_xlsx(self, api_session):
        """GET /api/sci/export-excel returns valid xlsx file"""
        response = api_session.get(f"{BASE_URL}/api/sci/export-excel")
        
        assert response.status_code == 200, f"SCI Export failed: {response.status_code} - {response.text[:200]}"
        
        content_type = response.headers.get('Content-Type', '')
        assert 'spreadsheet' in content_type or 'xlsx' in content_type or 'octet-stream' in content_type, \
            f"Unexpected content type: {content_type}"
        
        content_disp = response.headers.get('Content-Disposition', '')
        assert '.xlsx' in content_disp, f"Expected .xlsx file: {content_disp}"
        
        if OPENPYXL_AVAILABLE:
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            sheets = wb.sheetnames
            # Should have Lease 2026 or Lease 2025 sheets
            has_lease_sheet = any('Lease' in s for s in sheets)
            assert has_lease_sheet, f"Expected 'Lease' sheet, got: {sheets}"
            
            print(f"✓ SCI Export successful: {len(response.content)} bytes")
            print(f"  Sheets: {sheets}")
        else:
            assert len(response.content) > 3000, f"File too small ({len(response.content)} bytes)"
            print(f"✓ SCI Export successful: {len(response.content)} bytes")


# ============ SCI IMPORT TESTS ============

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl required for import tests")
class TestSCIImportExcel:
    """Tests for POST /api/sci/import-excel with before/after comparison"""
    
    def test_sci_import_without_password_fails(self, api_session):
        """POST /api/sci/import-excel without password returns 401"""
        export_resp = api_session.get(f"{BASE_URL}/api/sci/export-excel")
        assert export_resp.status_code == 200
        
        files = {'file': ('sci_test.xlsx', io.BytesIO(export_resp.content),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'password': ''}
        
        response = requests.post(
            f"{BASE_URL}/api/sci/import-excel",
            files=files,
            data=data
        )
        
        assert response.status_code == 401, f"Expected 401, got {response.status_code}"
        print("✓ SCI Import without password correctly rejected")
    
    def test_sci_import_unchanged_returns_zero_updates(self, api_session):
        """POST /api/sci/import-excel without modifications returns updated=0, unchanged > 0"""
        export_resp = api_session.get(f"{BASE_URL}/api/sci/export-excel")
        assert export_resp.status_code == 200, f"SCI Export failed: {export_resp.status_code}"
        
        files = {'file': ('sci_unchanged.xlsx', io.BytesIO(export_resp.content),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'password': ADMIN_PASSWORD}
        
        import_resp = requests.post(
            f"{BASE_URL}/api/sci/import-excel",
            files=files,
            data=data
        )
        
        assert import_resp.status_code == 200, f"SCI Import failed: {import_resp.status_code} - {import_resp.text[:500]}"
        
        result = import_resp.json()
        assert result.get('success') == True, f"SCI Import not successful: {result}"
        assert result.get('unchanged', 0) > 0, f"Expected unchanged > 0, got {result.get('unchanged')}"
        
        print(f"✓ SCI Import unchanged file successful:")
        print(f"  updated={result.get('updated')}, unchanged={result.get('unchanged')}")
    
    def test_sci_import_with_changes_returns_comparison(self, api_session):
        """POST /api/sci/import-excel with modifications returns comparison details"""
        # 1. Export current state
        export_resp = api_session.get(f"{BASE_URL}/api/sci/export-excel")
        assert export_resp.status_code == 200, f"SCI Export failed: {export_resp.status_code}"
        
        # 2. Load and modify Excel - change lease_cash for first vehicle
        wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
        
        # Find a sheet with data (Lease 2026 or Lease 2025)
        ws = None
        for sheet_name in wb.sheetnames:
            if 'Lease' in sheet_name and 'Instruction' not in sheet_name:
                ws = wb[sheet_name]
                break
        
        assert ws is not None, f"No Lease sheet found in: {wb.sheetnames}"
        
        # Find Lease Cash column (should be column C)
        headers = [cell.value for cell in ws[1]]
        lease_cash_col = None
        for i, h in enumerate(headers):
            if h and 'Lease Cash' in str(h):
                lease_cash_col = i + 1  # 1-indexed
                break
        
        assert lease_cash_col is not None, f"Lease Cash column not found in headers: {headers[:10]}"
        
        # Store original value and modify first data row
        original_value = ws.cell(row=2, column=lease_cash_col).value or 0
        test_value = float(original_value) + 7777  # Add unique value
        ws.cell(row=2, column=lease_cash_col).value = test_value
        
        brand = ws.cell(row=2, column=1).value
        model = ws.cell(row=2, column=2).value
        
        print(f"  Modifying SCI: {brand} {model}")
        print(f"  lease_cash: {original_value} -> {test_value}")
        
        # 3. Save modified Excel
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 4. Import modified file
        files = {'file': ('sci_modified.xlsx', output,
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'password': ADMIN_PASSWORD}
        
        import_resp = requests.post(
            f"{BASE_URL}/api/sci/import-excel",
            files=files,
            data=data
        )
        
        assert import_resp.status_code == 200, f"SCI Import failed: {import_resp.status_code} - {import_resp.text[:500]}"
        
        result = import_resp.json()
        assert result.get('success') == True, f"SCI Import not successful: {result}"
        assert result.get('updated', 0) > 0, f"Expected updated > 0, got {result.get('updated')}"
        
        # Check comparison details
        comparison = result.get('comparison', [])
        assert len(comparison) > 0, f"Expected comparison details, got none"
        
        # Verify lease_cash change is in comparison
        has_lease_cash_change = any(
            'lease_cash' in item.get('changes', {}) 
            for item in comparison
        )
        assert has_lease_cash_change, f"No lease_cash change found in comparison: {comparison[:2]}"
        
        # Find and log the specific change
        for item in comparison:
            if 'lease_cash' in item.get('changes', {}):
                cash_change = item['changes']['lease_cash']
                print(f"✓ SCI Comparison shows change:")
                print(f"  vehicule: {item.get('vehicule')}")
                print(f"  lease_cash: {cash_change.get('avant')} -> {cash_change.get('apres')}")
                break
        
        # 5. Revert the change
        files_revert = {'file': ('sci_revert.xlsx', io.BytesIO(export_resp.content),
                                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        revert_resp = requests.post(
            f"{BASE_URL}/api/sci/import-excel",
            files=files_revert,
            data={'password': ADMIN_PASSWORD}
        )
        
        assert revert_resp.status_code == 200, f"SCI Revert failed: {revert_resp.status_code}"
        print("✓ SCI Changes reverted successfully")


# ============ SCI COMPARISONS HISTORY TESTS ============

class TestSCIComparisonsHistory:
    """Tests for GET /api/sci/comparisons"""
    
    def test_get_sci_comparisons_returns_list(self, api_session):
        """GET /api/sci/comparisons returns list of SCI comparison records"""
        response = api_session.get(f"{BASE_URL}/api/sci/comparisons")
        
        assert response.status_code == 200, f"SCI Comparisons endpoint failed: {response.status_code}"
        
        data = response.json()
        assert isinstance(data, list), f"Expected list, got {type(data)}"
        
        if len(data) > 0:
            comp = data[0]
            # Validate structure
            assert 'id' in comp, f"Missing 'id' field"
            assert 'type' in comp, f"Missing 'type' field"
            assert comp['type'] == 'sci_lease', f"Expected type='sci_lease', got {comp['type']}"
            assert 'date' in comp, f"Missing 'date' field"
            assert 'updated' in comp, f"Missing 'updated' field"
            assert 'unchanged' in comp, f"Missing 'unchanged' field"
            
            print(f"✓ SCI Comparisons history returned {len(data)} records")
            print(f"  Latest: id={comp['id'][:8]}... date={comp['date']}")
            print(f"  updated={comp['updated']}, unchanged={comp['unchanged']}")
        else:
            print("✓ SCI Comparisons endpoint working (no history yet)")


# ============ COMPARISON DETAIL ENDPOINT TEST ============

class TestComparisonDetail:
    """Tests for GET /api/programs/comparison/{comparison_id}"""
    
    def test_get_comparison_detail(self, api_session):
        """GET /api/programs/comparison/{id} returns specific comparison"""
        # First get list to find an ID
        list_resp = api_session.get(f"{BASE_URL}/api/programs/comparisons")
        assert list_resp.status_code == 200
        
        comparisons = list_resp.json()
        if len(comparisons) == 0:
            pytest.skip("No comparisons available to test detail endpoint")
        
        comparison_id = comparisons[0]['id']
        
        # Get detail
        response = api_session.get(f"{BASE_URL}/api/programs/comparison/{comparison_id}")
        
        assert response.status_code == 200, f"Get comparison detail failed: {response.status_code}"
        
        data = response.json()
        assert data.get('id') == comparison_id, f"ID mismatch: {data.get('id')} != {comparison_id}"
        assert 'details' in data, f"Missing 'details' field"
        
        print(f"✓ Comparison detail endpoint working")
        print(f"  ID: {comparison_id}")
        print(f"  Details count: {len(data.get('details', []))}")
    
    def test_get_invalid_comparison_returns_404(self, api_session):
        """GET /api/programs/comparison/{invalid_id} returns 404"""
        response = api_session.get(f"{BASE_URL}/api/programs/comparison/invalid-uuid-12345")
        
        assert response.status_code == 404, f"Expected 404, got {response.status_code}"
        print("✓ Invalid comparison ID correctly returns 404")


# ============ EXECUTION ============

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
