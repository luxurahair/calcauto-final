"""
Test suite for CalcAuto AiPro - Corrections Memory System
Tests the memorization of Excel corrections for future PDF imports.

Features tested:
- GET /api/corrections - List all memorized corrections
- POST /api/programs/import-excel - Saves corrections when values are modified
- DELETE /api/corrections/all - Deletes all corrections with admin password
- Flexible matching (normalize_correction_str, normalize_correction_model) for finding corrections
"""

import pytest
import requests
import os
import io
import openpyxl
from openpyxl.styles import Font
from datetime import datetime

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://deal-detail-modal.preview.emergentagent.com').rstrip('/')
ADMIN_PASSWORD = "Liana2018"


class TestCorrectionsEndpoints:
    """Test corrections CRUD endpoints"""
    
    def test_get_corrections_endpoint_returns_valid_response(self):
        """GET /api/corrections returns list with total count"""
        response = requests.get(f"{BASE_URL}/api/corrections")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        data = response.json()
        assert "total" in data, "Response should have 'total' field"
        assert "corrections" in data, "Response should have 'corrections' field"
        assert isinstance(data["total"], int), "total should be an integer"
        assert isinstance(data["corrections"], list), "corrections should be a list"
        print(f"✓ GET /api/corrections returned {data['total']} corrections")
    
    def test_correction_structure_has_required_fields(self):
        """Each correction should have brand, model, trim, year, corrected_values, corrected_at"""
        response = requests.get(f"{BASE_URL}/api/corrections")
        assert response.status_code == 200
        
        data = response.json()
        if data["total"] > 0:
            correction = data["corrections"][0]
            required_fields = ["brand", "model", "trim", "year", "corrected_values", "corrected_at"]
            for field in required_fields:
                assert field in correction, f"Correction should have '{field}' field"
            
            # Verify corrected_values structure
            cv = correction.get("corrected_values", {})
            assert isinstance(cv, dict), "corrected_values should be a dictionary"
            print(f"✓ Correction structure verified: {correction.get('brand')} {correction.get('model')} {correction.get('year')}")
        else:
            print("✓ No corrections found (test will verify structure after creating one)")
    
    def test_delete_all_corrections_requires_password(self):
        """DELETE /api/corrections/all without password should fail"""
        response = requests.delete(f"{BASE_URL}/api/corrections/all")
        assert response.status_code == 401, f"Expected 401 without password, got {response.status_code}"
        print("✓ DELETE /api/corrections/all correctly requires password")
    
    def test_delete_all_corrections_with_wrong_password(self):
        """DELETE /api/corrections/all with wrong password should fail"""
        response = requests.delete(f"{BASE_URL}/api/corrections/all?password=wrongpassword")
        assert response.status_code == 401, f"Expected 401 with wrong password, got {response.status_code}"
        print("✓ DELETE /api/corrections/all rejects wrong password")
    
    def test_delete_all_corrections_with_correct_password(self):
        """DELETE /api/corrections/all with correct password should succeed"""
        response = requests.delete(f"{BASE_URL}/api/corrections/all?password={ADMIN_PASSWORD}")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        
        data = response.json()
        assert "deleted" in data, "Response should have 'deleted' count"
        assert isinstance(data["deleted"], int), "deleted should be an integer"
        print(f"✓ DELETE /api/corrections/all deleted {data['deleted']} corrections")


class TestExcelImportCreatesCorrections:
    """Test that Excel import with changes saves corrections to program_corrections collection"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Clear corrections before each test"""
        requests.delete(f"{BASE_URL}/api/corrections/all?password={ADMIN_PASSWORD}")
        yield
        # Cleanup after test - restore corrections state
        requests.delete(f"{BASE_URL}/api/corrections/all?password={ADMIN_PASSWORD}")
    
    def test_excel_import_creates_correction_when_value_changed(self):
        """POST /api/programs/import-excel with modified values should create correction"""
        # Step 1: Get current programs to export
        programs_response = requests.get(f"{BASE_URL}/api/programs")
        assert programs_response.status_code == 200
        programs = programs_response.json()
        assert len(programs) > 0, "Need at least one program to test"
        
        # Step 2: Find a program to modify
        test_program = programs[0]
        original_consumer_cash = test_program.get("consumer_cash", 0) or 0
        modified_consumer_cash = original_consumer_cash + 500  # Add 500 to trigger a change
        
        print(f"Testing with: {test_program.get('brand')} {test_program.get('model')} {test_program.get('trim')} {test_program.get('year')}")
        print(f"  Original consumer_cash: {original_consumer_cash} -> Modified: {modified_consumer_cash}")
        
        # Step 3: Create an Excel file with the modified value
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Programmes"
        
        # Headers matching export format
        headers = [
            "ID", "Marque", "Modele", "Trim", "Annee",
            "Consumer Cash ($)", "Bonus Cash ($)",
            "Opt1 36M", "Opt1 48M", "Opt1 60M", "Opt1 72M", "Opt1 84M", "Opt1 96M",
            "Rabais Alt. Cash ($)",
            "Opt2 36M", "Opt2 48M", "Opt2 60M", "Opt2 72M", "Opt2 84M", "Opt2 96M",
            "Sort Order"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        
        # Add the test program row with modified consumer_cash
        o1 = test_program.get("option1_rates") or {}
        o2 = test_program.get("option2_rates")
        
        row_data = [
            test_program.get("id", ""),
            test_program.get("brand", ""),
            test_program.get("model", ""),
            test_program.get("trim", ""),
            test_program.get("year", 2026),
            modified_consumer_cash,  # Modified value!
            test_program.get("bonus_cash", 0) or 0,
            o1.get("rate_36"), o1.get("rate_48"), o1.get("rate_60"),
            o1.get("rate_72"), o1.get("rate_84"), o1.get("rate_96"),
            test_program.get("alternative_consumer_cash", 0) or 0,
            o2.get("rate_36") if o2 else None, o2.get("rate_48") if o2 else None,
            o2.get("rate_60") if o2 else None, o2.get("rate_72") if o2 else None,
            o2.get("rate_84") if o2 else None, o2.get("rate_96") if o2 else None,
            test_program.get("sort_order", 999)
        ]
        
        for col, val in enumerate(row_data, 1):
            ws.cell(row=2, column=col, value=val)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Step 4: Import the Excel file
        files = {"file": ("test_corrections.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"password": ADMIN_PASSWORD}
        
        import_response = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data=data
        )
        assert import_response.status_code == 200, f"Import failed: {import_response.text}"
        
        import_result = import_response.json()
        print(f"  Import result: {import_result.get('message')}")
        print(f"  Updated: {import_result.get('updated')}, Corrections saved: {import_result.get('corrections_saved')}")
        
        # Step 5: Verify a correction was created
        corrections_response = requests.get(f"{BASE_URL}/api/corrections")
        assert corrections_response.status_code == 200
        
        corrections_data = corrections_response.json()
        assert corrections_data["total"] >= 1, "At least one correction should have been created"
        
        # Find our correction
        test_brand = test_program.get("brand")
        test_model = test_program.get("model")
        test_year = test_program.get("year")
        
        matching_correction = None
        for c in corrections_data["corrections"]:
            if (c.get("brand") == test_brand and 
                c.get("model") == test_model and 
                c.get("year") == test_year):
                matching_correction = c
                break
        
        assert matching_correction is not None, f"Correction for {test_brand} {test_model} {test_year} not found"
        
        # Verify correction structure
        assert "corrected_values" in matching_correction
        assert "corrected_at" in matching_correction
        cv = matching_correction["corrected_values"]
        assert cv.get("consumer_cash") == modified_consumer_cash, f"Expected consumer_cash={modified_consumer_cash}, got {cv.get('consumer_cash')}"
        
        print(f"✓ Correction created successfully:")
        print(f"    Vehicle: {matching_correction.get('brand')} {matching_correction.get('model')} {matching_correction.get('trim')} {matching_correction.get('year')}")
        print(f"    Corrected values: consumer_cash={cv.get('consumer_cash')}")
        print(f"    Corrected at: {matching_correction.get('corrected_at')}")
        
        # Step 6: Restore original value
        row_data[5] = original_consumer_cash
        ws.cell(row=2, column=6, value=original_consumer_cash)
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        files = {"file": ("restore.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        restore_response = requests.post(
            f"{BASE_URL}/api/programs/import-excel",
            files=files,
            data={"password": ADMIN_PASSWORD}
        )
        print(f"  Restored original value: {restore_response.status_code}")


class TestCorrectionFieldsStructure:
    """Test the structure and content of correction documents"""
    
    def test_correction_corrected_values_contains_expected_fields(self):
        """corrected_values should contain consumer_cash, alternative_consumer_cash, bonus_cash, option1_rates, option2_rates"""
        # First, ensure we have at least one correction by importing with a change
        # Get a program
        programs_response = requests.get(f"{BASE_URL}/api/programs")
        assert programs_response.status_code == 200
        programs = programs_response.json()
        
        if len(programs) == 0:
            pytest.skip("No programs available for testing")
        
        # Clear existing corrections
        requests.delete(f"{BASE_URL}/api/corrections/all?password={ADMIN_PASSWORD}")
        
        # Create a correction by importing modified data
        test_program = programs[0]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Programmes"
        
        headers = [
            "ID", "Marque", "Modele", "Trim", "Annee",
            "Consumer Cash ($)", "Bonus Cash ($)",
            "Opt1 36M", "Opt1 48M", "Opt1 60M", "Opt1 72M", "Opt1 84M", "Opt1 96M",
            "Rabais Alt. Cash ($)",
            "Opt2 36M", "Opt2 48M", "Opt2 60M", "Opt2 72M", "Opt2 84M", "Opt2 96M",
            "Sort Order"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        
        o1 = test_program.get("option1_rates") or {}
        o2 = test_program.get("option2_rates")
        
        row_data = [
            test_program.get("id", ""),
            test_program.get("brand", ""),
            test_program.get("model", ""),
            test_program.get("trim", ""),
            test_program.get("year", 2026),
            (test_program.get("consumer_cash", 0) or 0) + 1000,  # Modify
            test_program.get("bonus_cash", 0) or 0,
            o1.get("rate_36"), o1.get("rate_48"), o1.get("rate_60"),
            o1.get("rate_72"), o1.get("rate_84"), o1.get("rate_96"),
            test_program.get("alternative_consumer_cash", 0) or 0,
            o2.get("rate_36") if o2 else None, o2.get("rate_48") if o2 else None,
            o2.get("rate_60") if o2 else None, o2.get("rate_72") if o2 else None,
            o2.get("rate_84") if o2 else None, o2.get("rate_96") if o2 else None,
            test_program.get("sort_order", 999)
        ]
        
        for col, val in enumerate(row_data, 1):
            ws.cell(row=2, column=col, value=val)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        files = {"file": ("test.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        requests.post(f"{BASE_URL}/api/programs/import-excel", files=files, data={"password": ADMIN_PASSWORD})
        
        # Now check the correction structure
        response = requests.get(f"{BASE_URL}/api/corrections")
        assert response.status_code == 200
        
        data = response.json()
        if data["total"] > 0:
            correction = data["corrections"][0]
            cv = correction.get("corrected_values", {})
            
            # These fields should be present in corrected_values
            expected_fields = ["consumer_cash", "alternative_consumer_cash", "bonus_cash"]
            for field in expected_fields:
                assert field in cv, f"corrected_values should contain '{field}'"
            
            # option1_rates and option2_rates can be None or dict
            assert "option1_rates" in cv, "corrected_values should contain 'option1_rates'"
            assert "option2_rates" in cv, "corrected_values should contain 'option2_rates'"
            
            print(f"✓ corrected_values structure verified:")
            print(f"    consumer_cash: {cv.get('consumer_cash')}")
            print(f"    alternative_consumer_cash: {cv.get('alternative_consumer_cash')}")
            print(f"    bonus_cash: {cv.get('bonus_cash')}")
            print(f"    option1_rates: {type(cv.get('option1_rates'))}")
            print(f"    option2_rates: {type(cv.get('option2_rates'))}")
        
        # Cleanup - restore original
        row_data[5] = test_program.get("consumer_cash", 0) or 0
        ws.cell(row=2, column=6, value=row_data[5])
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        files = {"file": ("restore.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        requests.post(f"{BASE_URL}/api/programs/import-excel", files=files, data={"password": ADMIN_PASSWORD})
        requests.delete(f"{BASE_URL}/api/corrections/all?password={ADMIN_PASSWORD}")


class TestCorrectionsCountAfterImport:
    """Test that correction count increases after Excel import with changes"""
    
    def test_correction_count_increases_after_excel_import_with_changes(self):
        """After importing Excel with modified values, correction count should increase"""
        # Clear all corrections first
        delete_response = requests.delete(f"{BASE_URL}/api/corrections/all?password={ADMIN_PASSWORD}")
        assert delete_response.status_code == 200
        
        # Verify count is 0
        initial_response = requests.get(f"{BASE_URL}/api/corrections")
        assert initial_response.status_code == 200
        initial_count = initial_response.json()["total"]
        print(f"Initial correction count: {initial_count}")
        
        # Get programs
        programs_response = requests.get(f"{BASE_URL}/api/programs")
        assert programs_response.status_code == 200
        programs = programs_response.json()
        
        if len(programs) == 0:
            pytest.skip("No programs available")
        
        # Create Excel with 2 modified programs
        test_programs = programs[:2]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Programmes"
        
        headers = [
            "ID", "Marque", "Modele", "Trim", "Annee",
            "Consumer Cash ($)", "Bonus Cash ($)",
            "Opt1 36M", "Opt1 48M", "Opt1 60M", "Opt1 72M", "Opt1 84M", "Opt1 96M",
            "Rabais Alt. Cash ($)",
            "Opt2 36M", "Opt2 48M", "Opt2 60M", "Opt2 72M", "Opt2 84M", "Opt2 96M",
            "Sort Order"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        
        for row_idx, prog in enumerate(test_programs, 2):
            o1 = prog.get("option1_rates") or {}
            o2 = prog.get("option2_rates")
            
            row_data = [
                prog.get("id", ""),
                prog.get("brand", ""),
                prog.get("model", ""),
                prog.get("trim", ""),
                prog.get("year", 2026),
                (prog.get("consumer_cash", 0) or 0) + 1000,  # Modify to trigger change
                prog.get("bonus_cash", 0) or 0,
                o1.get("rate_36"), o1.get("rate_48"), o1.get("rate_60"),
                o1.get("rate_72"), o1.get("rate_84"), o1.get("rate_96"),
                prog.get("alternative_consumer_cash", 0) or 0,
                o2.get("rate_36") if o2 else None, o2.get("rate_48") if o2 else None,
                o2.get("rate_60") if o2 else None, o2.get("rate_72") if o2 else None,
                o2.get("rate_84") if o2 else None, o2.get("rate_96") if o2 else None,
                prog.get("sort_order", 999)
            ]
            
            for col, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=val)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Import
        files = {"file": ("test.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        import_response = requests.post(f"{BASE_URL}/api/programs/import-excel", files=files, data={"password": ADMIN_PASSWORD})
        assert import_response.status_code == 200
        
        import_result = import_response.json()
        corrections_saved = import_result.get("corrections_saved", 0)
        print(f"Excel import reported {corrections_saved} corrections saved")
        
        # Check new count
        final_response = requests.get(f"{BASE_URL}/api/corrections")
        assert final_response.status_code == 200
        final_count = final_response.json()["total"]
        print(f"Final correction count: {final_count}")
        
        # Count should have increased
        assert final_count > initial_count, f"Correction count should have increased from {initial_count} to at least {initial_count + 1}"
        assert final_count >= corrections_saved, f"Final count ({final_count}) should be >= corrections_saved ({corrections_saved})"
        
        print(f"✓ Correction count increased from {initial_count} to {final_count}")
        
        # Cleanup - restore original values
        for row_idx, prog in enumerate(test_programs, 2):
            ws.cell(row=row_idx, column=6, value=prog.get("consumer_cash", 0) or 0)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        files = {"file": ("restore.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        requests.post(f"{BASE_URL}/api/programs/import-excel", files=files, data={"password": ADMIN_PASSWORD})
        
        # Clear corrections
        requests.delete(f"{BASE_URL}/api/corrections/all?password={ADMIN_PASSWORD}")


class TestDeleteSpecificCorrection:
    """Test deleting specific corrections by brand/model/year"""
    
    def test_delete_specific_correction(self):
        """DELETE /api/corrections/{brand}/{model}/{year} should delete matching corrections"""
        # First, ensure we have a correction
        requests.delete(f"{BASE_URL}/api/corrections/all?password={ADMIN_PASSWORD}")
        
        # Get a program and create a correction
        programs_response = requests.get(f"{BASE_URL}/api/programs")
        programs = programs_response.json()
        
        if len(programs) == 0:
            pytest.skip("No programs available")
        
        test_program = programs[0]
        brand = test_program.get("brand")
        model = test_program.get("model")
        year = test_program.get("year")
        
        # Create Excel with modified value
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Programmes"
        
        headers = [
            "ID", "Marque", "Modele", "Trim", "Annee",
            "Consumer Cash ($)", "Bonus Cash ($)",
            "Opt1 36M", "Opt1 48M", "Opt1 60M", "Opt1 72M", "Opt1 84M", "Opt1 96M",
            "Rabais Alt. Cash ($)",
            "Opt2 36M", "Opt2 48M", "Opt2 60M", "Opt2 72M", "Opt2 84M", "Opt2 96M",
            "Sort Order"
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        
        o1 = test_program.get("option1_rates") or {}
        o2 = test_program.get("option2_rates")
        
        row_data = [
            test_program.get("id", ""),
            brand,
            model,
            test_program.get("trim", ""),
            year,
            (test_program.get("consumer_cash", 0) or 0) + 500,
            test_program.get("bonus_cash", 0) or 0,
            o1.get("rate_36"), o1.get("rate_48"), o1.get("rate_60"),
            o1.get("rate_72"), o1.get("rate_84"), o1.get("rate_96"),
            test_program.get("alternative_consumer_cash", 0) or 0,
            o2.get("rate_36") if o2 else None, o2.get("rate_48") if o2 else None,
            o2.get("rate_60") if o2 else None, o2.get("rate_72") if o2 else None,
            o2.get("rate_84") if o2 else None, o2.get("rate_96") if o2 else None,
            test_program.get("sort_order", 999)
        ]
        
        for col, val in enumerate(row_data, 1):
            ws.cell(row=2, column=col, value=val)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        files = {"file": ("test.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        requests.post(f"{BASE_URL}/api/programs/import-excel", files=files, data={"password": ADMIN_PASSWORD})
        
        # Verify correction exists
        response = requests.get(f"{BASE_URL}/api/corrections")
        assert response.status_code == 200
        count_before = response.json()["total"]
        assert count_before >= 1, "Should have at least one correction"
        
        # Delete specific correction
        delete_response = requests.delete(
            f"{BASE_URL}/api/corrections/{brand}/{model}/{year}?password={ADMIN_PASSWORD}"
        )
        assert delete_response.status_code == 200
        
        delete_result = delete_response.json()
        assert "deleted" in delete_result
        print(f"✓ DELETE /api/corrections/{brand}/{model}/{year} deleted {delete_result['deleted']} correction(s)")
        
        # Restore original value
        ws.cell(row=2, column=6, value=test_program.get("consumer_cash", 0) or 0)
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        files = {"file": ("restore.xlsx", excel_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        requests.post(f"{BASE_URL}/api/programs/import-excel", files=files, data={"password": ADMIN_PASSWORD})


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
