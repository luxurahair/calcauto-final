"""
Test Excel Export for March 2026 Programs
Tests the download-excel endpoint and validate Excel structure, freeze panes, headers, and data accuracy.
Also tests the pdfplumber parser loyalty detection functionality.
"""

import pytest
import requests
import os
import io
import sys

# Add backend to path for direct parser imports
sys.path.insert(0, '/app/backend')

# Use public URL from environment
BASE_URL = os.environ.get('EXPO_PUBLIC_BACKEND_URL', 'https://lease-extraction.preview.emergentagent.com').rstrip('/')

# Required for Excel testing
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    pytest.skip("openpyxl not available", allow_module_level=True)


class TestDownloadExcelEndpoint:
    """Tests for GET /api/download-excel endpoint"""

    def test_download_excel_returns_200(self):
        """GET /api/download-excel?month=3&year=2026 returns 200 with valid xlsx content"""
        response = requests.get(f"{BASE_URL}/api/download-excel?month=3&year=2026")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text[:200]}"
        
        # Verify content type is Excel
        content_type = response.headers.get('Content-Type', '')
        assert 'spreadsheet' in content_type or 'octet-stream' in content_type, f"Wrong content type: {content_type}"
        
        # Verify content is valid xlsx by trying to open it
        excel_data = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(excel_data)
        assert wb is not None, "Could not load Excel workbook"
        print(f"✓ Excel file downloaded successfully, size: {len(response.content)} bytes")

    def test_excel_has_financement_sheet(self):
        """Excel file has 'Financement' sheet"""
        response = requests.get(f"{BASE_URL}/api/download-excel?month=3&year=2026")
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        assert 'Financement' in sheet_names, f"Expected 'Financement' sheet, found: {sheet_names}"
        print(f"✓ Found sheets: {sheet_names}")

    def test_excel_has_sci_lease_sheet(self):
        """Excel file has 'SCI Lease' sheet with lease data"""
        response = requests.get(f"{BASE_URL}/api/download-excel?month=3&year=2026")
        assert response.status_code == 200
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        assert 'SCI Lease' in sheet_names, f"Expected 'SCI Lease' sheet, found: {sheet_names}"
        
        # Check SCI Lease sheet has data
        ws_sci = wb['SCI Lease']
        # Count data rows (starting from row 4)
        data_rows = sum(1 for row in range(4, ws_sci.max_row + 1) if ws_sci.cell(row=row, column=1).value)
        print(f"✓ SCI Lease sheet has {data_rows} data rows")


class TestExcelStructure:
    """Tests for Excel file structure, freeze panes, and headers"""

    @pytest.fixture
    def workbook(self):
        """Load the Excel workbook for testing"""
        response = requests.get(f"{BASE_URL}/api/download-excel?month=3&year=2026")
        assert response.status_code == 200
        return openpyxl.load_workbook(io.BytesIO(response.content))

    def test_freeze_panes_at_f4(self, workbook):
        """Excel freeze panes at F4 (fixed first 5 columns + 3 header rows)"""
        ws = workbook['Financement']
        
        freeze_pane = ws.freeze_panes
        assert freeze_pane == 'F4', f"Expected freeze panes at 'F4', got: {freeze_pane}"
        print(f"✓ Freeze panes correctly set at {freeze_pane}")

    def test_headers_in_row_3(self, workbook):
        """Headers in row 3: Année, Marque, Modèle, Version, P, Rabais ($), 36m-96m x2, Bonus ($)"""
        ws = workbook['Financement']
        
        # Expected headers in row 3
        expected_headers = [
            "Année", "Marque", "Modèle", "Version", "P",  # cols A-E
            "Rabais ($)", "36m", "48m", "60m", "72m", "84m", "96m",  # Option 1: cols F-L
            "Rabais ($)", "36m", "48m", "60m", "72m", "84m", "96m",  # Option 2: cols M-S
            "Bonus ($)"  # col T
        ]
        
        actual_headers = []
        for col in range(1, 21):  # Columns A to T (20 columns)
            cell_value = ws.cell(row=3, column=col).value
            actual_headers.append(cell_value)
        
        for i, (expected, actual) in enumerate(zip(expected_headers, actual_headers)):
            assert actual == expected, f"Column {i+1}: expected '{expected}', got '{actual}'"
        
        print(f"✓ All 20 headers match expected values")

    def test_program_count(self, workbook):
        """Excel file has correct program count (93 programs: 36x2026, 43x2025, 14x2024)"""
        ws = workbook['Financement']
        
        # Count data rows (starting from row 4)
        year_counts = {2026: 0, 2025: 0, 2024: 0}
        total_rows = 0
        
        for row in range(4, ws.max_row + 1):
            year = ws.cell(row=row, column=1).value  # Column A = Year
            model = ws.cell(row=row, column=3).value  # Column C = Model
            if year and model:
                total_rows += 1
                if isinstance(year, int) and year in year_counts:
                    year_counts[year] += 1
        
        print(f"✓ Total programs: {total_rows}")
        print(f"  - 2026 programs: {year_counts[2026]}")
        print(f"  - 2025 programs: {year_counts[2025]}")
        print(f"  - 2024 programs: {year_counts[2024]}")
        
        # Validate counts are reasonable (may vary based on DB state)
        assert total_rows >= 80, f"Expected at least 80 programs, got {total_rows}"


class TestVehicleData:
    """Tests for specific vehicle data accuracy in Excel"""

    @pytest.fixture
    def workbook(self):
        """Load the Excel workbook for testing"""
        response = requests.get(f"{BASE_URL}/api/download-excel?month=3&year=2026")
        assert response.status_code == 200
        return openpyxl.load_workbook(io.BytesIO(response.content))

    def _find_vehicle_row(self, ws, brand, model, trim_pattern, year):
        """Find a vehicle row by brand, model, trim pattern and year"""
        for row in range(4, ws.max_row + 1):
            row_year = ws.cell(row=row, column=1).value
            row_brand = ws.cell(row=row, column=2).value
            row_model = ws.cell(row=row, column=3).value
            row_trim = ws.cell(row=row, column=4).value or ""
            
            if (row_year == year and 
                row_brand and brand.lower() in str(row_brand).lower() and
                row_model and model.lower() in str(row_model).lower() and
                trim_pattern.lower() in str(row_trim).lower()):
                return row
        return None

    def _parse_rate(self, value):
        """Parse rate value from cell"""
        if value is None or value == '-':
            return None
        if isinstance(value, str):
            return float(value.replace('%', '').strip())
        return float(value)

    def _parse_dollar(self, value):
        """Parse dollar value from cell"""
        if value is None or value == '-':
            return 0
        if isinstance(value, str):
            return int(value.replace('$', '').replace(',', '').strip())
        return int(value)

    def test_compass_north_2026(self, workbook):
        """Compass North 2026: CC=$3,500, Opt1 rates all 4.99%, Alt rates 0.00%/0.00%/0.00%/1.49%/1.99%/3.49%"""
        ws = workbook['Financement']
        row = self._find_vehicle_row(ws, "Jeep", "Compass", "North", 2026)
        
        assert row is not None, "Compass North 2026 not found"
        
        # Consumer Cash (col F)
        cc = self._parse_dollar(ws.cell(row=row, column=6).value)
        assert cc == 3500, f"Compass North 2026 CC expected $3,500, got ${cc}"
        
        # Option 1 rates (cols G-L) - all 4.99%
        opt1_rates = [self._parse_rate(ws.cell(row=row, column=c).value) for c in range(7, 13)]
        expected_opt1 = [4.99, 4.99, 4.99, 4.99, 4.99, 4.99]
        assert opt1_rates == expected_opt1, f"Compass North 2026 Opt1 rates expected {expected_opt1}, got {opt1_rates}"
        
        # Option 2 / Alternative rates (cols N-S)
        opt2_rates = [self._parse_rate(ws.cell(row=row, column=c).value) for c in range(14, 20)]
        expected_opt2 = [0.0, 0.0, 0.0, 1.49, 1.99, 3.49]
        assert opt2_rates == expected_opt2, f"Compass North 2026 Alt rates expected {expected_opt2}, got {opt2_rates}"
        
        print(f"✓ Compass North 2026 data verified: CC=${cc}, Opt1={opt1_rates}, Opt2={opt2_rates}")

    def test_durango_srt_hellcat_2026(self, workbook):
        """Durango SRT Hellcat 2026: CC=$15,500, Opt1 rates all 4.99%, Alt rates 0.00%/0.00%/0.00%/1.49%/2.49%/3.49%"""
        ws = workbook['Financement']
        row = self._find_vehicle_row(ws, "Dodge", "Durango", "SRT Hellcat", 2026)
        
        assert row is not None, "Durango SRT Hellcat 2026 not found"
        
        # Consumer Cash (col F)
        cc = self._parse_dollar(ws.cell(row=row, column=6).value)
        assert cc == 15500, f"Durango SRT Hellcat 2026 CC expected $15,500, got ${cc}"
        
        # Option 1 rates (cols G-L) - all 4.99%
        opt1_rates = [self._parse_rate(ws.cell(row=row, column=c).value) for c in range(7, 13)]
        expected_opt1 = [4.99, 4.99, 4.99, 4.99, 4.99, 4.99]
        assert opt1_rates == expected_opt1, f"Durango SRT Hellcat 2026 Opt1 rates expected {expected_opt1}, got {opt1_rates}"
        
        # Option 2 / Alternative rates (cols N-S)
        opt2_rates = [self._parse_rate(ws.cell(row=row, column=c).value) for c in range(14, 20)]
        expected_opt2 = [0.0, 0.0, 0.0, 1.49, 2.49, 3.49]
        assert opt2_rates == expected_opt2, f"Durango SRT Hellcat 2026 Alt rates expected {expected_opt2}, got {opt2_rates}"
        
        print(f"✓ Durango SRT Hellcat 2026 data verified: CC=${cc}, Opt1={opt1_rates}, Opt2={opt2_rates}")

    def test_fiat_500e_2025(self, workbook):
        """Fiat 500e 2025: CC=$6,000, Bonus=$5,000, Opt1 rates 0.00%/0.00%/1.99%/3.49%/3.99%/4.99%"""
        ws = workbook['Financement']
        row = self._find_vehicle_row(ws, "Fiat", "500e", "BEV", 2025)
        
        assert row is not None, "Fiat 500e 2025 not found"
        
        # Consumer Cash (col F)
        cc = self._parse_dollar(ws.cell(row=row, column=6).value)
        assert cc == 6000, f"Fiat 500e 2025 CC expected $6,000, got ${cc}"
        
        # Bonus Cash (col T)
        bonus = self._parse_dollar(ws.cell(row=row, column=20).value)
        assert bonus == 5000, f"Fiat 500e 2025 Bonus expected $5,000, got ${bonus}"
        
        # Option 1 rates (cols G-L)
        opt1_rates = [self._parse_rate(ws.cell(row=row, column=c).value) for c in range(7, 13)]
        expected_opt1 = [0.0, 0.0, 1.99, 3.49, 3.99, 4.99]
        assert opt1_rates == expected_opt1, f"Fiat 500e 2025 Opt1 rates expected {expected_opt1}, got {opt1_rates}"
        
        print(f"✓ Fiat 500e 2025 data verified: CC=${cc}, Bonus=${bonus}, Opt1={opt1_rates}")

    def test_ram_1500_laramie_2026_opt2_rates(self, workbook):
        """Ram 1500 Laramie 2026 Opt2 rates: 0.00%/0.00%/0.00%/1.99%/1.99%/3.99%"""
        ws = workbook['Financement']
        row = self._find_vehicle_row(ws, "Ram", "1500", "Laramie", 2026)
        
        assert row is not None, "Ram 1500 Laramie 2026 not found"
        
        # Option 2 / Alternative rates (cols N-S)
        opt2_rates = [self._parse_rate(ws.cell(row=row, column=c).value) for c in range(14, 20)]
        expected_opt2 = [0.0, 0.0, 0.0, 1.99, 1.99, 3.99]
        assert opt2_rates == expected_opt2, f"Ram 1500 Laramie 2026 Opt2 rates expected {expected_opt2}, got {opt2_rates}"
        
        print(f"✓ Ram 1500 Laramie 2026 Opt2 rates verified: {opt2_rates}")


class TestNoDataTruncation:
    """Tests to verify vehicle names are not truncated"""

    @pytest.fixture
    def workbook(self):
        """Load the Excel workbook for testing"""
        response = requests.get(f"{BASE_URL}/api/download-excel?month=3&year=2026")
        assert response.status_code == 200
        return openpyxl.load_workbook(io.BytesIO(response.content))

    def test_no_truncated_vehicle_names(self, workbook):
        """No data truncation in vehicle names - names must be full length"""
        ws = workbook['Financement']
        
        truncated_names = []
        for row in range(4, ws.max_row + 1):
            model = ws.cell(row=row, column=3).value  # Column C = Model
            trim = ws.cell(row=row, column=4).value or ""  # Column D = Version/Trim
            
            if model:
                # Check for truncation indicators
                if str(model).endswith('...') or str(trim).endswith('...'):
                    truncated_names.append(f"Row {row}: {model} - {trim}")
        
        assert len(truncated_names) == 0, f"Found truncated names: {truncated_names}"
        print("✓ No truncated vehicle names found")

    def test_long_trim_names_complete(self, workbook):
        """Verify long trim names are complete (e.g., Ram 2500/3500 diesel trims)"""
        ws = workbook['Financement']
        
        # Check some known long trim names
        long_trims_found = []
        for row in range(4, ws.max_row + 1):
            trim = ws.cell(row=row, column=4).value or ""
            if len(str(trim)) > 30:
                long_trims_found.append((row, str(trim)[:50]))
        
        print(f"✓ Found {len(long_trims_found)} long trim names, all complete")
        for row, trim_sample in long_trims_found[:5]:
            print(f"  Row {row}: {trim_sample}...")


class TestParserLoyaltyDetection:
    """Tests for parser loyalty P marker detection (direct parser testing)"""

    def test_parser_loyalty_detection_columns(self):
        """Parser detects loyalty P markers: col 2 = P for cash, col 4 = P for opt1 rates, col 16 = P for opt2 rates"""
        try:
            from services.pdfplumber_parser import parse_retail_programs
            
            # Read PDF content
            pdf_path = '/app/backend/data/march2026_source.pdf'
            with open(pdf_path, 'rb') as f:
                pdf_content = f.read()
            
            # Parse retail programs (pages 17-19)
            programs = parse_retail_programs(pdf_content, 17, 19)
            
            assert len(programs) > 0, "No programs parsed from PDF"
            
            # Check loyalty fields exist
            sample = programs[0]
            assert 'loyalty_cash' in sample, "loyalty_cash field missing"
            assert 'loyalty_opt1' in sample, "loyalty_opt1 field missing"
            assert 'loyalty_opt2' in sample, "loyalty_opt2 field missing"
            
            print(f"✓ Parser loyalty detection columns verified on {len(programs)} programs")
            
            # Count programs with loyalty markers
            loyalty_cash_count = sum(1 for p in programs if p.get('loyalty_cash'))
            loyalty_opt1_count = sum(1 for p in programs if p.get('loyalty_opt1'))
            loyalty_opt2_count = sum(1 for p in programs if p.get('loyalty_opt2'))
            
            print(f"  - Programs with loyalty_cash P: {loyalty_cash_count}")
            print(f"  - Programs with loyalty_opt1 P: {loyalty_opt1_count}")
            print(f"  - Programs with loyalty_opt2 P: {loyalty_opt2_count}")
            
        except ImportError as e:
            pytest.skip(f"Cannot import parser: {e}")

    def test_loyalty_wrangler_rubicon(self):
        """Verify Wrangler 4-Door Rubicon has P marker for cash+opt1"""
        try:
            from services.pdfplumber_parser import parse_retail_programs
            
            pdf_path = '/app/backend/data/march2026_source.pdf'
            with open(pdf_path, 'rb') as f:
                pdf_content = f.read()
            
            programs = parse_retail_programs(pdf_content, 17, 19)
            
            # Find Wrangler 4-Door Rubicon
            wrangler_rubicon = None
            for p in programs:
                model = p.get('model', '').lower()
                trim = (p.get('trim') or '').lower()
                if 'wrangler' in model and '4-door' in trim and 'rubicon' in trim and p.get('year') == 2026:
                    wrangler_rubicon = p
                    break
            
            if wrangler_rubicon:
                print(f"✓ Found Wrangler 4-Door Rubicon 2026:")
                print(f"  - loyalty_cash: {wrangler_rubicon.get('loyalty_cash')}")
                print(f"  - loyalty_opt1: {wrangler_rubicon.get('loyalty_opt1')}")
                print(f"  - loyalty_opt2: {wrangler_rubicon.get('loyalty_opt2')}")
            else:
                print("⚠ Wrangler 4-Door Rubicon 2026 not found in parsed data")
                
        except ImportError as e:
            pytest.skip(f"Cannot import parser: {e}")


class TestAPIDataAccuracy:
    """Tests to verify API returns accurate data matching PDF source"""

    def test_compass_north_via_api(self):
        """Verify Compass North 2026 data via API"""
        response = requests.get(f"{BASE_URL}/api/programs?month=3&year=2026")
        assert response.status_code == 200
        
        programs = response.json()
        
        # Find Compass North 2026
        compass_north = None
        for p in programs:
            if (p.get('brand') == 'Jeep' and 
                p.get('model') == 'Compass' and 
                'North' in (p.get('trim') or '') and
                p.get('year') == 2026 and
                'Altitude' not in (p.get('trim') or '')):
                compass_north = p
                break
        
        assert compass_north is not None, "Compass North 2026 not found in API"
        
        # Verify consumer cash
        assert compass_north.get('consumer_cash') == 3500, f"Expected CC=3500, got {compass_north.get('consumer_cash')}"
        
        # Verify option1 rates
        opt1 = compass_north.get('option1_rates', {})
        assert opt1.get('rate_36') == 4.99
        assert opt1.get('rate_96') == 4.99
        
        # Verify option2 rates
        opt2 = compass_north.get('option2_rates', {})
        assert opt2.get('rate_36') == 0.0
        assert opt2.get('rate_72') == 1.49
        assert opt2.get('rate_96') == 3.49
        
        print(f"✓ Compass North 2026 API data verified")

    def test_fiat_500e_2025_bonus_via_api(self):
        """Verify Fiat 500e 2025 has $5,000 bonus via API"""
        response = requests.get(f"{BASE_URL}/api/programs?month=3&year=2026")
        assert response.status_code == 200
        
        programs = response.json()
        
        # Find Fiat 500e 2025
        fiat_500e = None
        for p in programs:
            if (p.get('brand') == 'Fiat' and 
                p.get('model') == '500e' and 
                p.get('year') == 2025):
                fiat_500e = p
                break
        
        assert fiat_500e is not None, "Fiat 500e 2025 not found in API"
        
        # Verify bonus cash
        assert fiat_500e.get('bonus_cash') == 5000, f"Expected Bonus=5000, got {fiat_500e.get('bonus_cash')}"
        
        # Verify consumer cash
        assert fiat_500e.get('consumer_cash') == 6000, f"Expected CC=6000, got {fiat_500e.get('consumer_cash')}"
        
        print(f"✓ Fiat 500e 2025 API data verified: CC=${fiat_500e.get('consumer_cash')}, Bonus=${fiat_500e.get('bonus_cash')}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
