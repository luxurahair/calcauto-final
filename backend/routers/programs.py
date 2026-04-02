from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
from datetime import datetime
import uuid
import re
from database import db, ADMIN_PASSWORD, logger
from models import (
    VehicleProgram, VehicleProgramCreate, VehicleProgramUpdate,
    CalculationRequest, PaymentComparison, CalculationResponse,
    ProgramPeriod, ImportRequest, FinancingRates
)
from dependencies import calculate_monthly_payment, get_rate_for_term
import pypdf
import io

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter()


def normalize_str(s: str) -> str:
    """Normalise une chaine pour matching flexible.
    Retire les codes produit (CPOS, WLJH74, JLXL74, etc.) mais garde les descripteurs (excluding, PHEV, Gas, etc.)."""
    if not s:
        return ""
    s = s.strip().lower()
    # Retirer les codes CPOS
    s = re.sub(r'\(cpos[^)]*\)', '', s, flags=re.IGNORECASE)
    # Retirer les codes produit: parentheses contenant LETTRES+CHIFFRES (ex: WLJH74, DJ7X91, KMJL74, DT6P98)
    s = re.sub(r'\([A-Z]{2,}[0-9]+[^)]*\)', '', s, flags=re.IGNORECASE)
    # Retirer (ETM) specifiquement
    s = re.sub(r'\(etm\)', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    s = s.rstrip(' ,')
    return s


def normalize_model(model: str) -> str:
    """Normalise le nom du modele pour matching flexible.
    Ex: 'Grand Cherokee/Grand Cherokee L' -> 'grand cherokee/l'"""
    if not model:
        return ""
    m = model.strip().lower()
    m = m.replace("grand cherokee/grand cherokee l", "grand cherokee/l")
    m = m.replace("grand wagoneer / grand wagoneer l", "grand wagoneer/l")
    m = m.replace("grand wagoneer/grand wagoneer l", "grand wagoneer/l")
    m = m.replace("wagoneer / wagoneer l", "wagoneer/l")
    m = m.replace("wagoneer/wagoneer l", "wagoneer/l")
    m = m.replace("2500 / 3500", "2500/3500")
    m = re.sub(r'\s+', ' ', m).strip()
    return m


def find_best_match(brand: str, model: str, trim: str, year: int, all_before: dict) -> tuple:
    """Trouve le meilleur match dans all_before avec strategies multiples.
    Retourne (prog, match_query) ou (None, None)."""
    # Strategie 1: Match exact
    key = f"{brand}|{model}|{trim}|{year}"
    if key in all_before:
        return all_before[key], {"brand": brand, "model": model, "trim": trim, "year": year}

    # Strategie 2: Match avec trim None vs ""
    for trim_variant in [None, ""]:
        if trim_variant != trim:
            key2 = f"{brand}|{model}|{trim_variant}|{year}"
            if key2 in all_before:
                return all_before[key2], {"brand": brand, "model": model, "trim": trim_variant, "year": year}

    # Strategie 3: Match normalise (modele + trim)
    norm_model = normalize_model(model)
    norm_trim = normalize_str(trim)

    for db_key, prog in all_before.items():
        parts = db_key.split("|")
        if len(parts) != 4:
            continue
        db_brand, db_model, db_trim, db_year = parts
        if db_brand.lower() != brand.lower() or str(db_year) != str(year):
            continue

        db_norm_model = normalize_model(db_model)
        db_norm_trim = normalize_str(db_trim or "")

        if db_norm_model == norm_model and db_norm_trim == norm_trim:
            return prog, {"brand": db_brand, "model": db_model, "trim": db_trim if db_trim != "None" else None, "year": int(db_year)}

    # Strategie 4: Match par modele normalise + trim partiel
    for db_key, prog in all_before.items():
        parts = db_key.split("|")
        if len(parts) != 4:
            continue
        db_brand, db_model, db_trim, db_year = parts
        if db_brand.lower() != brand.lower() or str(db_year) != str(year):
            continue

        db_norm_model = normalize_model(db_model)
        db_norm_trim = normalize_str(db_trim or "")

        # Verifier si le modele normalise match et les trims sont similaires
        if db_norm_model == norm_model:
            # Trim partiel: le trim DB est contenu dans le trim Excel (ou inversement)
            if norm_trim and db_norm_trim and (db_norm_trim in norm_trim or norm_trim in db_norm_trim):
                return prog, {"brand": db_brand, "model": db_model, "trim": db_trim if db_trim != "None" else None, "year": int(db_year)}

    # Strategie 5: Match par modele partiel (un contient l'autre)
    for db_key, prog in all_before.items():
        parts = db_key.split("|")
        if len(parts) != 4:
            continue
        db_brand, db_model, db_trim, db_year = parts
        if db_brand.lower() != brand.lower() or str(db_year) != str(year):
            continue

        db_norm_model = normalize_model(db_model)
        db_norm_trim = normalize_str(db_trim or "")

        # Si les deux trims sont vides/None
        if not norm_trim and not db_norm_trim:
            if db_norm_model in norm_model or norm_model in db_norm_model:
                return prog, {"brand": db_brand, "model": db_model, "trim": db_trim if db_trim != "None" else None, "year": int(db_year)}

    return None, None


async def compute_sort_order(brand: str, model: str, trim: Optional[str], year: int = 2026) -> int:
    """Compute sort_order for a program based on stored trim_orders in MongoDB.
    Uses exact (brand, model, trim) match against trim_orders collection."""
    # Find the trim_orders document for this brand/model/year
    trim_order_doc = await db.trim_orders.find_one({"brand": brand, "model": model, "year": year})
    if not trim_order_doc:
        # Fallback: try without year filter
        trim_order_doc = await db.trim_orders.find_one({"brand": brand, "model": model})

    if trim_order_doc:
        trims_list = trim_order_doc.get("trims", [])
        trim_val = trim if trim else "__none__"
        if trim_val in trims_list:
            return trims_list.index(trim_val)

    # Not found in trim_orders - return high value
    return 999

@router.post("/pdf-info")
async def get_pdf_info(file: UploadFile = File(...)):
    """Récupère les informations du PDF (nombre de pages)"""
    try:
        contents = await file.read()
        pdf_reader = pypdf.PdfReader(io.BytesIO(contents))
        total_pages = len(pdf_reader.pages)
        
        return {
            "success": True,
            "total_pages": total_pages,
            "filename": file.filename
        }
    except Exception as e:
        logger.error(f"Error reading PDF: {str(e)}")
        return {
            "success": False,
            "message": f"Erreur lors de la lecture du PDF: {str(e)}"
        }

# Get available program periods (for history)
@router.get("/periods", response_model=List[ProgramPeriod])
async def get_periods():
    """Récupère les périodes de programmes disponibles (pour l'historique)"""
    pipeline = [
        {"$group": {
            "_id": {"month": "$program_month", "year": "$program_year"},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id.year": -1, "_id.month": -1}}
    ]
    periods = await db.programs.aggregate(pipeline).to_list(100)
    return [ProgramPeriod(month=p["_id"]["month"], year=p["_id"]["year"], count=p["count"]) 
            for p in periods if p["_id"].get("month") is not None and p["_id"].get("year") is not None]


# ============ EXPORT / IMPORT EXCEL ============

@router.get("/programs/export-excel")
async def export_programs_excel(month: Optional[int] = None, year: Optional[int] = None):
    """Exporte tous les programmes en Excel pour correction manuelle"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")

    if month and year:
        query = {"program_month": month, "program_year": year}
    else:
        latest = await db.programs.find_one(sort=[("program_year", -1), ("program_month", -1)])
        if latest:
            query = {"program_month": latest.get("program_month"), "program_year": latest.get("program_year")}
        else:
            query = {}

    programs = await db.programs.find(query).sort([("year", -1), ("sort_order", 1), ("brand", 1), ("model", 1)]).to_list(1000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programmes"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "ID", "Marque", "Modele", "Trim", "Annee",
        "Consumer Cash ($)", "Bonus Cash ($)",
        "Opt1 36M", "Opt1 48M", "Opt1 60M", "Opt1 72M", "Opt1 84M", "Opt1 96M",
        "Rabais Alt. Cash ($)",
        "Opt2 36M", "Opt2 48M", "Opt2 60M", "Opt2 72M", "Opt2 84M", "Opt2 96M",
        "Sort Order"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14

    # Figer colonnes A-D et ligne 1
    ws.freeze_panes = 'E2'

    year_2026_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    year_2025_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    for row_idx, prog in enumerate(programs, 2):
        o1 = prog.get("option1_rates") or {}
        o2 = prog.get("option2_rates")
        row_fill = year_2026_fill if prog.get("year") == 2026 else year_2025_fill

        row_data = [
            prog.get("id", ""),
            prog.get("brand", ""),
            prog.get("model", ""),
            prog.get("trim", ""),
            prog.get("year", 2026),
            prog.get("consumer_cash", 0) or 0,
            prog.get("bonus_cash", 0) or 0,
            o1.get("rate_36"), o1.get("rate_48"), o1.get("rate_60"),
            o1.get("rate_72"), o1.get("rate_84"), o1.get("rate_96"),
            prog.get("alternative_consumer_cash", 0) or 0,
            o2.get("rate_36") if o2 else None, o2.get("rate_48") if o2 else None,
            o2.get("rate_60") if o2 else None, o2.get("rate_72") if o2 else None,
            o2.get("rate_84") if o2 else None, o2.get("rate_96") if o2 else None,
            prog.get("sort_order", 999)
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.fill = row_fill
            if col_idx >= 6:
                cell.alignment = center
            if 8 <= col_idx <= 13:
                cell.fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
            elif col_idx == 14:
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            elif 15 <= col_idx <= 20:
                cell.fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")

    ws2 = wb.create_sheet("Instructions")
    instructions = [
        "INSTRUCTIONS - CORRECTION DES PROGRAMMES",
        "",
        "1. Modifiez les valeurs dans l'onglet 'Programmes'",
        "2. Consumer Cash = rabais avant taxes Option 1",
        "3. Bonus Cash = rabais apres taxes (0 dans la plupart des cas)",
        "4. Les colonnes Opt1 = taux de financement Option 1 (ex: 4.99)",
        "5. Les colonnes Opt2 = taux de financement Option 2 (vide = pas d'option 2)",
        "6. NE PAS modifier la colonne ID",
        "7. NE PAS inclure les Delivery Credits ('E' Only)",
        "8. Sauvegardez et reimportez le fichier dans l'application",
        "",
        "IMPORTANT: Ce fichier corrige devient la SOURCE DE VERITE",
    ]
    for i, text in enumerate(instructions, 1):
        cell = ws2.cell(row=i, column=1, value=text)
        if i == 1:
            cell.font = Font(bold=True, size=14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=programmes_calcauto.xlsx"}
    )


@router.post("/programs/import-excel")
async def import_programs_excel(file: UploadFile = File(...), password: str = Form("")):
    """Importe un Excel corrige avec systeme de comparaison avant/apres.
    Utilise cle composite (brand+model+trim+year) pour le matching."""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")

    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe admin incorrect")

    # 1. SNAPSHOT AVANT: capturer l'etat actuel de tous les programmes
    all_before = {}
    async for prog in db.programs.find({}, {"_id": 0}):
        key = f"{prog.get('brand','')}|{prog.get('model','')}|{prog.get('trim','')}|{prog.get('year','')}"
        all_before[key] = prog

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Detect program_month from title row (e.g. "PROGRAMMES DE FINANCEMENT RETAIL - MARS 2026")
    import_month = None
    import_year = None
    title_val = str(ws.cell(row=1, column=1).value or '').upper()
    month_names_map = {"JANVIER": 1, "FÉVRIER": 2, "FEVRIER": 2, "MARS": 3, "AVRIL": 4, "MAI": 5, "JUIN": 6,
                       "JUILLET": 7, "AOÛT": 8, "AOUT": 8, "SEPTEMBRE": 9, "OCTOBRE": 10, "NOVEMBRE": 11, "DÉCEMBRE": 12, "DECEMBRE": 12}
    for mname, mnum in month_names_map.items():
        if mname in title_val:
            import_month = mnum
            break
    import re as re_mod
    year_match = re_mod.search(r'20\d{2}', title_val)
    if year_match:
        import_year = int(year_match.group())

    updated = 0
    unchanged = 0
    created = 0
    corrections_saved = 0
    errors = []
    rows_processed = 0
    comparison_details = []

    def parse_cash(val):
        """Parse '$3,500' or '-' or 3500 → float"""
        if val is None or str(val).strip() in ('-', '–', ''):
            return 0
        s = str(val).replace('$', '').replace(',', '').replace(' ', '').strip()
        try:
            return float(s)
        except ValueError:
            return 0

    def parse_rate(val):
        """Parse '4.99%' or '-' or 4.99 → float or None"""
        if val is None or str(val).strip() in ('-', '–', ''):
            return None
        s = str(val).replace('%', '').replace(' ', '').strip()
        try:
            return float(s)
        except ValueError:
            return None

    # Detect header row: find row with "Marque" in first column
    data_start_row = 2
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        first_val = str(row[0].value or '').strip().lower()
        if first_val in ('marque', 'brand'):
            data_start_row = row[0].row + 1
            break
        elif first_val and first_val not in ('', 'none') and 'programme' not in first_val and 'véhicule' not in first_val and 'option' not in first_val:
            # First row with actual data
            data_start_row = row[0].row
            break

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=False), data_start_row):
        values = [cell.value for cell in row]
        if not values or not values[0]:
            continue
        
        # Skip header-like rows
        first_val = str(values[0]).strip().lower()
        if first_val in ('marque', 'brand', 'véhicule', '') or 'programme' in first_val or 'option' in first_val:
            continue

        rows_processed += 1

        try:
            brand = str(values[0]).strip() if values[0] else ""
            model = str(values[1]).strip() if values[1] else ""
            trim = str(values[2]).strip() if len(values) > 2 and values[2] else ""
            
            year_val = values[3] if len(values) > 3 else 2026
            try:
                year = int(year_val) if year_val else 2026
            except (ValueError, TypeError):
                year = 2026

            # Col E (index 4): Consumer Cash (Option 1 rabais)
            consumer_cash = parse_cash(values[4]) if len(values) > 4 else 0
            
            # Col F-K (index 5-10): Option 1 rates
            o1_rates = {}
            rate_terms = ["rate_36", "rate_48", "rate_60", "rate_72", "rate_84", "rate_96"]
            has_o1 = False
            for i, term in enumerate(rate_terms):
                col_idx = 5 + i
                if col_idx < len(values):
                    rate = parse_rate(values[col_idx])
                    if rate is not None:
                        o1_rates[term] = rate
                        has_o1 = True
            
            # Col L (index 11): Alternative Consumer Cash (Option 2 rabais)
            alternative_consumer_cash = parse_cash(values[11]) if len(values) > 11 else 0
            
            # Col M-R (index 12-17): Option 2 rates
            o2_rates = None
            has_o2 = False
            o2_temp = {}
            for i, term in enumerate(rate_terms):
                col_idx = 12 + i
                if col_idx < len(values):
                    rate = parse_rate(values[col_idx])
                    if rate is not None:
                        o2_temp[term] = rate
                        has_o2 = True
            if has_o2:
                o2_rates = o2_temp
            
            # Col S (index 18): Bonus Cash
            bonus_cash = parse_cash(values[18]) if len(values) > 18 else 0

            # Match par cle composite (brand+model+trim+year) avec matching flexible
            old_prog, match_query = find_best_match(brand, model, trim, year, all_before)

            if not old_prog:
                # Programme non trouve = le creer depuis l'Excel (source de verite)
                new_prog = {
                    "id": str(uuid.uuid4()),
                    "brand": brand,
                    "model": model,
                    "trim": trim,
                    "year": year,
                    "consumer_cash": consumer_cash,
                    "bonus_cash": bonus_cash,
                    "alternative_consumer_cash": alternative_consumer_cash,
                    "option1_rates": o1_rates if has_o1 else {"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99},
                    "option2_rates": o2_rates,
                    "program_month": import_month,
                    "program_year": import_year or year,
                    "created_at": datetime.utcnow().isoformat(),
                    "source": "excel_import"
                }
                await db.programs.insert_one(new_prog)
                created += 1
                continue

            # Calculer les differences AVANT d'appliquer
            changes = {}
            if (old_prog.get("consumer_cash") or 0) != consumer_cash:
                changes["consumer_cash"] = {"avant": old_prog.get("consumer_cash", 0) or 0, "apres": consumer_cash}
            if (old_prog.get("alternative_consumer_cash") or 0) != alternative_consumer_cash:
                changes["alternative_consumer_cash"] = {"avant": old_prog.get("alternative_consumer_cash", 0) or 0, "apres": alternative_consumer_cash}
            if (old_prog.get("bonus_cash") or 0) != bonus_cash:
                changes["bonus_cash"] = {"avant": old_prog.get("bonus_cash", 0) or 0, "apres": bonus_cash}
            old_o1 = old_prog.get("option1_rates") or {}
            if has_o1 and old_o1 != o1_rates:
                changes["option1_rates"] = {"avant": old_o1, "apres": o1_rates}
            old_o2 = old_prog.get("option2_rates")
            if old_o2 != o2_rates:
                changes["option2_rates"] = {"avant": old_o2, "apres": o2_rates}

            update_fields = {
                "consumer_cash": consumer_cash,
                "bonus_cash": bonus_cash,
                "alternative_consumer_cash": alternative_consumer_cash,
                "updated_at": datetime.utcnow()
            }
            if has_o1:
                update_fields["option1_rates"] = o1_rates
            if has_o2:
                update_fields["option2_rates"] = o2_rates

            update_ops = {"$set": update_fields}
            if not has_o2:
                update_ops["$unset"] = {"option2_rates": ""}

            result = await db.programs.update_one(match_query, update_ops)

            if changes:
                updated += 1
                comparison_details.append({
                    "vehicule": f"{brand} {model} {trim} {year}",
                    "changes": changes
                })

                # Memoriser les corrections
                await db.program_corrections.update_one(
                    {"brand": brand, "model": model, "trim": trim, "year": year},
                    {"$set": {
                        "brand": brand, "model": model, "trim": trim, "year": year,
                        "corrected_values": {
                            "consumer_cash": consumer_cash,
                            "alternative_consumer_cash": alternative_consumer_cash,
                            "bonus_cash": bonus_cash,
                            "option1_rates": o1_rates if has_o1 else None,
                            "option2_rates": o2_rates,
                        },
                        "changes_history": changes,
                        "corrected_at": datetime.utcnow()
                    }},
                    upsert=True
                )
                corrections_saved += 1
            else:
                unchanged += 1

        except Exception as e:
            errors.append(f"Ligne {row_idx}: {str(e)}")

    # 2. Sauvegarder la comparaison dans MongoDB
    comparison_id = str(uuid.uuid4())
    comparison_doc = {
        "id": comparison_id,
        "type": "programs",
        "date": datetime.utcnow().isoformat(),
        "rows_processed": rows_processed,
        "updated": updated,
        "unchanged": unchanged,
        "corrections_saved": corrections_saved,
        "errors": errors[:20],
        "details": comparison_details
    }
    await db.import_comparisons.insert_one(comparison_doc)

    # Force logout
    await db.tokens.delete_many({})
    logger.info(f"[EXCEL IMPORT] {updated} modifies, {unchanged} inchanges, {created} crees, {len(errors)} erreurs")

    return {
        "success": True,
        "message": f"Import termine: {updated} modifies, {unchanged} inchanges, {created} crees, {corrections_saved} corrections memorisees.",
        "comparison_id": comparison_id,
        "updated": updated,
        "created": created,
        "unchanged": unchanged,
        "rows_processed": rows_processed,
        "corrections_saved": corrections_saved,
        "errors": errors[:10],
        "comparison": comparison_details[:50]
    }


@router.get("/programs/comparisons")
async def get_import_comparisons():
    """Retourne l'historique des comparaisons d'imports"""
    comparisons = await db.import_comparisons.find(
        {"type": "programs"},
        {"_id": 0}
    ).sort("date", -1).to_list(20)
    return comparisons


@router.get("/programs/comparison/{comparison_id}")
async def get_comparison_detail(comparison_id: str):
    """Retourne le detail d'une comparaison specifique"""
    comp = await db.import_comparisons.find_one(
        {"id": comparison_id},
        {"_id": 0}
    )
    if not comp:
        raise HTTPException(status_code=404, detail="Comparaison non trouvee")
    return comp



# Programs CRUD
@router.put("/programs/reorder")
async def reorder_programs(data: Dict[str, Any]):
    """Réordonne les programmes - data = {password, orders: [{id, sort_order}]}"""
    password = data.get("password", "")
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")

    orders = data.get("orders", [])
    updated = 0
    for item in orders:
        prog_id = item.get("id")
        new_sort_order = item.get("sort_order")
        if prog_id is not None and new_sort_order is not None:
            result = await db.programs.update_one(
                {"id": prog_id},
                {"$set": {"sort_order": new_sort_order}}
            )
            if result.modified_count > 0:
                updated += 1

    return {"message": f"Réordonné {updated} programmes", "updated": updated}


@router.post("/programs", response_model=VehicleProgram)
async def create_program(program: VehicleProgramCreate):
    program_dict = program.dict()
    if program_dict.get("program_month") is None:
        program_dict["program_month"] = datetime.utcnow().month
    if program_dict.get("program_year") is None:
        program_dict["program_year"] = datetime.utcnow().year
    program_obj = VehicleProgram(**program_dict)
    await db.programs.insert_one(program_obj.dict())
    return program_obj

@router.get("/programs", response_model=List[VehicleProgram])
async def get_programs(month: Optional[int] = None, year: Optional[int] = None):
    """
    Récupère les programmes de financement
    Si month/year sont fournis, filtre par période
    Sinon, retourne la période la plus récente
    """
    if month and year:
        query = {"program_month": month, "program_year": year}
    else:
        # Trouver la période la plus récente
        latest = await db.programs.find_one(sort=[("program_year", -1), ("program_month", -1)])
        if latest:
            query = {"program_month": latest.get("program_month"), "program_year": latest.get("program_year")}
        else:
            query = {}
    
    programs = await db.programs.find(query).sort([("sort_order", 1), ("year", -1), ("model", 1), ("trim", 1)]).to_list(1000)
    return [VehicleProgram(**p) for p in programs]

@router.get("/programs/{program_id}", response_model=VehicleProgram)
async def get_program(program_id: str):
    program = await db.programs.find_one({"id": program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    return VehicleProgram(**program)

@router.put("/programs/{program_id}", response_model=VehicleProgram)
async def update_program(program_id: str, update: VehicleProgramUpdate):
    program = await db.programs.find_one({"id": program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    update_data = {k: v for k, v in update.dict(exclude_unset=True).items()}
    update_data["updated_at"] = datetime.utcnow()
    
    # Handle explicitly setting option2_rates to null
    unset_fields = {}
    if "option2_rates" in update_data and update_data["option2_rates"] is None:
        del update_data["option2_rates"]
        unset_fields["option2_rates"] = ""
    
    update_ops = {"$set": update_data}
    if unset_fields:
        update_ops["$unset"] = unset_fields
    
    await db.programs.update_one({"id": program_id}, update_ops)
    updated = await db.programs.find_one({"id": program_id})
    return VehicleProgram(**updated)

@router.delete("/programs/{program_id}")
async def delete_program(program_id: str):
    result = await db.programs.delete_one({"id": program_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Program not found")
    return {"message": "Program deleted successfully"}

# Import programs with password protection
@router.post("/import")
async def import_programs(request: ImportRequest):
    """
    Importe des programmes de financement (protégé par mot de passe)
    Remplace tous les programmes pour le mois/année spécifié
    """
    if request.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    
    # Supprimer les programmes existants pour cette période
    await db.programs.delete_many({
        "program_month": request.program_month,
        "program_year": request.program_year
    })
    
    # Insérer les nouveaux programmes
    inserted = 0
    for prog_data in request.programs:
        # Assurer que les taux sont au bon format
        if prog_data.get("option1_rates") and isinstance(prog_data["option1_rates"], dict):
            prog_data["option1_rates"] = FinancingRates(**prog_data["option1_rates"]).dict()
        if prog_data.get("option2_rates") and isinstance(prog_data["option2_rates"], dict):
            prog_data["option2_rates"] = FinancingRates(**prog_data["option2_rates"]).dict()
        
        prog_data["program_month"] = request.program_month
        prog_data["program_year"] = request.program_year
        prog_data["bonus_cash"] = prog_data.get("bonus_cash", 0)
        
        # Compute sort_order from trim_orders collection
        sort_order = await compute_sort_order(
            prog_data.get("brand", ""),
            prog_data.get("model", ""),
            prog_data.get("trim"),
            prog_data.get("year", 2026)
        )
        prog_data["sort_order"] = sort_order
        
        prog = VehicleProgram(**prog_data)
        await db.programs.insert_one(prog.dict())
        inserted += 1
    
    return {"message": f"Importé {inserted} programmes pour {request.program_month}/{request.program_year}"}

# Calculate financing options
@router.post("/calculate", response_model=CalculationResponse)
async def calculate_financing(request: CalculationRequest):
    """
    Calcule et compare les options de financement
    
    Option 1: Prix - Consumer Cash (rabais avant taxes), avec taux Option 1
    Option 2: Prix complet, avec taux réduits (si disponible)
    
    Note: Bonus Cash est affiché mais non inclus dans le calcul du financement
    (car appliqué après taxes, comme comptant)
    """
    vehicle_price = request.vehicle_price
    
    if not request.program_id:
        raise HTTPException(status_code=400, detail="Program ID is required")
    
    program = await db.programs.find_one({"id": request.program_id})
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    
    program_obj = VehicleProgram(**program)
    consumer_cash = program_obj.consumer_cash
    bonus_cash = program_obj.bonus_cash
    
    comparisons = []
    terms = [36, 48, 60, 72, 84, 96]
    
    for term in terms:
        # Option 1: Avec Consumer Cash (rabais avant taxes) + taux Option 1
        option1_rate = get_rate_for_term(program_obj.option1_rates, term)
        principal1 = vehicle_price - consumer_cash  # Rabais avant taxes
        monthly1 = calculate_monthly_payment(principal1, option1_rate, term)
        total1 = round(monthly1 * term, 2)
        
        comparison = PaymentComparison(
            term_months=term,
            option1_rate=option1_rate,
            option1_monthly=monthly1,
            option1_total=total1,
            option1_rebate=consumer_cash
        )
        
        # Option 2: Sans rabais + taux réduits (si disponible)
        if program_obj.option2_rates:
            option2_rate = get_rate_for_term(program_obj.option2_rates, term)
            principal2 = vehicle_price  # Pas de rabais
            monthly2 = calculate_monthly_payment(principal2, option2_rate, term)
            total2 = round(monthly2 * term, 2)
            
            comparison.option2_rate = option2_rate
            comparison.option2_monthly = monthly2
            comparison.option2_total = total2
            
            # Déterminer la meilleure option (coût total le plus bas)
            if total1 < total2:
                comparison.best_option = "1"
                comparison.savings = round(total2 - total1, 2)
            elif total2 < total1:
                comparison.best_option = "2"
                comparison.savings = round(total1 - total2, 2)
            else:
                comparison.best_option = "1"
                comparison.savings = 0
        
        comparisons.append(comparison)
    
    return CalculationResponse(
        vehicle_price=vehicle_price,
        consumer_cash=consumer_cash,
        bonus_cash=bonus_cash,
        brand=program_obj.brand,
        model=program_obj.model,
        trim=program_obj.trim,
        year=program_obj.year,
        comparisons=comparisons
    )

# Seed initial data from PDF pages 20-21 (Février 2026)
@router.post("/seed")
async def seed_data():
    """
    Seed les données initiales à partir du PDF Février 2026
    Pages 20 (2026) et 21 (2025)
    """
    # Clear existing data
    await db.programs.delete_many({})
    
    # Taux standard 4.99% pour la plupart des véhicules
    std = {"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}
    
    # Mois/Année du programme
    prog_month = 2
    prog_year = 2026
    
    programs_data = [
        # ==================== 2026 MODELS (Page 20) ====================
        
        # CHRYSLER 2026
        {"brand": "Chrysler", "model": "Grand Caravan", "trim": "SXT", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49}, 
         "bonus_cash": 0},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "PHEV", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49}, 
         "bonus_cash": 0},
        
        # Pacifica (excluding PHEV): Taux bas (0% jusqu'à 72mo, 1.99% à 84mo, 3.49% à 96mo), PAS d'Option 2
        {"brand": "Chrysler", "model": "Pacifica", "trim": "(excluding PHEV)", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 3.49},
         "option2_rates": None, "bonus_cash": 0},
        
        # JEEP COMPASS 2026 - Taux Option 2 corrigés
        {"brand": "Jeep", "model": "Compass", "trim": "Sport", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49}, 
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "North", "year": 2026,
         "consumer_cash": 3500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "North w/ Altitude Package (ADZ)", "year": 2026,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Trailhawk", "year": 2026,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Limited", "year": 2026,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        # JEEP CHEROKEE 2026
        {"brand": "Jeep", "model": "Cherokee", "trim": "Base (KMJL74)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Cherokee", "trim": "(excluding Base)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # JEEP WRANGLER 2026 - Corrigé avec Option 2
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door (JL) non Rubicon", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door Rubicon (JL)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99}, 
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (excl. 392 et 4xe)", "year": 2026,
         "consumer_cash": 5250, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99}, 
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door MOAB 392", "year": 2026,
         "consumer_cash": 6000, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # JEEP GLADIATOR 2026
        {"brand": "Jeep", "model": "Gladiator", "trim": "Sport S, Willys, Sahara, Willys '41", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Gladiator", "trim": "(excl. Sport S, Willys, Sahara, Willys '41)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49}, 
         "bonus_cash": 0},
        
        # JEEP GRAND CHEROKEE 2026 - Taux variables Option 1 et Option 2 corrigés
        {"brand": "Jeep", "model": "Grand Cherokee/L", "trim": "Laredo/Laredo X", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 1.99, "rate_48": 2.99, "rate_60": 3.49, "rate_72": 3.99, "rate_84": 4.49, "rate_96": 4.99},
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Grand Cherokee/L", "trim": "Altitude", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 1.99, "rate_48": 2.99, "rate_60": 3.49, "rate_72": 3.99, "rate_84": 4.49, "rate_96": 4.99},
         "option2_rates": None,
         "bonus_cash": 0},
        
        {"brand": "Jeep", "model": "Grand Cherokee/L", "trim": "Limited/Limited Reserve/Summit", "year": 2026,
         "consumer_cash": 0,
         "option1_rates": {"rate_36": 1.99, "rate_48": 2.99, "rate_60": 3.49, "rate_72": 3.99, "rate_84": 4.49, "rate_96": 4.99},
         "option2_rates": None,
         "bonus_cash": 0},
        
        # JEEP GRAND WAGONEER 2026
        {"brand": "Jeep", "model": "Grand Wagoneer/L", "trim": None, "year": 2026,
         "consumer_cash": 0, "option1_rates": std, 
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.99}, 
         "bonus_cash": 0},
        
        # DODGE DURANGO 2026
        {"brand": "Dodge", "model": "Durango", "trim": "SXT, GT, GT Plus", "year": 2026,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Dodge", "model": "Durango", "trim": "GT Hemi V8 Plus, GT Hemi V8 Premium", "year": 2026,
         "consumer_cash": 9000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Dodge", "model": "Durango", "trim": "SRT Hellcat", "year": 2026,
         "consumer_cash": 15500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.49, "rate_84": 2.49, "rate_96": 3.49},
         "bonus_cash": 0},
        
        # DODGE CHARGER 2026
        {"brand": "Dodge", "model": "Charger", "trim": "2-Door & 4-Door (ICE)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # RAM 2026 - Corrigé
        {"brand": "Ram", "model": "ProMaster", "trim": None, "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Tradesman, Express, Warlock", "year": 2026,
         "consumer_cash": 6500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Big Horn", "year": 2026,
         "consumer_cash": 6000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Sport, Rebel", "year": 2026,
         "consumer_cash": 8250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Laramie (DT6P98)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Laramie, Limited, Longhorn, Tungsten, RHO (excl. DT6P98)", "year": 2026,
         "consumer_cash": 11500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500 Power Wagon Crew Cab", "trim": "(DJ7X91 2UP)", "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "Gas Models", "year": 2026,
         "consumer_cash": 7000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 1.99, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "Diesel Models", "year": 2026,
         "consumer_cash": 5000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 0.99, "rate_60": 0.99, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "Chassis Cab", "trim": None, "year": 2026,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # ==================== 2025 MODELS (Page 21) ====================
        
        # CHRYSLER 2025 - Corrigé
        {"brand": "Chrysler", "model": "Grand Caravan", "trim": "SXT", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "Hybrid", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 1.99, "rate_72": 2.99, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "Select Models (excl. Hybrid)", "year": 2025,
         "consumer_cash": 0, 
         "option1_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 1.99, "rate_72": 2.99, "rate_84": 3.99, "rate_96": 4.99},
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        {"brand": "Chrysler", "model": "Pacifica", "trim": "(excl. Select & Hybrid)", "year": 2025,
         "consumer_cash": 750, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        # JEEP COMPASS 2025 - Corrigé
        {"brand": "Jeep", "model": "Compass", "trim": "Sport", "year": 2025,
         "consumer_cash": 5500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Compass", "trim": "North", "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0.99, "rate_60": 1.99, "rate_72": 1.99, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Altitude, Trailhawk, Trailhawk Elite", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Compass", "trim": "Limited", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.49, "rate_96": 2.49},
         "bonus_cash": 1000},
        
        # JEEP WRANGLER 2025 - Corrigé
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (JL) 4xe (JLXL74)", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (JL) 4xe (excl. JLXL74)", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door (JL) non Rubicon", "year": 2025,
         "consumer_cash": 750, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "2-Door Rubicon (JL)", "year": 2025,
         "consumer_cash": 8500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door Rubicon w/ 2.0L", "year": 2025,
         "consumer_cash": 0, 
         "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wrangler", "trim": "4-Door (JL) (excl. Rubicon 2.0L & 4xe)", "year": 2025,
         "consumer_cash": 8500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        # JEEP GLADIATOR 2025
        {"brand": "Jeep", "model": "Gladiator", "trim": None, "year": 2025,
         "consumer_cash": 11000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        # JEEP GRAND CHEROKEE 2025 - Corrigé
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "4xe (WL)", "year": 2025,
         "consumer_cash": 4000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 1.99, "rate_60": 2.49, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Laredo (WLJH74 2*A)", "year": 2025,
         "consumer_cash": 6000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Altitude (WLJH74 2*B)", "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "Summit (WLJT74 23S)", "year": 2025,
         "consumer_cash": 0, 
         "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee", "trim": "(WL) (excl. Laredo, Altitude, Summit, 4xe)", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        # JEEP GRAND CHEROKEE L 2025 - Corrigé
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Laredo (WLJH75 2*A)", "year": 2025,
         "consumer_cash": 6000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Altitude (WLJH75 2*B)", "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "Overland (WLJS75)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Cherokee L", "trim": "(WL) (excl. Laredo, Altitude, Overland)", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        # JEEP WAGONEER 2025 - Corrigé
        {"brand": "Jeep", "model": "Wagoneer/L", "trim": None, "year": 2025,
         "consumer_cash": 7500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Grand Wagoneer/L", "trim": None, "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 3.99},
         "bonus_cash": 1000},
        
        {"brand": "Jeep", "model": "Wagoneer S", "trim": "Limited & Premium (BEV)", "year": 2025,
         "consumer_cash": 8000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        # DODGE DURANGO 2025 - Corrigé
        {"brand": "Dodge", "model": "Durango", "trim": "GT, GT Plus", "year": 2025,
         "consumer_cash": 8000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Durango", "trim": "R/T, R/T Plus, R/T 20th Anniversary", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Durango", "trim": "SRT Hellcat", "year": 2025,
         "consumer_cash": 16000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 2.99, "rate_96": 3.49},
         "bonus_cash": 1000},
        
        # DODGE CHARGER 2025 - Corrigé
        {"brand": "Dodge", "model": "Charger Daytona", "trim": "R/T (BEV)", "year": 2025,
         "consumer_cash": 3000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Charger Daytona", "trim": "R/T Plus (BEV)", "year": 2025,
         "consumer_cash": 5000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Charger Daytona", "trim": "Scat Pack (BEV)", "year": 2025,
         "consumer_cash": 7000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 1000},
        
        # DODGE HORNET 2025
        {"brand": "Dodge", "model": "Hornet", "trim": "RT (PHEV)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Hornet", "trim": "RT Plus (PHEV)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Hornet", "trim": "GT (Gas)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        {"brand": "Dodge", "model": "Hornet", "trim": "GT Plus (Gas)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 1000},
        
        # RAM 2025 - Corrigé
        {"brand": "Ram", "model": "ProMaster", "trim": None, "year": 2025,
         "consumer_cash": 0, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "1500", "trim": "Tradesman, Warlock, Express (DT)", "year": 2025,
         "consumer_cash": 9250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Big Horn (DT) w/ Off-Roader Value Package (4KF)", "year": 2025,
         "consumer_cash": 0, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Big Horn (DT) (excl. Off-Roader)", "year": 2025,
         "consumer_cash": 9250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Sport, Rebel (DT)", "year": 2025,
         "consumer_cash": 10000, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "1500", "trim": "Laramie, Limited, Longhorn, Tungsten, RHO (DT)", "year": 2025,
         "consumer_cash": 12250, "option1_rates": std,
         "option2_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 0, "rate_72": 0, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 3000},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "Gas Models (excl. Chassis Cab, Diesel)", "year": 2025,
         "consumer_cash": 9500, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        {"brand": "Ram", "model": "2500/3500", "trim": "6.7L High Output Diesel (ETM)", "year": 2025,
         "consumer_cash": 7000, "option1_rates": std,
         "option2_rates": {"rate_36": 0.99, "rate_48": 0.99, "rate_60": 0.99, "rate_72": 0.99, "rate_84": 1.99, "rate_96": 2.99},
         "bonus_cash": 0},
        
        {"brand": "Ram", "model": "Chassis Cab", "trim": None, "year": 2025,
         "consumer_cash": 5000, "option1_rates": std, "option2_rates": None, "bonus_cash": 0},
        
        # FIAT 2025
        {"brand": "Fiat", "model": "500e", "trim": "BEV", "year": 2025,
         "consumer_cash": 6000,
         "option1_rates": {"rate_36": 0, "rate_48": 0, "rate_60": 1.99, "rate_72": 3.49, "rate_84": 3.99, "rate_96": 4.99},
         "option2_rates": None, "bonus_cash": 5000},
    ]
    
    for prog_data in programs_data:
        if prog_data.get("option2_rates") and isinstance(prog_data["option2_rates"], dict):
            prog_data["option2_rates"] = FinancingRates(**prog_data["option2_rates"]).dict()
        if isinstance(prog_data["option1_rates"], dict):
            prog_data["option1_rates"] = FinancingRates(**prog_data["option1_rates"]).dict()
        prog_data["program_month"] = prog_month
        prog_data["program_year"] = prog_year
        prog = VehicleProgram(**prog_data)
        await db.programs.insert_one(prog.dict())
    
    return {"message": f"Seeded {len(programs_data)} programs for {prog_month}/{prog_year}"}



# ============ Trim Order Management ============

@router.get("/trim-orders")
async def get_trim_orders():
    """Récupère les ordres de tri des versions (trims) stockés en MongoDB."""
    orders = await db.trim_orders.find({}, {"_id": 0}).to_list(200)
    return orders


@router.post("/trim-orders/recalculate")
async def recalculate_sort_orders(password: str = ""):
    """Recalcule le sort_order de tous les programmes à partir des trim_orders stockés."""
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")

    programs = await db.programs.find({}).to_list(2000)
    updated = 0
    for prog in programs:
        sort_order = await compute_sort_order(
            prog.get("brand", ""),
            prog.get("model", ""),
            prog.get("trim"),
            prog.get("year", 2026)
        )
        await db.programs.update_one(
            {"_id": prog["_id"]},
            {"$set": {"sort_order": sort_order}}
        )
        updated += 1

    return {"message": f"Recalculé sort_order pour {updated} programmes"}
