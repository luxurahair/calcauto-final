from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from database import db, ROOT_DIR, logger
import json
import io

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter()

# ============ HELPER: Find latest data file ============

def _get_latest_data_file(prefix: str, target_month: int = None, target_year: int = None) -> str:
    """Find a specific or the most recent data file matching prefix.
    If target_month/year are provided, find that specific month's file.
    Otherwise, find the most recent."""
    import glob, re
    data_dir = ROOT_DIR / "data"
    
    month_order_en = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                      "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
    month_order_fr = {"janvier": 1, "février": 2, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
                      "juillet": 7, "août": 8, "aout": 8, "septembre": 9, "octobre": 10, "novembre": 11, "décembre": 12, "decembre": 12}
    
    all_months = {**month_order_en, **month_order_fr}
    
    best_file = None
    best_score = (0, 0)
    target_file = None
    
    pattern = str(data_dir / f"{prefix}_*.json")
    for filepath in glob.glob(pattern):
        fname = filepath.split("/")[-1]
        part = fname.replace(f"{prefix}_", "").replace(".json", "")
        year_match = re.search(r'(\d{4})$', part)
        if not year_match:
            continue
        year = int(year_match.group(1))
        month_part = part[:year_match.start()].lower().strip()
        month_num = all_months.get(month_part, 0)
        
        # If targeting a specific month, check for exact match
        if target_month and target_year:
            if year == target_year and month_num == target_month:
                target_file = filepath
        
        score = (year, month_num)
        if score > best_score:
            best_score = score
            best_file = filepath
    
    if target_month and target_year:
        return target_file  # None if not found for that specific month
    
    return best_file


# ============ SCI LEASE ENDPOINTS ============

@router.get("/sci/residuals")
async def get_sci_residuals(month: int = None, year: int = None):
    """Retourne les valeurs résiduelles SCI. month/year optionnels pour historique.
    Fusionne les km_adjustments dynamiques si disponibles."""
    residuals_path = _get_latest_data_file("sci_residuals", month, year)
    if not residuals_path:
        raise HTTPException(status_code=404, detail="Residual data not found" + (f" for {month}/{year}" if month else ""))
    with open(residuals_path, 'r') as f:
        data = json.load(f)

    # Override km_adjustments with the latest dynamic file if available
    km_path = _get_latest_data_file("km_adjustments", month, year)
    if km_path:
        try:
            with open(km_path, 'r') as f:
                km_data = json.load(f)
            data['km_adjustments'] = km_data
        except Exception:
            pass  # Keep existing km_adjustments from residuals file

    return data

@router.get("/sci/lease-rates")
async def get_sci_lease_rates(month: int = None, year: int = None):
    """Retourne les taux de location SCI et lease cash. month/year optionnels pour historique."""
    rates_path = _get_latest_data_file("sci_lease_rates", month, year)
    if not rates_path:
        raise HTTPException(status_code=404, detail="Lease rates data not found" + (f" for {month}/{year}" if month else ""))
    with open(rates_path, 'r') as f:
        data = json.load(f)
    return data

@router.get("/sci/vehicle-hierarchy")
async def get_sci_vehicle_hierarchy():
    """Retourne la hiérarchie des véhicules SCI: marque -> modèle -> trim -> body_style"""
    residuals_path = _get_latest_data_file("sci_residuals")
    if not residuals_path:
        raise HTTPException(status_code=404, detail="Residual data not found")
    with open(residuals_path, 'r') as f:
        data = json.load(f)
    
    hierarchy = {}
    for v in data.get("vehicles", []):
        brand = v["brand"]
        model = v["model_name"]
        trim = v.get("trim", "")
        body = v.get("body_style", "")
        year = v.get("model_year", 2026)
        
        if brand not in hierarchy:
            hierarchy[brand] = {}
        if model not in hierarchy[brand]:
            hierarchy[brand][model] = {"years": set(), "trims": {}}
        hierarchy[brand][model]["years"].add(year)
        if trim not in hierarchy[brand][model]["trims"]:
            hierarchy[brand][model]["trims"][trim] = []
        if body and body not in hierarchy[brand][model]["trims"][trim]:
            hierarchy[brand][model]["trims"][trim].append(body)
    
    # Convert sets to sorted lists
    result = {}
    for brand, models in hierarchy.items():
        result[brand] = {}
        for model, info in models.items():
            result[brand][model] = {
                "years": sorted(list(info["years"]), reverse=True),
                "trims": info["trims"]
            }
    
    return result

@router.post("/sci/calculate-lease")
async def calculate_lease(payload: dict):
    """
    Calcul de location SCI Quebec — Formule exacte (annuite en avance).
    
    Methode:
    1. Cap cost = prix + frais dossier - lease cash
    2. Net cap cost = cap + solde + du_echange - val_echange - comptant - bonus
    3. PMT arriere = (ncc * mr * factor - residuel * mr) / (factor - 1)
    4. PMT avance = PMT arriere / (1 + mr)
    5. Taxes QC SUR le paiement (5% TPS + 9.975% TVQ)
    6. Credit taxe echange reparti sur les paiements
    """
    import math
    
    try:
        msrp = float(payload.get("msrp", 0))
        selling_price = float(payload.get("selling_price", 0))
        term = int(payload.get("term", 36))
        annual_rate = float(payload.get("annual_rate", 0))
        residual_pct = float(payload.get("residual_pct", 0))
        km_per_year = int(payload.get("km_per_year", 24000))
        lease_cash = float(payload.get("lease_cash", 0))
        bonus_cash = float(payload.get("bonus_cash", 0))
        cash_down = float(payload.get("cash_down", 0))
        trade_value = float(payload.get("trade_value", 0))
        trade_owed = float(payload.get("trade_owed", 0))
        frais_dossier = float(payload.get("frais_dossier", 259.95))
        solde_reporte = float(payload.get("solde_reporte", 0))
        rabais_concess = float(payload.get("rabais_concess", 0))
        accessoires = float(payload.get("accessoires", 0))
        
        if msrp <= 0 or selling_price <= 0 or term <= 0:
            raise HTTPException(status_code=400, detail="Invalid input values")
        
        # Constantes fiscales QC
        TPS = 0.05
        TVQ = 0.09975
        TAUX_TAXE = TPS + TVQ
        
        # Ajustement km
        residuals_path = _get_latest_data_file("sci_residuals")
        km_adj = 0
        if residuals_path:
            with open(residuals_path, 'r') as f:
                res_data = json.load(f)
            adjustments = res_data.get("km_adjustments", {}).get("adjustments", {})
            km_key = str(km_per_year)
            term_key = str(term)
            if km_key in adjustments and term_key in adjustments[km_key]:
                km_adj = adjustments[km_key][term_key]
        
        # 1. Residuel ajuste
        adjusted_residual_pct = residual_pct + km_adj
        residual_value = msrp * (adjusted_residual_pct / 100)
        
        # 2. Cout capitalise
        sp = selling_price + accessoires - rabais_concess
        cap_cost = sp + frais_dossier - lease_cash
        
        # 3. Solde reporte
        solde_net = 0
        if solde_reporte < 0:
            solde_net = abs(solde_reporte) * (1 + TAUX_TAXE)
        elif solde_reporte > 0:
            solde_net = solde_reporte
        
        # 4. Net cap cost
        net_cap_cost = cap_cost + solde_net + trade_owed - trade_value - cash_down - bonus_cash
        
        # 5. PMT en avance (formule SCI exacte)
        monthly_rate = annual_rate / 100 / 12
        
        if monthly_rate == 0:
            monthly_before_tax = (net_cap_cost - residual_value) / term
            finance_charge = 0
        else:
            factor = math.pow(1 + monthly_rate, term)
            pmt_arrears = (net_cap_cost * monthly_rate * factor - residual_value * monthly_rate) / (factor - 1)
            monthly_before_tax = pmt_arrears / (1 + monthly_rate)
            finance_charge = monthly_before_tax - (net_cap_cost - residual_value) / term
        
        # 6. Taxes SUR le paiement
        tps_on_payment = monthly_before_tax * TPS
        tvq_on_payment = monthly_before_tax * TVQ
        taxes_mensuelles = tps_on_payment + tvq_on_payment
        
        # 7. Credit taxe echange
        credit_taxe = 0
        credit_perdu = 0
        if trade_value > 0:
            depreciation_trade = trade_value / term
            credit_potentiel = depreciation_trade * TAUX_TAXE
            credit_taxe = min(credit_potentiel, taxes_mensuelles)
            credit_perdu = max(0, credit_potentiel - taxes_mensuelles)
        
        # 8. Paiement final
        monthly_payment = max(0, monthly_before_tax + taxes_mensuelles - credit_taxe)
        biweekly_payment = monthly_payment * 12 / 26
        weekly_payment = monthly_payment * 12 / 52
        total_cost = monthly_payment * term
        cout_emprunt = finance_charge * term
        
        return {
            "success": True,
            "msrp": msrp,
            "selling_price": selling_price,
            "lease_cash": lease_cash,
            "bonus_cash": bonus_cash,
            "residual_pct": round(adjusted_residual_pct, 2),
            "residual_value": round(residual_value, 2),
            "km_adjustment": km_adj,
            "annual_rate": annual_rate,
            "term": term,
            "cap_cost": round(cap_cost, 2),
            "net_cap_cost": round(net_cap_cost, 2),
            "monthly_before_tax": round(monthly_before_tax, 2),
            "tps_on_payment": round(tps_on_payment, 2),
            "tvq_on_payment": round(tvq_on_payment, 2),
            "credit_taxe_echange": round(credit_taxe, 2),
            "credit_perdu": round(credit_perdu, 2),
            "monthly_payment": round(monthly_payment, 2),
            "biweekly_payment": round(biweekly_payment, 2),
            "weekly_payment": round(weekly_payment, 2),
            "total_lease_cost": round(total_cost, 2),
            "cout_emprunt": round(cout_emprunt, 2),
            "cash_down": cash_down,
            "trade_value": trade_value,
            "trade_owed": trade_owed,
            "frais_dossier": frais_dossier,
            "solde_reporte": solde_reporte,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lease calculation error: {str(e)}")



# ============ LEASE SCI EXCEL EXPORT/IMPORT ============

@router.get("/sci/export-excel")
async def export_sci_lease_excel():
    """Exporte les taux de location SCI en Excel"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")

    rates_path = _get_latest_data_file("sci_lease_rates")
    if not rates_path:
        raise HTTPException(status_code=404, detail="Fichier de taux SCI introuvable")

    with open(rates_path, 'r') as f:
        data = json.load(f)

    terms = data.get("terms", [24, 27, 36, 39, 42, 48, 51, 54, 60])
    wb = openpyxl.Workbook()

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    std_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    alt_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    cash_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    for year_key, sheet_title in [("vehicles_2026", "Lease 2026"), ("vehicles_2025", "Lease 2025")]:
        vehicles = data.get(year_key, [])
        if not vehicles:
            continue

        ws = wb.create_sheet(title=sheet_title) if wb.sheetnames != ["Sheet"] else wb.active
        if ws.title == "Sheet":
            ws.title = sheet_title

        headers = ["Marque", "Modele", "Lease Cash ($)"]
        for t in terms:
            headers.append(f"Std {t}M")
        headers.append("Alt Lease Cash ($)")
        for t in terms:
            headers.append(f"Alt {t}M")

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hfont
            c.fill = hfill
            c.alignment = center
            c.border = border

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 55

        # Figer colonnes A-B et ligne 1
        ws.freeze_panes = 'C2'

        for ri, v in enumerate(vehicles, 2):
            std = v.get("standard_rates") or {}
            alt = v.get("alternative_rates") or {}

            row = [v.get("brand", ""), v.get("model", ""), v.get("lease_cash", 0) or 0]
            for t in terms:
                row.append(std.get(str(t)))
            row.append(v.get("alternative_lease_cash", 0) or 0)
            for t in terms:
                row.append(alt.get(str(t)))

            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = border
                if ci >= 3:
                    c.alignment = center
                if ci == 3:
                    c.fill = cash_fill
                elif 4 <= ci <= 3 + len(terms):
                    c.fill = std_fill
                elif ci == 4 + len(terms):
                    c.fill = cash_fill
                elif ci > 4 + len(terms):
                    c.fill = alt_fill

    # Instructions
    ws2 = wb.create_sheet("Instructions")
    for i, t in enumerate([
        "INSTRUCTIONS - TAUX DE LOCATION SCI",
        "",
        "1. Modifiez les taux dans les onglets Lease 2026 et Lease 2025",
        "2. Lease Cash = rabais avant taxes pour location standard",
        "3. Std = taux de location standard SCI",
        "4. Alt Lease Cash = rabais avant taxes pour location alternative",
        "5. Alt = taux de location alternative SCI",
        "6. Vide = pas de taux disponible pour ce terme",
        "7. Sauvegardez et reimportez dans l'application",
    ], 1):
        c = ws2.cell(row=i, column=1, value=t)
        if i == 1: c.font = Font(bold=True, size=14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sci_lease_rates.xlsx"}
    )


@router.post("/sci/import-excel")
async def import_sci_lease_excel(file: UploadFile = File(...), password: str = Form(""), program_month: int = Form(0), program_year: int = Form(0)):
    """Importe un Excel corrigé pour les taux SCI. 
    program_month/year optionnels pour cibler un mois spécifique (sinon le plus récent)."""
    from database import ADMIN_PASSWORD
    import uuid
    from datetime import datetime

    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe admin incorrect")

    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")

    # Cibler le bon fichier: mois spécifique OU le plus récent
    if program_month > 0 and program_year > 0:
        rates_path = _get_latest_data_file("sci_lease_rates", program_month, program_year)
        if not rates_path:
            # Créer le fichier pour ce mois s'il n'existe pas
            en_months = ["", "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            rates_path = str(ROOT_DIR / "data" / f"sci_lease_rates_{en_months[program_month]}{program_year}.json")
            fr_months = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            initial_data = {"program_period": f"{fr_months[program_month]} {program_year}", "terms": [24, 27, 36, 39, 42, 48, 51, 54, 60], "vehicles_2026": [], "vehicles_2025": []}
            with open(rates_path, 'w') as f:
                json.dump(initial_data, f, indent=2, ensure_ascii=False)
    else:
        rates_path = _get_latest_data_file("sci_lease_rates")
    
    if not rates_path:
        raise HTTPException(status_code=404, detail="Fichier de taux SCI introuvable")
    with open(rates_path, 'r') as f:
        data = json.load(f)

    # 1. SNAPSHOT AVANT: copie profonde de l'etat actuel
    import copy
    data_before = copy.deepcopy(data)

    terms = data.get("terms", [24, 27, 36, 39, 42, 48, 51, 54, 60])
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))

    updated_total = 0
    unchanged_total = 0
    comparison_details = []

    for year_key, possible_titles in [("vehicles_2026", ["Lease 2026"]), ("vehicles_2025", ["Lease 2025"])]:
        ws = None
        for title in possible_titles:
            if title in wb.sheetnames:
                ws = wb[title]
                break
        if not ws:
            continue

        vehicles = data.get(year_key, [])
        vehicles_before = data_before.get(year_key, [])

        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if ri >= len(vehicles):
                break
            if not row or not row[0]:
                continue

            v = vehicles[ri]
            old_v = vehicles_before[ri] if ri < len(vehicles_before) else {}

            v["brand"] = str(row[0]).strip() if row[0] else v.get("brand", "")
            v["model"] = str(row[1]).strip() if row[1] else v.get("model", "")
            v["lease_cash"] = float(row[2]) if row[2] is not None else 0

            std = {}
            has_std = False
            for i, t in enumerate(terms):
                val = row[3 + i]
                if val is not None:
                    std[str(t)] = float(val)
                    has_std = True
            v["standard_rates"] = std if has_std else None

            alt_cash_idx = 3 + len(terms)
            v["alternative_lease_cash"] = float(row[alt_cash_idx]) if row[alt_cash_idx] is not None else 0

            alt = {}
            has_alt = False
            for i, t in enumerate(terms):
                val = row[alt_cash_idx + 1 + i]
                if val is not None:
                    alt[str(t)] = float(val)
                    has_alt = True
            v["alternative_rates"] = alt if has_alt else None

            # Comparer avant/apres pour ce vehicule
            changes = {}
            if (old_v.get("lease_cash") or 0) != v["lease_cash"]:
                changes["lease_cash"] = {"avant": old_v.get("lease_cash", 0) or 0, "apres": v["lease_cash"]}
            if (old_v.get("alternative_lease_cash") or 0) != v.get("alternative_lease_cash", 0):
                changes["alternative_lease_cash"] = {"avant": old_v.get("alternative_lease_cash", 0) or 0, "apres": v.get("alternative_lease_cash", 0)}
            if old_v.get("standard_rates") != v.get("standard_rates"):
                changes["standard_rates"] = {"avant": old_v.get("standard_rates"), "apres": v.get("standard_rates")}
            if old_v.get("alternative_rates") != v.get("alternative_rates"):
                changes["alternative_rates"] = {"avant": old_v.get("alternative_rates"), "apres": v.get("alternative_rates")}

            if changes:
                updated_total += 1
                comparison_details.append({
                    "vehicule": f"{v.get('brand','')} {v.get('model','')}",
                    "changes": changes
                })
            else:
                unchanged_total += 1

        data[year_key] = vehicles

    with open(rates_path, 'w') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    # Sauvegarder la comparaison dans MongoDB
    comparison_id = str(uuid.uuid4())
    comparison_doc = {
        "id": comparison_id,
        "type": "sci_lease",
        "date": datetime.utcnow().isoformat(),
        "updated": updated_total,
        "unchanged": unchanged_total,
        "details": comparison_details
    }
    await db.import_comparisons.insert_one(comparison_doc)

    # Force logout
    r = await db.tokens.delete_many({})
    logger.info(f"[SCI IMPORT] {updated_total} modifies, {unchanged_total} inchanges, {r.deleted_count} tokens invalides")

    return {
        "success": True,
        "message": f"Import SCI termine: {updated_total} modifies, {unchanged_total} inchanges.",
        "comparison_id": comparison_id,
        "updated": updated_total,
        "unchanged": unchanged_total,
        "comparison": comparison_details[:50]
    }


@router.get("/sci/comparisons")
async def get_sci_comparisons():
    """Retourne l'historique des comparaisons d'imports SCI"""
    comparisons = await db.import_comparisons.find(
        {"type": "sci_lease"},
        {"_id": 0}
    ).sort("date", -1).to_list(20)
    return comparisons

