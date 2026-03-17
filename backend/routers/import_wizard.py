from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Header
from typing import Optional, List, Dict, Any
from datetime import datetime
import uuid
import json
import re
import io
import asyncio
import pypdf
import pdfplumber
from services.pdfplumber_parser import extract_stable_all
from database import db, ADMIN_PASSWORD, OPENAI_API_KEY, SMTP_EMAIL, SMTP_PASSWORD, SMTP_HOST, SMTP_PORT, ROOT_DIR, logger
from models import (
    PDFExtractRequest, ProgramPreview, ExtractedDataResponse,
    SaveProgramsRequest, FinancingRates, VehicleProgram
)
from dependencies import get_current_user
from services.email_service import send_email

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

import re as re_module

def _merge_previous_sci_rates(new_vehicles_2026, new_vehicles_2025, current_month, current_year):
    """Copie les taux du mois précédent pour les véhicules dont les taux sont vides.
    Ceci évite de perdre les taux importés via Excel quand un nouveau mois est extrait."""
    import glob
    
    month_order = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                   "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
    month_order_fr = {"janvier": 1, "février": 2, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5, "juin": 6,
                      "juillet": 7, "août": 8, "aout": 8, "septembre": 9, "octobre": 10, "novembre": 11, "décembre": 12, "decembre": 12}
    all_months = {**month_order, **month_order_fr}
    
    data_dir = ROOT_DIR / "data"
    best_file = None
    best_score = (0, 0)
    
    for filepath in glob.glob(str(data_dir / "sci_lease_rates_*.json")):
        fname = filepath.split("/")[-1]
        part = fname.replace("sci_lease_rates_", "").replace(".json", "")
        year_match = re.search(r'(\d{4})$', part)
        if not year_match:
            continue
        year = int(year_match.group(1))
        month_part = part[:year_match.start()].lower().strip()
        month_num = all_months.get(month_part, 0)
        
        score = (year, month_num)
        # Find the best file that is BEFORE the current month
        if score < (current_year, current_month) and score > best_score:
            best_score = score
            best_file = filepath
    
    if not best_file:
        logger.info("[SCI Merge] No previous SCI file found to merge rates from")
        return
    
    try:
        with open(best_file, 'r', encoding='utf-8') as f:
            prev_data = json.load(f)
        
        # Build lookup: (brand_lower, model_lower) -> vehicle with rates
        def build_lookup(vehicles):
            lookup = {}
            for v in (vehicles or []):
                key = (v.get("brand", "").lower(), v.get("model", "").lower())
                sr = v.get("standard_rates") or v.get("standard_rate") or {}
                ar = v.get("alternative_rates") or v.get("alt_rates") or {}
                if sr or ar:
                    lookup[key] = {"standard_rates": sr, "alternative_rates": ar}
            return lookup
        
        prev_2026 = build_lookup(prev_data.get("vehicles_2026", []))
        prev_2025 = build_lookup(prev_data.get("vehicles_2025", []))
        merged_count = 0
        
        for v in (new_vehicles_2026 or []):
            sr = v.get("standard_rates") or {}
            ar = v.get("alternative_rates") or {}
            if not sr and not ar:
                key = (v.get("brand", "").lower(), v.get("model", "").lower())
                prev = prev_2026.get(key)
                if prev:
                    v["standard_rates"] = prev["standard_rates"]
                    v["alternative_rates"] = prev["alternative_rates"]
                    merged_count += 1
        
        for v in (new_vehicles_2025 or []):
            sr = v.get("standard_rates") or {}
            ar = v.get("alternative_rates") or {}
            if not sr and not ar:
                key = (v.get("brand", "").lower(), v.get("model", "").lower())
                prev = prev_2025.get(key)
                if prev:
                    v["standard_rates"] = prev["standard_rates"]
                    v["alternative_rates"] = prev["alternative_rates"]
                    merged_count += 1
        
        logger.info(f"[SCI Merge] Merged rates from {best_file} for {merged_count} vehicles")
    except Exception as e:
        logger.error(f"[SCI Merge] Error merging rates: {e}")

def _strip_delivery_credit(text: str) -> str:
    """
    ╔══════════════════════════════════════════════════════════╗
    ║  SUPPRIMER la colonne Delivery Credit (261Q03, 'E' Only)║
    ║  du texte PDF AVANT envoi à l'IA.                       ║
    ║  L'IA ne verra JAMAIS ces montants.                     ║
    ╚══════════════════════════════════════════════════════════╝
    """
    import re
    # Remove header references
    text = re.sub(r'Delivery\s*Credit\*?', '', text, flags=re.IGNORECASE)
    text = re.sub(r'261Q03', '', text)
    text = re.sub(r"['\"]E['\"]\s*Only", '', text, flags=re.IGNORECASE)
    text = re.sub(r"TYPE\s+OF\s+SALE\s*[:\s]*'E'\s*Only", '', text, flags=re.IGNORECASE)
    text = re.sub(r"Stackable\s+only\s+with\s+Consumer\s+Cash\s+and\s+Bonus\s+Cash", '', text, flags=re.IGNORECASE)
    text = re.sub(r"BEFORE\s+TAX", '', text, flags=re.IGNORECASE)
    
    # Remove dollar amounts at the END of lines that are Delivery Credit values
    # Pattern: lines ending with $X,XXX (after Bonus Cash which is also $ at end)
    # Strategy: each data line typically ends with 2 dollar amounts: Bonus Cash then Delivery Credit
    # We remove the very last dollar amount on lines that have 2+ dollar amounts at the end
    lines = text.split('\n')
    cleaned = []
    for line in lines:
        # Find all dollar amounts in the line
        dollar_pattern = r'\$[\d,]+(?:\.\d+)?'
        dollars = list(re.finditer(dollar_pattern, line))
        
        if len(dollars) >= 2:
            # Check if the last 2 dollar amounts are close together (Bonus Cash + Delivery Credit)
            last = dollars[-1]
            second_last = dollars[-2]
            gap = last.start() - second_last.end()
            # If gap is small (< 30 chars), they're adjacent columns → remove the last one (Delivery Credit)
            if gap < 30:
                line = line[:last.start()] + line[last.end():]
        elif len(dollars) == 1:
            # Single dollar at end of a data line could be Delivery Credit
            # Check context: if the line has rate patterns (X.XX%) before it, the trailing $ is likely DC
            if re.search(r'\d+\.\d+%', line):
                last = dollars[-1]
                # Check if there's rate data before this dollar (indicating it's at the end of the row)
                before = line[:last.start()].strip()
                if before.endswith('-') or before.endswith('%'):
                    line = line[:last.start()] + line[last.end():]
        
        cleaned.append(line)
    
    return '\n'.join(cleaned)

def normalize_correction_str(s: str) -> str:
    """Normalise une chaine pour matching flexible des corrections."""
    if not s:
        return ""
    s = s.strip().lower()
    s = re_module.sub(r'\(cpos[^)]*\)', '', s)
    s = re_module.sub(r'\([A-Z]{2,}[0-9]+[^)]*\)', '', s, flags=re_module.IGNORECASE)
    s = re_module.sub(r'\(etm\)', '', s, flags=re_module.IGNORECASE)
    s = re_module.sub(r'\s+', ' ', s).strip()
    s = s.rstrip(' ,')
    return s

def normalize_correction_model(model: str) -> str:
    """Normalise le nom du modele pour matching flexible."""
    if not model:
        return ""
    m = model.strip().lower()
    m = m.replace("grand cherokee/grand cherokee l", "grand cherokee/l")
    m = m.replace("grand wagoneer / grand wagoneer l", "grand wagoneer/l")
    m = m.replace("grand wagoneer/grand wagoneer l", "grand wagoneer/l")
    m = m.replace("wagoneer / wagoneer l", "wagoneer/l")
    m = m.replace("wagoneer/wagoneer l", "wagoneer/l")
    m = re_module.sub(r'\s+', ' ', m).strip()
    return m

async def find_best_correction(brand: str, model: str, trim: str, year: int) -> dict:
    """Cherche la meilleure correction memorisee avec matching flexible."""
    # Strategie 1: Match exact
    correction = await db.program_corrections.find_one(
        {"brand": brand, "model": model, "trim": trim, "year": year},
        {"_id": 0}
    )
    if correction:
        return correction

    # Strategie 2: Match normalise
    norm_model = normalize_correction_model(model)
    norm_trim = normalize_correction_str(trim)

    all_corrections = await db.program_corrections.find(
        {"year": year},
        {"_id": 0}
    ).to_list(500)

    for c in all_corrections:
        if c.get("brand", "").lower() != brand.lower():
            continue
        c_model = normalize_correction_model(c.get("model", ""))
        c_trim = normalize_correction_str(c.get("trim", ""))
        if c_model == norm_model and c_trim == norm_trim:
            return c
        # Trim partiel
        if c_model == norm_model and norm_trim and c_trim and (c_trim in norm_trim or norm_trim in c_trim):
            return c

    return None

router = APIRouter()

# ============ Excel Generation Function ============

# ============ STANDARD EXTRACTION PROMPT (Source de verite pour la structure Excel) ============
# Ce prompt definit la structure FIGEE du fichier Excel genere.
# Colonnes: A(Marque) B(Modele) C(Trim) D(Annee) | E(Rabais Opt1) F-K(Taux Opt1) | L(Rabais Opt2) M-R(Taux Opt2) | S(Bonus)
# L'IA peut ajouter des lignes (nouveaux vehicules) mais NE DOIT JAMAIS modifier la structure.

def build_extraction_prompt(pdf_text: str) -> str:
    """
    ╔══════════════════════════════════════════════════════════════╗
    ║  PROMPT VERROUILLÉ — NE JAMAIS MODIFIER                    ║
    ║  Structure figée pour extraction PDF FCA Canada QBC Retail  ║
    ║  Même gabarit chaque mois. Mêmes règles. Mêmes colonnes.   ║
    ╚══════════════════════════════════════════════════════════════╝
    """
    return f"""EXTRAIS TOUS LES VÉHICULES de ce PDF de programmes de financement FCA Canada QBC Retail.

TEXTE COMPLET DU PDF:
{pdf_text}

=== STRUCTURE EXCEL FIGÉE (IDENTIQUE CHAQUE MOIS) ===
- Colonne A: Marque (brand)
- Colonne B: Modèle (model)  
- Colonne C: Version/Trim (trim)
- Colonne D: Année (year)
- Colonne E: Rabais Option 1 — Consumer Cash en $ (consumer_cash)
- Colonnes F-K: Taux Option 1 — 36m, 48m, 60m, 72m, 84m, 96m (option1_rates)
- Colonne L: Rabais Option 2 — Alternative Consumer Cash en $ (alt_consumer_cash)
- Colonnes M-R: Taux Option 2 — 36m, 48m, 60m, 72m, 84m, 96m (option2_rates)
- Colonne S: Bonus Cash en $ (bonus_cash)

=== STRUCTURE DU PDF FCA (TOUJOURS LE MÊME GABARIT) ===
Le PDF a TOUJOURS cette structure de colonnes de GAUCHE à DROITE:
1. Nom du véhicule (sous la barre de couleur de la marque)
2. Consumer Cash ($) — Rabais Option 1
3. 6 colonnes de taux Option 1: 36M, 48M, 60M, 72M, 84M, 96M
4. Alternative Consumer Cash ($) — Rabais Option 2 (souvent vide)
5. 6 colonnes de taux Option 2: 36M, 48M, 60M, 72M, 84M, 96M (souvent "- - - - - -")
6. Bonus Cash ($) — colonne jaune, pour TOUS les clients
7. Delivery Credit ($) — dernière colonne rouge — TOUJOURS IGNORER

=== RÈGLES DE LECTURE (FIGÉES) ===
- "- - - - - -" ou colonnes vides = option non disponible → null
- "P" ou "P$X,XXX" avant un montant = le montant est valide, extraire le chiffre
- 6 taux = toujours dans l'ordre 36M, 48M, 60M, 72M, 84M, 96M
- La marque apparaît comme BARRE DE COULEUR à gauche (ex: barre verte = FIAT, barre rouge = DODGE)
- Chaque ligne SOUS la barre = un véhicule de cette marque

=== OPTION 2 — RÈGLES ===
- BEAUCOUP de véhicules n'ont PAS d'Option 2
- Colonnes Option 2 VIDES → alt_consumer_cash = 0, option2_rates = null
- NE PAS inventer ou copier des taux d'un autre véhicule
- Chaque véhicule traité INDIVIDUELLEMENT

=== BONUS CASH vs DELIVERY CREDIT — DEUX COLONNES DISTINCTES ===
1. BONUS CASH (colonne jaune) = bonus RÉEL pour TOUS les clients (TYPE OF SALE: 1, L or E)
   → EXTRAIRE le montant exact dans bonus_cash
   
2. DELIVERY CREDIT (dernière colonne, rouge, 261Q03) = 'E' Only = EMPLOYÉS SEULEMENT
   → TOUJOURS IGNORER. NE JAMAIS mettre dans bonus_cash.

*** bonus_cash = UNIQUEMENT la colonne jaune "Bonus Cash". JAMAIS le Delivery Credit ***

=== CODES PROGRAMMES FCA — VÉHICULES À EXTRAIRE ===
Voici la liste COMPLÈTE des véhicules FCA. Chaque véhicule du PDF DOIT correspondre à un de ces modèles:

CHRYSLER:
- Grand Caravan (CVP, SXT, etc.)
- Pacifica (Select, Limited, Pinnacle, Select PHEV)

DODGE:
- Charger (base, GT, Scat Pack)
- Durango (base, GT, R/T)
- Hornet (base, R/T, GT Plus)

FIAT:
- 500e BEV (brand="Fiat", model="500e", trim="BEV")
  ATTENTION: "FIAT" apparaît comme barre de couleur, la ligne dit juste "500e BEV"

JEEP:
- Compass (Sport, North, Altitude, Limited, Trailhawk)
- Cherokee (base, Limited, Overland)
- Wrangler (Sport, Willys, Sahara, Rubicon, 4Xe)
- Gladiator (Sport, Mojave, Rubicon)
- Grand Cherokee (Limited, Overland, Summit, 4Xe, L variants)
- Grand Wagoneer (base, Series I, L, Summit)
- Wagoneer (base, Series I, S, L)

RAM:
- 1500 (Tradesman, Big Horn, Sport, Rebel, Laramie, Limited, Tungsten, RHO)
- 2500 (Tradesman, Big Horn, Rebel, Laramie, Limited)
- 3500 (Tradesman, Big Horn, Laramie, Limited)
- ProMaster Cargo Van (Low Roof, High Roof, Super High Roof, divers WB)
- Chassis Cab (4500, 5500)

=== SECTIONS DU PDF ===
- "2026 MODELS" → year: 2026
- "2025 MODELS" → year: 2025
EXTRAIS les véhicules des DEUX sections!

=== JSON REQUIS ===
{{{{
    "programs": [
        {{{{
            "brand": "Chrysler",
            "model": "Grand Caravan", 
            "trim": "SXT",
            "year": 2026,
            "consumer_cash": 0,
            "alt_consumer_cash": 0,
            "bonus_cash": 0,
            "option1_rates": {{{{"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}}}},
            "option2_rates": null
        }}}},
        {{{{
            "brand": "Fiat",
            "model": "500e",
            "trim": "BEV",
            "year": 2025,
            "consumer_cash": 0,
            "alt_consumer_cash": 0,
            "bonus_cash": 5000,
            "option1_rates": {{{{"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}}}},
            "option2_rates": {{{{"rate_36": 0.00, "rate_48": 0.00, "rate_60": 1.99, "rate_72": 2.99, "rate_84": 3.99, "rate_96": 4.99}}}}
        }}}}
    ]
}}}}

=== VÉRIFICATION FINALE (CHECKLIST) ===
Avant de retourner le JSON, vérifie:
[ ] TOUS les véhicules des sections 2026 ET 2025 sont inclus
[ ] La Fiat 500e BEV est présente (souvent oubliée car dernière marque)
[ ] bonus_cash = montant de la colonne jaune (PAS le Delivery Credit rouge)
[ ] Delivery Credit = IGNORÉ (colonne rouge, 'E' Only)
[ ] Option 2: si vide dans le PDF → null (ne pas inventer)
[ ] consumer_cash et alt_consumer_cash corrects par véhicule
[ ] 6 taux par option (36, 48, 60, 72, 84, 96)
[ ] Chaque brand correspond à la barre de couleur du PDF"""



def generate_excel_from_programs(programs: List[Dict[str, Any]], program_month: int, program_year: int, sci_lease_data: Dict = None) -> bytes:
    """Génère un fichier Excel avec onglet Programmes + onglet SCI Lease (si données fournies).
    Colonnes: Année | Marque | Modèle | Version | Loyauté |
              Rabais Opt1 | 36m-96m Opt1 | Rabais Alt | 36m-96m Opt2 | Bonus
    Premières colonnes et en-têtes figés pour faciliter la correction.
    """
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financement"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    option1_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    option2_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    bonus_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    loyalty_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    month_names = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

    # ── Row 1: Title ──
    ws.merge_cells('A1:T1')
    ws['A1'] = f"PROGRAMMES DE FINANCEMENT RETAIL - {month_names[program_month].upper()} {program_year}"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center")

    # ── Row 2: Group headers ──
    # A-E: Véhicule
    ws.merge_cells('A2:E2')
    ws['A2'] = "VEHICULE"
    ws['A2'].font = header_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = header_alignment
    for c in range(2, 6):
        ws.cell(row=2, column=c).fill = header_fill

    # F-L: Option 1
    ws.merge_cells('F2:L2')
    ws['F2'] = "OPTION 1 - Consumer Cash + Taux Standard"
    ws['F2'].font = header_font
    ws['F2'].fill = option1_fill
    ws['F2'].alignment = header_alignment
    for c in range(7, 13):
        ws.cell(row=2, column=c).fill = option1_fill

    # M-S: Option 2
    ws.merge_cells('M2:S2')
    ws['M2'] = "OPTION 2 - Alternative Cash + Taux Reduit"
    ws['M2'].font = header_font
    ws['M2'].fill = option2_fill
    ws['M2'].alignment = header_alignment
    for c in range(14, 20):
        ws.cell(row=2, column=c).fill = option2_fill

    # T: Bonus
    ws['T2'] = "BONUS"
    ws['T2'].font = header_font
    ws['T2'].fill = bonus_fill
    ws['T2'].alignment = header_alignment

    ws.row_dimensions[2].height = 30

    # ── Row 3: Column headers ──
    # Cols: A=Année B=Marque C=Modèle D=Version E=Loyauté
    #       F=Rabais($) G-L=36m..96m
    #       M=Rabais Alt($) N-S=36m..96m
    #       T=Bonus($)
    headers = [
        "Année", "Marque", "Modèle", "Version", "P",
        "Rabais ($)", "36m", "48m", "60m", "72m", "84m", "96m",
        "Rabais ($)", "36m", "48m", "60m", "72m", "84m", "96m",
        "Bonus ($)"
    ]
    col_fills = (
        [header_fill]*4 + [loyalty_fill]
        + [option1_fill]*7
        + [option2_fill]*7
        + [bonus_fill]
    )
    for col, (header, fill) in enumerate(zip(headers, col_fills), 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.fill = fill
    ws.row_dimensions[3].height = 25

    # ── Data rows ──
    # Light background fills for data cells
    opt1_data_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    opt2_data_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    bonus_data_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    loyalty_data_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    def format_rate(rate):
        if rate is None or rate == "-":
            return "-"
        try:
            r = float(rate)
            return f"{r:.2f}%"
        except Exception:
            return "-"

    for row_idx, prog in enumerate(programs, 4):
        opt1_rates = prog.get("option1_rates") or {}
        opt2_rates = prog.get("option2_rates") or {}
        consumer_cash = prog.get("consumer_cash", 0) or 0
        alt_consumer_cash = prog.get("alt_consumer_cash", 0) or 0
        bonus_cash = prog.get("bonus_cash", 0) or 0

        # Build loyalty string from P flags
        loyalty_parts = []
        if prog.get("loyalty_cash"):
            loyalty_parts.append("$")
        if prog.get("loyalty_opt1"):
            loyalty_parts.append("O1")
        if prog.get("loyalty_opt2"):
            loyalty_parts.append("O2")
        loyalty_str = "P" if loyalty_parts else ""

        data = [
            prog.get("year", ""),
            prog.get("brand", ""),
            prog.get("model", ""),
            prog.get("trim", "") or "",
            loyalty_str,
            # Option 1
            f"${consumer_cash:,.0f}" if consumer_cash else "-",
            format_rate(opt1_rates.get("rate_36")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_48")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_60")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_72")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_84")) if opt1_rates else "-",
            format_rate(opt1_rates.get("rate_96")) if opt1_rates else "-",
            # Option 2
            f"${alt_consumer_cash:,.0f}" if alt_consumer_cash else "-",
            format_rate(opt2_rates.get("rate_36")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_48")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_60")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_72")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_84")) if opt2_rates else "-",
            format_rate(opt2_rates.get("rate_96")) if opt2_rates else "-",
            # Bonus
            f"${bonus_cash:,.0f}" if bonus_cash else "-",
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            # Trim column (col 4) left-aligned for readability
            if col == 4:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")
            if col == 5:
                cell.fill = loyalty_data_fill
                if value == "P":
                    cell.font = Font(bold=True, color="E65100")
            elif 6 <= col <= 12:
                cell.fill = opt1_data_fill
            elif 13 <= col <= 19:
                cell.fill = opt2_data_fill
            elif col == 20:
                cell.fill = bonus_data_fill

    # ── Column widths ──
    col_widths = [7, 12, 22, 55, 4, 12, 8, 8, 8, 8, 8, 8, 12, 8, 8, 8, 8, 8, 8, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ── Freeze panes: fix Année+Marque+Modèle+Version+Loyauté columns + header rows ──
    ws.freeze_panes = 'F4'
    
    # ============ ONGLET 2: SCI Lease (si données fournies) ============
    if sci_lease_data:
        ws2 = wb.create_sheet("SCI Lease")
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1A237E")
        header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        cat_fill_std = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
        cat_fill_alt = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        thin_border_l = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # SCI Lease terms
        lease_terms = ["24m", "27m", "36m", "39m", "42m", "48m", "51m", "54m", "60m"]
        
        # Row 1: Title
        ws2.merge_cells('A1:U1')
        title_cell = ws2['A1']
        title_cell.value = f"TAUX DE LOCATION SCI LEASE - {month_names[program_month].upper()} {program_year}"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 35
        
        # Row 2: Category headers
        # A-B: Vehicle info
        ws2.merge_cells('A2:B2')
        ws2['A2'].value = "VÉHICULE"
        ws2['A2'].font = header_font
        ws2['A2'].fill = header_fill
        ws2['B2'].fill = header_fill
        # C: Lease Cash
        ws2['C2'].value = "LEASE CASH"
        ws2['C2'].font = header_font
        ws2['C2'].fill = header_fill
        # D-L: Standard Rates
        ws2.merge_cells('D2:L2')
        ws2['D2'].value = "STANDARD RATES (Stackable avec Lease Cash)"
        ws2['D2'].font = header_font
        ws2['D2'].fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
        for c in range(5, 13):
            ws2.cell(row=2, column=c).fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
        # M-U: Alternative Rates
        ws2.merge_cells('M2:U2')
        ws2['M2'].value = "ALTERNATIVE RATES (Stackable avec Alt Lease Cash)"
        ws2['M2'].font = header_font
        ws2['M2'].fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
        for c in range(14, 22):
            ws2.cell(row=2, column=c).fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
        ws2.row_dimensions[2].height = 30
        
        # Row 3: Column headers
        col_headers = ["Marque", "Modèle", "Rabais ($)"] + lease_terms + lease_terms
        for col, header in enumerate(col_headers, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border_l
            cell.alignment = Alignment(horizontal="center")
            if col <= 3:
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            elif col <= 12:
                cell.fill = cat_fill_std
            else:
                cell.fill = cat_fill_alt
        ws2.row_dimensions[3].height = 25
        
        # Data rows
        row_idx = 4
        term_keys = ["24", "27", "36", "39", "42", "48", "51", "54", "60"]
        
        for year_label, vehicles_key in [("2026", "vehicles_2026"), ("2025", "vehicles_2025")]:
            vehicles = sci_lease_data.get(vehicles_key, [])
            if not vehicles:
                continue
            
            # Year separator
            ws2.merge_cells(f'A{row_idx}:U{row_idx}')
            year_cell = ws2.cell(row=row_idx, column=1, value=f"MODÈLES {year_label}")
            year_cell.font = Font(bold=True, size=11, color="FFFFFF")
            year_cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
            year_cell.alignment = Alignment(horizontal="center")
            row_idx += 1
            
            for v in vehicles:
                brand = v.get("brand", "")
                model = v.get("model", "")
                lease_cash = v.get("lease_cash", 0) or 0
                std_rates = v.get("standard_rates") or {}
                alt_rates = v.get("alternative_rates") or {}
                
                def format_lease_rate(val):
                    if val is None:
                        return "-"
                    return f"{val}%"
                
                data = [
                    brand, model,
                    f"${lease_cash:,.0f}" if lease_cash else "-",
                ] + [
                    format_lease_rate(std_rates.get(t)) for t in term_keys
                ] + [
                    format_lease_rate(alt_rates.get(t)) for t in term_keys
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws2.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border_l
                    # Model column (col 2) left-aligned for readability
                    if col == 2:
                        cell.alignment = Alignment(horizontal="left", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="center")
                    if col <= 3:
                        pass
                    elif col <= 12:
                        cell.fill = cat_fill_std
                    else:
                        cell.fill = cat_fill_alt
                
                row_idx += 1
        
        # Column widths for SCI Lease sheet
        sci_widths = [12, 60, 12] + [7]*9 + [7]*9
        for col, width in enumerate(sci_widths, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Freeze: rows 1-3 + columns A-C
        ws2.freeze_panes = 'D4'
    
    # ── Rapport de validation sheet ──
    try:
        from services.pdfplumber_parser import validate_extraction
        report = validate_extraction(programs, sci_lease_data)
        ws3 = wb.create_sheet("Rapport")
        ws3.column_dimensions['A'].width = 14
        ws3.column_dimensions['B'].width = 70

        ws3.cell(1, 1, "RAPPORT DE VALIDATION").font = Font(bold=True, size=14)
        ws3.cell(2, 1, f"Mois: {program_month}/{program_year}").font = Font(size=11)

        stats = report.get('stats', {})
        r = 4
        ws3.cell(r, 1, "Programmes").font = Font(bold=True)
        ws3.cell(r, 2, stats.get('total_programs', 0))
        r += 1
        ws3.cell(r, 1, "Par année").font = Font(bold=True)
        ws3.cell(r, 2, str(stats.get('by_year', {})))
        r += 1
        ws3.cell(r, 1, "Par marque").font = Font(bold=True)
        ws3.cell(r, 2, str(stats.get('by_brand', {})))
        r += 1
        ws3.cell(r, 1, "Loyauté (P)").font = Font(bold=True)
        ws3.cell(r, 2, stats.get('loyalty_count', 0))
        r += 1
        ws3.cell(r, 1, "Bonus Cash").font = Font(bold=True)
        ws3.cell(r, 2, stats.get('bonus_count', 0))
        r += 1
        ws3.cell(r, 1, "SCI Lease").font = Font(bold=True)
        ws3.cell(r, 2, str(stats.get('sci_lease', {})))

        warnings = report.get('warnings', [])
        errors = report.get('errors', [])

        r += 2
        ws3.cell(r, 1, "RÉSULTAT").font = Font(bold=True, size=12)
        if not warnings and not errors:
            ws3.cell(r, 2, "Aucun problème détecté").font = Font(color="00AA00", size=12, bold=True)
        else:
            ws3.cell(r, 2, f"{len(errors)} erreur(s), {len(warnings)} avertissement(s)").font = Font(color="FF0000" if errors else "FF8800", size=12, bold=True)

        if errors:
            r += 2
            ws3.cell(r, 1, "ERREURS").font = Font(bold=True, color="FF0000")
            for e in errors:
                r += 1
                ws3.cell(r, 2, e).font = Font(color="FF0000")

        if warnings:
            r += 2
            ws3.cell(r, 1, "AVERTISSEMENTS").font = Font(bold=True, color="FF8800")
            for w in warnings:
                r += 1
                ws3.cell(r, 2, w).font = Font(color="FF8800")
    except Exception as val_err:
        logger.error(f"[Excel] Validation sheet error: {val_err}")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def send_excel_email(excel_data: bytes, admin_email: str, program_month: int, program_year: int, program_count: int):
    """Envoie le fichier Excel par email"""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    
    month_names = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                   "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    
    msg = MIMEMultipart()
    msg['From'] = SMTP_EMAIL
    msg['To'] = admin_email
    msg['Subject'] = f"CalcAuto AiPro - Extraction PDF {month_names[program_month]} {program_year}"
    
    body = f"""
Bonjour,

L'extraction du PDF des programmes de financement est terminée.

📊 Résumé:
• Période: {month_names[program_month]} {program_year}
• Programmes extraits: {program_count}

Le fichier Excel est joint à cet email pour vérification.

⚠️ IMPORTANT: Veuillez vérifier les données dans le fichier Excel avant de confirmer l'import dans l'application.

---
CalcAuto AiPro
    """
    
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    # Attach Excel file
    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment.set_payload(excel_data)
    encoders.encode_base64(attachment)
    filename = f"programmes_{month_names[program_month].lower()}_{program_year}.xlsx"
    attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
    msg.attach(attachment)
    
    # Send email
    try:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        logger.error(f"Error sending Excel email: {str(e)}")
        return False

# ============ PDF Import with pdfplumber ============

@router.post("/scan-pdf")
async def scan_pdf(
    file: UploadFile = File(...),
    password: str = Form(...)
):
    """
    Scanne le PDF et détecte TOUTES les sections du TOC (page 2).
    Retourne la liste pour la boîte du haut.
    """
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    try:
        from services.pdfplumber_parser import auto_detect_pages
        pdf_content = await file.read()
        result = auto_detect_pages(pdf_content)
        return {
            "success": True,
            "sections": result.get('sections', []),  # ← boîte du haut va afficher ça
            "total_pages": result['total_pages']
        }
    except Exception as e:
        logger.error(f"[ScanPDF] Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur scan: {str(e)}")


@router.get("/program-meta")
async def get_program_meta(month: Optional[int] = None, year: Optional[int] = None):
    """
    Retourne les métadonnées du programme du mois (événement, loyauté, etc.).
    Auto-détecte le dernier fichier si month/year non spécifiés.
    """
    en_month_abbrev = ["", "jan", "feb", "mar", "apr", "may", "jun",
                      "jul", "aug", "sep", "oct", "nov", "dec"]
    data_dir = ROOT_DIR / "data"

    if month and year:
        filename = f"program_meta_{en_month_abbrev[month]}{year}.json"
        filepath = data_dir / filename
    else:
        # Find latest program_meta file
        meta_files = sorted(data_dir.glob("program_meta_*.json"), reverse=True)
        filepath = meta_files[0] if meta_files else None

    if filepath and filepath.exists():
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)

    return {
        "event_names": [],
        "program_period": "",
        "program_month": "",
        "program_year": None,
        "loyalty_rate": 0.0,
        "no_payments_days": 0,
        "featured_rate": None,
        "featured_term": None,
        "key_message": "",
    }

@router.post("/verify-password")
async def verify_password(password: str = Form(...)):
    """Vérifie le mot de passe admin"""
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    return {"success": True, "message": "Mot de passe vérifié"}


from fastapi.responses import StreamingResponse

@router.get("/download-excel")
async def download_excel(month: int = 3, year: int = 2026):
    """Génère et télécharge le fichier Excel de vérification des programmes du mois."""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")

    programs = []
    async for doc in db.programs.find(
        {"program_month": month, "program_year": year},
        {"_id": 0}
    ):
        programs.append(doc)

    if not programs:
        raise HTTPException(status_code=404, detail=f"Aucun programme trouvé pour {month}/{year}")

    # Load SCI lease data
    en_month_abbrev = ["", "jan", "feb", "mar", "apr", "may", "jun",
                      "jul", "aug", "sep", "oct", "nov", "dec"]
    sci_data = None
    sci_path = ROOT_DIR / "data" / f"sci_lease_rates_{en_month_abbrev[month]}{year}.json"
    if sci_path.exists():
        with open(sci_path, 'r', encoding='utf-8') as f:
            sci_data = json.load(f)

    excel_data = generate_excel_from_programs(programs, month, year, sci_lease_data=sci_data)

    month_names = ["", "Janvier", "Fevrier", "Mars", "Avril", "Mai", "Juin",
                   "Juillet", "Aout", "Septembre", "Octobre", "Novembre", "Decembre"]
    filename = f"Programmes_{month_names[month]}_{year}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/validate-data")
async def validate_data(month: int = 3, year: int = 2026):
    """Valide les données extraites en BD et retourne un rapport de qualité."""
    from services.pdfplumber_parser import validate_extraction

    programs = []
    async for doc in db.programs.find(
        {"program_month": month, "program_year": year},
        {"_id": 0}
    ):
        programs.append(doc)

    if not programs:
        return {"success": False, "message": f"Aucun programme trouvé pour {month}/{year}"}

    # Load SCI lease data
    en_month_abbrev = ["", "jan", "feb", "mar", "apr", "may", "jun",
                      "jul", "aug", "sep", "oct", "nov", "dec"]
    sci_data = None
    sci_path = ROOT_DIR / "data" / f"sci_lease_rates_{en_month_abbrev[month]}{year}.json"
    if sci_path.exists():
        with open(sci_path, 'r', encoding='utf-8') as f:
            sci_data = json.load(f)

    report = validate_extraction(programs, sci_data)
    report['success'] = len(report['errors']) == 0
    report['programs_count'] = len(programs)
    return report


@router.post("/extract-pdf", response_model=ExtractedDataResponse)
async def extract_pdf(
    file: UploadFile = File(...),
    password: str = Form(...),
    program_month: int = Form(...),
    program_year: int = Form(...),
    start_page: Optional[int] = Form(None),
    end_page: Optional[int] = Form(None),
    lease_start_page: Optional[int] = Form(None),
    lease_end_page: Optional[int] = Form(None)
):
    """
    Extrait les données de financement d'un PDF via pdfplumber (déterministe).
    Utilise maintenant extract_stable_all pour 100% de stabilité sur tous les PDFs mensuels.
    """
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")

    try:
        from services.pdfplumber_parser import extract_stable_all   # ← Utilise la version stable

        pdf_content = await file.read()

        # EXTRACTION 100% STABLE (remplace toutes les anciennes fonctions)
        programs_data = await extract_stable_all(pdf_content)

        valid_programs = programs_data["programs"]
        sci_data_for_excel = programs_data["sci"]
        validation = programs_data["validation"]
        status = programs_data["status"]

        logger.info(f"[Sync] extract_stable_all → {len(valid_programs)} programmes - {status}")

        # Auto-save programs (seulement si extraction réussie - évite perte de données)
        excel_sent = False
        saved_count = 0
        try:
            if not valid_programs:
                logger.warning("[Sync] 0 programmes extraits - données existantes CONSERVÉES pour éviter la perte de données")
                return ExtractedDataResponse(
                    success=False,
                    message=f"0 programmes extraits. Vérifiez le PDF. Les données existantes n'ont PAS été effacées.",
                    programs=[],
                    raw_text="",
                    sci_lease_count=0
                )
            await db.programs.delete_many({"program_month": program_month, "program_year": program_year})
            for prog in valid_programs:
                opt1 = prog.get("option1_rates")
                if opt1 is None or not isinstance(opt1, dict):
                    opt1 = {"rate_36": None, "rate_48": None, "rate_60": None,
                            "rate_72": None, "rate_84": None, "rate_96": None}
                program_doc = {
                    "id": str(uuid.uuid4()),
                    "brand": prog.get("brand", ""),
                    "model": prog.get("model", ""),
                    "trim": prog.get("trim", ""),
                    "year": prog.get("year", program_year),
                    "consumer_cash": prog.get("consumer_cash", 0) or 0,
                    "bonus_cash": prog.get("bonus_cash", 0) or 0,
                    "alt_consumer_cash": prog.get("alt_consumer_cash", 0) or 0,
                    "option1_rates": opt1,
                    "option2_rates": prog.get("option2_rates"),
                    "loyalty_cash": prog.get("loyalty_cash", False),
                    "loyalty_opt1": prog.get("loyalty_opt1", False),
                    "loyalty_opt2": prog.get("loyalty_opt2", False),
                    "program_month": program_month,
                    "program_year": program_year,
                    "created_at": datetime.utcnow().isoformat()
                }
                await db.programs.insert_one(program_doc)
                saved_count += 1
            logger.info(f"[Sync] Auto-saved {saved_count} programs")
        except Exception as save_error:
            logger.error(f"[Sync] Save error: {str(save_error)}")

        # Merge SCI rates si nécessaire
        if sci_data_for_excel:
            vehicles_2026 = sci_data_for_excel.get("vehicles_2026", [])
            vehicles_2025 = sci_data_for_excel.get("vehicles_2025", [])
            _merge_previous_sci_rates(vehicles_2026, vehicles_2025, program_month, program_year)

        # Run post-extraction validation
        try:
            from services.pdfplumber_parser import validate_extraction
            validation = validate_extraction(valid_programs, sci_data_for_excel)
            validation_msg = ""
            if validation['warnings']:
                validation_msg = f" | {len(validation['warnings'])} avertissement(s)"
            if validation['errors']:
                validation_msg += f" | {len(validation['errors'])} erreur(s)"
        except Exception as val_error:
            logger.error(f"[Sync] Validation error: {str(val_error)}")
            validation_msg = ""

        # Generate Excel and send email
        if EXCEL_AVAILABLE and valid_programs and SMTP_EMAIL:
            try:
                excel_data = generate_excel_from_programs(valid_programs, program_month, program_year, sci_lease_data=sci_data_for_excel)
                excel_sent = send_excel_email(excel_data, SMTP_EMAIL, program_month, program_year, len(valid_programs))
            except Exception as excel_error:
                logger.error(f"[Sync] Excel email error: {str(excel_error)}")

        lease_msg = f" + {len(sci_data_for_excel.get('vehicles_2026', []) + sci_data_for_excel.get('vehicles_2025', []))} taux SCI Lease" if sci_data_for_excel else ""
        return ExtractedDataResponse(
            success=True,
            message=f"Extrait et sauvegardé {len(valid_programs)} programmes{lease_msg}{validation_msg}" + (" - Excel envoyé par email!" if excel_sent else ""),
            programs=valid_programs,
            raw_text="",
            sci_lease_count=len(sci_data_for_excel.get('vehicles_2026', []) + sci_data_for_excel.get('vehicles_2025', [])) if sci_data_for_excel else 0
        )

    except Exception as e:
        logger.error(f"[Sync] Error extracting PDF: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur d'extraction: {str(e)}")
        
# ============ Async PDF Extraction (for environments with short timeouts) ============
async def _run_extraction_task(task_id: str, pdf_content: bytes, password: str,
                                program_month: int, program_year: int,
                                start_page: int, end_page: int,
                                lease_start_page: Optional[int], lease_end_page: Optional[int]):
    """Background task: extraction 100% stable avec extract_stable_all"""
    try:
        await db.extract_tasks.update_one(
            {"task_id": task_id},
            {"$set": {"status": "extracting", "message": "Extraction stable (extract_stable_all)..."}}
        )

        # EXTRACTION 100% STABLE
        programs_data = await extract_stable_all(pdf_content)

        valid_programs = programs_data["programs"]
        sci_data_for_excel = programs_data["sci"]

        logger.info(f"[Async] extract_stable_all → {len(valid_programs)} programmes - {programs_data.get('status', 'OK')}")

        # ── Step 2: Save programs to MongoDB (protect against data loss) ──
        saved_count = 0
        try:
            if not valid_programs:
                logger.warning("[Async] 0 programmes extraits - données existantes CONSERVÉES")
                await db.extract_tasks.update_one(
                    {"task_id": task_id},
                    {"$set": {"status": "error", "message": "0 programmes extraits"}}
                )
                return
            await db.programs.delete_many({"program_month": program_month, "program_year": program_year})
            for prog in valid_programs:
                opt1 = prog.get("option1_rates")
                if opt1 is None or not isinstance(opt1, dict):
                    opt1 = {"rate_36": None, "rate_48": None, "rate_60": None,
                            "rate_72": None, "rate_84": None, "rate_96": None}
                program_doc = {
                    "id": str(uuid.uuid4()),
                    "brand": prog.get("brand", ""),
                    "model": prog.get("model", ""),
                    "trim": prog.get("trim", ""),
                    "year": prog.get("year", program_year),
                    "consumer_cash": prog.get("consumer_cash", 0) or 0,
                    "bonus_cash": prog.get("bonus_cash", 0) or 0,
                    "alt_consumer_cash": prog.get("alt_consumer_cash", 0) or 0,
                    "option1_rates": opt1,
                    "option2_rates": prog.get("option2_rates"),
                    "loyalty_cash": prog.get("loyalty_cash", False),
                    "loyalty_opt1": prog.get("loyalty_opt1", False),
                    "loyalty_opt2": prog.get("loyalty_opt2", False),
                    "program_month": program_month,
                    "program_year": program_year,
                    "created_at": datetime.utcnow().isoformat()
                }
                await db.programs.insert_one(program_doc)
                saved_count += 1
            logger.info(f"[Async] Saved {saved_count} programs")
        except Exception as save_error:
            logger.error(f"[Async] Save error: {str(save_error)}")

        # ── Step 3: Merge SCI + Excel + email + validation (le reste de ton code original) ──
        if sci_data_for_excel:
            vehicles_2026 = sci_data_for_excel.get("vehicles_2026", [])
            vehicles_2025 = sci_data_for_excel.get("vehicles_2025", [])
            _merge_previous_sci_rates(vehicles_2026, vehicles_2025, program_month, program_year)

        # Generate Excel and send email
        excel_sent = False
        if EXCEL_AVAILABLE and valid_programs and SMTP_EMAIL:
            try:
                excel_data = generate_excel_from_programs(valid_programs, program_month, program_year, sci_lease_data=sci_data_for_excel)
                excel_sent = send_excel_email(excel_data, SMTP_EMAIL, program_month, program_year, len(valid_programs))
            except Exception as excel_error:
                logger.error(f"[Async] Excel email error: {str(excel_error)}")

        # Post-extraction validation
        validation_msg = ""
        try:
            from services.pdfplumber_parser import validate_extraction
            validation = validate_extraction(valid_programs, sci_data_for_excel)
            if validation['warnings']:
                validation_msg = f" | {len(validation['warnings'])} avertissement(s)"
            if validation['errors']:
                validation_msg += f" | {len(validation['errors'])} erreur(s)"
        except Exception as val_error:
            logger.error(f"[Async] Validation error: {str(val_error)}")

        # Mark task complete
        lease_msg = f" + {len(sci_data_for_excel.get('vehicles_2026', []) + sci_data_for_excel.get('vehicles_2025', []))} taux SCI Lease" if sci_data_for_excel else ""
        email_msg = " - Excel envoyé par email!" if excel_sent else ""
        await db.extract_tasks.update_one(
            {"task_id": task_id},
            {"$set": {
                "status": "complete",
                "message": f"Extrait et sauvegardé {len(valid_programs)} programmes{lease_msg}{email_msg}{validation_msg}",
                "programs": valid_programs,
                "sci_lease_count": len(sci_data_for_excel.get('vehicles_2026', []) + sci_data_for_excel.get('vehicles_2025', [])) if sci_data_for_excel else 0,
                "completed_at": datetime.utcnow().isoformat()
            }}
        )

    except Exception as e:
        logger.error(f"[Async] Task {task_id} failed: {str(e)}")
        await db.extract_tasks.update_one(
            {"task_id": task_id},
            {"$set": {"status": "error", "message": f"Erreur: {str(e)}"}}
        )
        

@router.post("/extract-pdf-async")
async def extract_pdf_async(
    file: UploadFile = File(...),
    password: str = Form(...),
    program_month: int = Form(...),
    program_year: int = Form(...),
    start_page: Optional[int] = Form(None),
    end_page: Optional[int] = Form(None),
    lease_start_page: Optional[int] = Form(None),
    lease_end_page: Optional[int] = Form(None),
    selected_sections: Optional[str] = Form(None)
):
    """Upload PDF and start extraction in background. Uses selected sections from frontend checkboxes."""
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")

    pdf_content = await file.read()

    # Parse selected sections from frontend (JSON array)
    sections_list = []
    if selected_sections:
        try:
            sections_list = json.loads(selected_sections)
            logger.info(f"[AsyncExtract] Received {len(sections_list)} selected sections from frontend")
        except json.JSONDecodeError:
            logger.warning("[AsyncExtract] Failed to parse selected_sections JSON")

    # Extract page ranges from selected sections
    if sections_list:
        for s in sections_list:
            stype = s.get('type', '')
            if stype == 'retail':
                start_page = s.get('start_page')
                end_page = s.get('end_page')
            elif stype == 'lease':
                lease_start_page = s.get('start_page')
                lease_end_page = s.get('end_page')
        logger.info(f"[AsyncExtract] From selections: retail={start_page}-{end_page}, lease={lease_start_page}-{lease_end_page}")
    else:
        # Fallback: auto-detect from TOC
        from services.pdfplumber_parser import auto_detect_pages
        detected = auto_detect_pages(pdf_content)
        start_page = detected.get('retail_start') or start_page or 1
        end_page = detected.get('retail_end') or end_page or start_page
        lease_start_page = detected.get('lease_start') or lease_start_page
        lease_end_page = detected.get('lease_end') or lease_end_page
        logger.info(f"[AsyncExtract] Auto-detected fallback: retail={start_page}-{end_page}, lease={lease_start_page}-{lease_end_page}")

    task_id = str(uuid.uuid4())

    await db.extract_tasks.insert_one({
        "task_id": task_id,
        "status": "queued",
        "message": "En file d'attente...",
        "programs": [],
        "sci_lease_count": 0,
        "created_at": datetime.utcnow().isoformat()
    })

    asyncio.create_task(_run_extraction_task(
        task_id, pdf_content, password,
        program_month, program_year,
        start_page, end_page,
        lease_start_page, lease_end_page
    ))

    return {"task_id": task_id, "status": "queued", "message": "Extraction demarree en arriere-plan"}


@router.get("/extract-task/{task_id}")
async def get_extract_task(task_id: str):
    """Poll extraction task status."""
    task = await db.extract_tasks.find_one({"task_id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Tâche non trouvée")
    return task

# ============ Residual Guide PDF Upload ============

@router.post("/upload-residual-guide")
async def upload_residual_guide(
    file: UploadFile = File(...),
    password: str = Form(...),
    effective_month: int = Form(...),
    effective_year: int = Form(...)
):
    """
    Upload et parse automatiquement un PDF du guide des valeurs résiduelles SCI.
    Génère un Excel de vérification et l'envoie par email.
    """
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    
    import tempfile
    import os as os_module
    
    tmp_path = None
    try:
        # Save uploaded file temporarily
        content = await file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        
        # Parse PDF using our existing parser logic
        import fitz
        TERMS = ["24", "27", "36", "39", "42", "48", "51", "54", "60"]
        BODY_STYLES = [
            "4D Wagon AWD", "4D Wagon", "3D Hatchback",
            "2D Coupe AWD", "4D Sedan AWD",
            "4D Utility 4WD", "4D Utility AWD", "4D Utility",
            "2D Utility 4WD", "2D Utility AWD", "2D Utility",
            "Crew Cab LWB 2WD", "Crew Cab LWB 4WD",
            "Crew Cab SWB 2WD", "Crew Cab SWB 4WD",
            "Crew Cab 4WD",
            "Quad Cab SWB 4WD", "Quad Cab SWB 2WD",
            "Mega Cab 4WD", "Mega Cab 2WD",
            "Reg Cab LWB 2WD", "Reg Cab LWB 4WD",
            "Reg Cab 2WD", "Reg Cab 4WD",
        ]
        
        def is_number(s):
            try:
                int(s)
                return True
            except:
                return False
        
        def is_body_style(s):
            return s in BODY_STYLES
        
        doc = fitz.open(tmp_path)
        all_vehicles = []
        
        for page_num in range(doc.page_count):
            page = doc[page_num]
            text = page.get_text()
            lines_raw = [l.strip() for l in text.split('\n') if l.strip()]
            
            clean = []
            for l in lines_raw:
                if '675 Cochrane' in l or 'scileasecorp' in l or 'T: 1-888' in l or 'F: 1-866' in l:
                    continue
                if l in ('LEASE RESIDUAL VALUES', 'STANDARD', 'FEBRUARY 2026', 'RESIDUAL VALUE GUIDE') or l.startswith('Effective:'):
                    continue
                clean.append(l)
            
            current_brand = None
            current_model = None
            current_year = None
            
            i = 0
            while i < len(clean):
                line = clean[i]
                
                if line in ('CHRYSLER', 'DODGE', 'JEEP', 'RAM', 'FIAT'):
                    current_brand = line.title()
                    i += 1
                    if i < len(clean) and (clean[i] == '24' or clean[i].startswith('24 27')):
                        if clean[i] == '24':
                            skip = 0
                            while i + skip < len(clean) and clean[i + skip] in TERMS:
                                skip += 1
                            i += skip
                        else:
                            i += 1
                    continue
                
                if line.startswith('MODEL YEAR'):
                    i += 1
                    continue
                
                import re
                year_match = re.match(r'^(20\d{2})\s+(.+)$', line)
                if year_match:
                    current_year = int(year_match.group(1))
                    model_raw = year_match.group(2).strip()
                    
                    if i + 1 < len(clean):
                        next_line = clean[i + 1]
                        if (not is_body_style(next_line) and not is_number(next_line) and
                            not re.match(r'^(20\d{2})\s+', next_line) and
                            next_line not in ('CHRYSLER', 'DODGE', 'JEEP', 'RAM', 'FIAT') and
                            next_line not in TERMS and not next_line.startswith('24 27')):
                            if i + 2 < len(clean) and (is_body_style(clean[i + 2]) or is_number(clean[i + 2])):
                                current_model = model_raw
                            else:
                                current_model = model_raw + ' ' + next_line
                                i += 1
                        else:
                            current_model = model_raw
                    else:
                        current_model = model_raw
                    
                    i += 1
                    continue
                
                if current_brand and current_model and current_year:
                    trim = line
                    
                    if i + 1 < len(clean):
                        next_line = clean[i + 1]
                        
                        if is_body_style(next_line):
                            body_style = next_line
                            vals = []
                            j = i + 2
                            while j < len(clean) and len(vals) < 9:
                                if is_number(clean[j]):
                                    vals.append(int(clean[j]))
                                    j += 1
                                else:
                                    break
                            
                            if len(vals) == 9:
                                all_vehicles.append({
                                    "brand": current_brand,
                                    "model_year": current_year,
                                    "model_name": current_model,
                                    "trim": trim,
                                    "body_style": body_style,
                                    "residual_percentages": dict(zip(TERMS, vals))
                                })
                                i = j
                                continue
                        
                        if not is_body_style(next_line) and not is_number(next_line):
                            combined_trim = trim + ' ' + next_line
                            if i + 2 < len(clean) and is_body_style(clean[i + 2]):
                                body_style = clean[i + 2]
                                vals = []
                                j = i + 3
                                while j < len(clean) and len(vals) < 9:
                                    if is_number(clean[j]):
                                        vals.append(int(clean[j]))
                                        j += 1
                                    else:
                                        break
                                
                                if len(vals) == 9:
                                    all_vehicles.append({
                                        "brand": current_brand,
                                        "model_year": current_year,
                                        "model_name": current_model,
                                        "trim": combined_trim,
                                        "body_style": body_style,
                                        "residual_percentages": dict(zip(TERMS, vals))
                                    })
                                    i = j
                                    continue
                
                i += 1
        
        doc.close()
        
        if len(all_vehicles) == 0:
            raise HTTPException(status_code=400, detail="Aucun véhicule trouvé dans le PDF. Vérifiez le format du document.")
        
        # Build JSON result
        month_names = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                       "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        
        result = {
            "effective_from": f"{effective_year}-{effective_month:02d}-01",
            "source": f"SCI Lease Corp Stellantis Residual Guide - {month_names[effective_month]} {effective_year}",
            "km_adjustments": {
                "standard_km": 24000,
                "adjustments": {
                    "18000": {"24": 1, "27": 1, "36": 2, "39": 2, "42": 2, "48": 3, "51": 3, "54": 3, "60": 4},
                    "12000": {"24": 2, "27": 2, "36": 3, "39": 3, "42": 3, "48": 4, "51": 4, "54": 4, "60": 5}
                },
                "max_km_per_year": 36000
            },
            "vehicles": all_vehicles
        }
        
        # Save JSON file
        month_lower = month_names[effective_month].lower()
        json_filename = f"sci_residuals_{month_lower}{effective_year}.json"
        json_path = ROOT_DIR / "data" / json_filename
        with open(json_path, 'w') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        
        # Also update the current reference file
        current_path = ROOT_DIR / "data" / f"sci_residuals_{month_lower}{effective_year}.json"
        logger.info(f"Residual guide saved: {json_path} ({len(all_vehicles)} vehicles)")
        
        # Generate Excel for verification
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Valeurs résiduelles"
        
        headers = ["Marque", "Année", "Modèle", "Trim", "Carrosserie", "24m", "27m", "36m", "39m", "42m", "48m", "51m", "54m", "60m"]
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(color="4ECDC4", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for idx, v in enumerate(all_vehicles, 2):
            ws.cell(row=idx, column=1, value=v["brand"]).border = thin_border
            ws.cell(row=idx, column=2, value=v["model_year"]).border = thin_border
            ws.cell(row=idx, column=3, value=v["model_name"]).border = thin_border
            ws.cell(row=idx, column=4, value=v["trim"]).border = thin_border
            ws.cell(row=idx, column=5, value=v["body_style"]).border = thin_border
            for ti, term in enumerate(TERMS):
                val = v["residual_percentages"].get(term, 0)
                cell = ws.cell(row=idx, column=6 + ti, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        # Auto-width columns
        for col in range(1, len(headers) + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, len(all_vehicles) + 2))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 3
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_data = excel_buffer.getvalue()
        
        # Send verification email
        email_sent = False
        if SMTP_EMAIL and SMTP_PASSWORD:
            try:
                import smtplib
                from email.mime.multipart import MIMEMultipart
                from email.mime.text import MIMEText
                from email.mime.base import MIMEBase
                from email import encoders
                
                msg = MIMEMultipart()
                msg['From'] = SMTP_EMAIL
                msg['To'] = SMTP_EMAIL
                msg['Subject'] = f"CalcAuto - Guide Résiduel {month_names[effective_month]} {effective_year} ({len(all_vehicles)} véhicules)"
                
                # Count by brand
                brands = {}
                for v in all_vehicles:
                    b = v["brand"]
                    brands[b] = brands.get(b, 0) + 1
                brand_summary = "\n".join([f"  • {b}: {c} véhicules" for b, c in sorted(brands.items())])
                
                body = f"""Bonjour,

Le guide des valeurs résiduelles a été importé avec succès.

📊 Résumé:
• Période: {month_names[effective_month]} {effective_year}
• Total: {len(all_vehicles)} véhicules extraits

Par marque:
{brand_summary}

Le fichier Excel est joint pour vérification.

⚠️ IMPORTANT: Vérifiez les données avant utilisation dans les calculs de location.

---
CalcAuto AiPro"""
                
                msg.attach(MIMEText(body, 'plain', 'utf-8'))
                
                attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                attachment.set_payload(excel_data)
                encoders.encode_base64(attachment)
                att_filename = f"residuels_{month_lower}_{effective_year}.xlsx"
                attachment.add_header('Content-Disposition', f'attachment; filename={att_filename}')
                msg.attach(attachment)
                
                server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
                server.starttls()
                server.login(SMTP_EMAIL, SMTP_PASSWORD)
                server.send_message(msg)
                server.quit()
                email_sent = True
                logger.info(f"Residual guide email sent to {SMTP_EMAIL}")
            except Exception as e:
                logger.error(f"Error sending residual email: {str(e)}")
        
        # Count by brand for response
        brands_count = {}
        for v in all_vehicles:
            brands_count[v["brand"]] = brands_count.get(v["brand"], 0) + 1
        
        return {
            "success": True,
            "total_vehicles": len(all_vehicles),
            "brands": brands_count,
            "json_file": json_filename,
            "email_sent": email_sent,
            "message": f"{len(all_vehicles)} véhicules extraits et sauvegardés"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing residual guide: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")
    finally:
        if tmp_path and os_module.path.exists(tmp_path):
            os_module.unlink(tmp_path)

@router.post("/save-programs")
async def save_programs(request: SaveProgramsRequest):
    """
    Sauvegarde les programmes validés dans la base de données
    Remplace les programmes existants pour le mois/année spécifié
    Garde seulement les 6 derniers mois d'historique
    """
    if request.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe incorrect")
    
    # Delete existing programs for this period
    await db.programs.delete_many({
        "program_month": request.program_month,
        "program_year": request.program_year
    })
    
    # Insert new programs
    inserted = 0
    skipped = 0
    corrections_applied = []
    default_rates = {"rate_36": 4.99, "rate_48": 4.99, "rate_60": 4.99, "rate_72": 4.99, "rate_84": 4.99, "rate_96": 4.99}
    
    for prog_data in request.programs:
        # Skip invalid entries (missing brand/model)
        if not prog_data.get("brand") or not prog_data.get("model"):
            skipped += 1
            continue
            
        # Ensure option1_rates has a default value if missing
        if not prog_data.get("option1_rates"):
            prog_data["option1_rates"] = default_rates.copy()
        elif isinstance(prog_data["option1_rates"], dict):
            prog_data["option1_rates"] = FinancingRates(**prog_data["option1_rates"]).dict()
            
        # Process option2_rates if present
        if prog_data.get("option2_rates") and isinstance(prog_data["option2_rates"], dict):
            prog_data["option2_rates"] = FinancingRates(**prog_data["option2_rates"]).dict()
        
        prog_data["program_month"] = request.program_month
        prog_data["program_year"] = request.program_year
        prog_data["bonus_cash"] = prog_data.get("bonus_cash", 0)
        prog_data["consumer_cash"] = prog_data.get("consumer_cash", 0)

        # Appliquer les corrections memorisees (matching flexible)
        correction = await find_best_correction(
            prog_data.get("brand", ""),
            prog_data.get("model", ""),
            prog_data.get("trim", ""),
            prog_data.get("year", 2026)
        )
        correction_applied = False
        correction_details = {}
        if correction and correction.get("corrected_values"):
            cv = correction["corrected_values"]
            if cv.get("consumer_cash") is not None:
                old_val = prog_data.get("consumer_cash", 0)
                prog_data["consumer_cash"] = cv["consumer_cash"]
                if old_val != cv["consumer_cash"]:
                    correction_details["consumer_cash"] = {"avant": old_val, "apres": cv["consumer_cash"]}
            if cv.get("alternative_consumer_cash") is not None:
                old_val = prog_data.get("alternative_consumer_cash", 0)
                prog_data["alternative_consumer_cash"] = cv["alternative_consumer_cash"]
                if old_val != cv["alternative_consumer_cash"]:
                    correction_details["alternative_consumer_cash"] = {"avant": old_val, "apres": cv["alternative_consumer_cash"]}
            if cv.get("bonus_cash") is not None:
                old_val = prog_data.get("bonus_cash", 0)
                prog_data["bonus_cash"] = cv["bonus_cash"]
                if old_val != cv["bonus_cash"]:
                    correction_details["bonus_cash"] = {"avant": old_val, "apres": cv["bonus_cash"]}
            if cv.get("option1_rates"):
                prog_data["option1_rates"] = cv["option1_rates"]
                correction_details["option1_rates"] = "corrige"
            if cv.get("option2_rates") is not None:
                prog_data["option2_rates"] = cv["option2_rates"]
                correction_details["option2_rates"] = "corrige"
            correction_applied = True
            # Incrementer le compteur d'application
            await db.program_corrections.update_one(
                {"brand": correction["brand"], "model": correction["model"], "trim": correction["trim"], "year": correction["year"]},
                {"$inc": {"times_applied": 1}, "$set": {"last_applied_at": datetime.utcnow().isoformat()}}
            )
            logger.info(f"[CORRECTION] Appliquee pour {prog_data.get('brand')} {prog_data.get('model')} {prog_data.get('trim')}")
        
        if correction_applied:
            corrections_applied.append({
                "vehicule": f"{prog_data.get('brand')} {prog_data.get('model')} {prog_data.get('trim')} {prog_data.get('year')}",
                "changes": correction_details
            })
        
        try:
            prog = VehicleProgram(**prog_data)
            await db.programs.insert_one(prog.dict())
            inserted += 1
        except Exception as e:
            logger.warning(f"Skipped invalid program: {prog_data.get('brand')} {prog_data.get('model')} - {str(e)}")
            skipped += 1
            continue
    
    # Clean up old programs (keep only 6 months)
    await cleanup_old_programs()
    
    # Calculate brands summary for report
    brands_summary = {}
    for prog_data in request.programs:
        brand = prog_data.get("brand", "Inconnu")
        if brand not in brands_summary:
            brands_summary[brand] = 0
        brands_summary[brand] += 1
    
    # Send automatic email report
    try:
        await send_import_report_email(
            programs_count=inserted,
            program_month=request.program_month,
            program_year=request.program_year,
            brands_summary=brands_summary,
            skipped_count=skipped
        )
        logger.info(f"Import report email sent to {SMTP_EMAIL}")
    except Exception as e:
        logger.warning(f"Failed to send import report email: {str(e)}")
    
    # Force logout all users after data change
    await db.tokens.delete_many({})
    logger.info("[PDF IMPORT] Tokens invalides pour forcer reconnexion")

    return {
        "success": True,
        "message": f"Sauvegardé {inserted} programmes pour {request.program_month}/{request.program_year}" + (f" ({skipped} ignorés)" if skipped > 0 else "") + (f" — {len(corrections_applied)} corrections appliquées" if corrections_applied else "") + " - Tous les utilisateurs deconnectes",
        "inserted": inserted,
        "skipped": skipped,
        "corrections_applied": len(corrections_applied),
        "correction_details": corrections_applied[:50]
    }


@router.get("/corrections")
async def list_corrections():
    """Retourne toutes les corrections memorisees avec statistiques."""
    corrections = await db.program_corrections.find({}, {"_id": 0}).sort("corrected_at", -1).to_list(500)
    return {
        "total": len(corrections),
        "corrections": corrections
    }


@router.delete("/corrections/{brand}/{model}/{year}")
async def delete_correction(brand: str, model: str, year: int, password: str = ""):
    """Supprime une correction memorisee."""
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe admin incorrect")
    result = await db.program_corrections.delete_many({"brand": brand, "model": model, "year": year})
    return {"deleted": result.deleted_count}


@router.delete("/corrections/all")
async def delete_all_corrections(password: str = ""):
    """Supprime toutes les corrections memorisees."""
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Mot de passe admin incorrect")
    result = await db.program_corrections.delete_many({})
    return {"deleted": result.deleted_count}


async def send_import_report_email(programs_count: int, program_month: int, program_year: int, brands_summary: dict, skipped_count: int = 0):
    """Envoie automatiquement un rapport par email après l'import des programmes"""
    months_fr = {
        1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
        5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
        9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"
    }
    month_name = months_fr.get(program_month, str(program_month))
    
    # Generate brands table
    brands_rows = ""
    total_programs = 0
    for brand in ['Chrysler', 'Jeep', 'Dodge', 'Ram', 'Fiat']:
        count = brands_summary.get(brand, 0)
        if count > 0:
            brands_rows += f"<tr><td style='padding: 10px; border-bottom: 1px solid #eee;'>{brand}</td><td style='padding: 10px; border-bottom: 1px solid #eee; text-align: center; font-weight: bold;'>{count}</td></tr>"
            total_programs += count
    
    # Add any other brands not in the standard list
    for brand, count in brands_summary.items():
        if brand not in ['Chrysler', 'Jeep', 'Dodge', 'Ram', 'Fiat'] and count > 0:
            brands_rows += f"<tr><td style='padding: 10px; border-bottom: 1px solid #eee;'>{brand}</td><td style='padding: 10px; border-bottom: 1px solid #eee; text-align: center; font-weight: bold;'>{count}</td></tr>"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; background: #fff; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); overflow: hidden; }}
            .header {{ background: linear-gradient(135deg, #1a5f4a 0%, #2d8f6f 100%); color: #fff; padding: 25px; text-align: center; }}
            .header h1 {{ margin: 0; font-size: 24px; }}
            .header .subtitle {{ margin-top: 8px; opacity: 0.9; }}
            .content {{ padding: 25px; }}
            .success-badge {{ background: #d4edda; color: #155724; padding: 15px; border-radius: 8px; text-align: center; margin-bottom: 20px; font-size: 16px; }}
            .success-badge strong {{ font-size: 18px; }}
            .stats-box {{ background: #f8f9fa; border-radius: 8px; padding: 20px; margin-bottom: 20px; text-align: center; }}
            .big-number {{ font-size: 56px; font-weight: bold; color: #1a5f4a; }}
            .big-label {{ color: #666; font-size: 14px; margin-top: 5px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th {{ background: #1a5f4a; color: #fff; padding: 12px; text-align: left; }}
            .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #eee; }}
            .warning-box {{ background: #fff3cd; border: 1px solid #ffc107; border-radius: 6px; padding: 12px; margin-top: 15px; color: #856404; font-size: 13px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>✅ Import Réussi!</h1>
                <div class="subtitle">CalcAuto AiPro - Rapport d'import automatique</div>
            </div>
            <div class="content">
                <div class="success-badge">
                    <strong>🎉 Programmes {month_name} {program_year}</strong><br>
                    importés avec succès!
                </div>
                
                <div class="stats-box">
                    <div class="big-number">{programs_count}</div>
                    <div class="big-label">programmes de financement</div>
                </div>
                
                <h3 style="color: #1a5f4a; border-bottom: 2px solid #1a5f4a; padding-bottom: 8px;">📊 Répartition par marque</h3>
                <table>
                    <thead>
                        <tr>
                            <th>Marque</th>
                            <th style="text-align: center;">Nombre</th>
                        </tr>
                    </thead>
                    <tbody>
                        {brands_rows}
                        <tr style="background: #f8f9fa; font-weight: bold;">
                            <td style="padding: 12px;">TOTAL</td>
                            <td style="padding: 12px; text-align: center;">{programs_count}</td>
                        </tr>
                    </tbody>
                </table>
                
                {f'<div class="warning-box">⚠️ {skipped_count} programme(s) ignoré(s) en raison de données invalides.</div>' if skipped_count > 0 else ''}
                
                <p style="margin-top: 25px; color: #666; font-size: 14px;">
                    Les nouveaux programmes sont maintenant disponibles dans l'application CalcAuto AiPro. 
                    Vos clients peuvent commencer à utiliser les nouveaux taux immédiatement.
                </p>
            </div>
            <div class="footer">
                <p style="margin: 0;"><strong>CalcAuto AiPro</strong></p>
                <p style="margin: 8px 0 0;">Rapport généré automatiquement le {datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    subject = f"✅ Import {month_name} {program_year} - {programs_count} programmes"
    
    send_email(SMTP_EMAIL, subject, html_body)

async def cleanup_old_programs():
    """Supprime les programmes de plus de 6 mois"""
    # Get all unique periods
    pipeline = [
        {"$group": {
            "_id": {"month": "$program_month", "year": "$program_year"}
        }},
        {"$sort": {"_id.year": -1, "_id.month": -1}}
    ]
    periods = await db.programs.aggregate(pipeline).to_list(100)
    
    # Keep only the 6 most recent periods
    if len(periods) > 6:
        periods_to_delete = periods[6:]
        for p in periods_to_delete:
            await db.programs.delete_many({
                "program_month": p["_id"]["month"],
                "program_year": p["_id"]["year"]
            })
            logger.info(f"Deleted old programs for {p['_id']['month']}/{p['_id']['year']}")

